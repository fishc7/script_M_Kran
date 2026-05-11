import sys
import locale
import os

# Устанавливаем кодировку UTF-8 для Windows
if sys.platform.startswith('win'):
    # Устанавливаем локаль для корректной работы с русскими символами
    try:
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_ALL, 'Russian_Russia.1251')
        except:
            pass
    
    # Устанавливаем переменные окружения для кодировки
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    os.environ['PYTHONLEGACYWINDOWSSTDIO'] = '1'
    
    # Настраиваем стандартные потоки для UTF-8
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())
    sys.stdin = codecs.getreader('utf-8')(sys.stdin.detach())

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, send_from_directory, got_request_exception
import sqlite3
import pandas as pd
import os
import json
import subprocess
import threading
import time
from datetime import datetime, timedelta
import logging
from pathlib import Path
from werkzeug.utils import secure_filename
import zipfile
import io
from io import BytesIO
import re
import unicodedata
# Добавляем путь к текущей директории для импорта script_runner
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from script_runner import init_script_runner, get_script_runner

# Настраиваем пути для Flask
template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'templates')
static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static')

app = Flask(__name__, 
           template_folder=template_dir,
           static_folder=static_dir)
app.secret_key = 'm_kran_secret_key_2025'

# Подключаем NAKS blueprint, чтобы пункт "Парсинг НАКС" отображался в меню
try:
    from modules.routes.naks_routes import naks_bp
    app.register_blueprint(naks_bp)
    logging.info("NAKS blueprint зарегистрирован")
except Exception as e:
    logging.warning(f"NAKS blueprint не зарегистрирован: {e}")


@app.context_processor
def inject_template_flags():
    """Глобальные флаги для безопасного рендеринга шаблонов."""
    return {
        'naks_parser_available': 'naks.naks_parser_page' in app.view_functions,
        'statistics_available': 'statistics' in app.view_functions
    }

# Настраиваем кодировку для Flask
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = True
app.config['TEMPLATES_AUTO_RELOAD'] = True
try:
    # Очищаем кэш шаблонов на старте
    app.jinja_env.cache = {}
except Exception:
    pass

try:
    logging.info(f'TEMPLATE_DIR = {template_dir}')
    logging.info(f'STATIC_DIR   = {static_dir}')
except Exception:
    pass

# Диагностика путей и шаблонов
@app.route('/api/debug_template_info')
def api_debug_template_info():
    try:
        info = {
            'template_dir': template_dir,
            'static_dir': static_dir,
        }
        jm_path = os.path.join(template_dir, 'joint_manager.html')
        info['joint_manager_path'] = jm_path
        info['joint_manager_exists'] = os.path.exists(jm_path)
        if os.path.exists(jm_path):
            try:
                stat = os.stat(jm_path)
                info['joint_manager_mtime'] = stat.st_mtime
                info['joint_manager_size'] = stat.st_size
                with open(jm_path, 'r', encoding='utf-8', errors='ignore') as f:
                    head = f.read(200)
                info['joint_manager_head'] = head
            except Exception as e:
                info['joint_manager_read_error'] = str(e)
        return jsonify({'success': True, 'info': info})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/clear_template_cache', methods=['POST'])
def api_clear_template_cache():
    try:
        app.jinja_env.cache = {}
        return jsonify({'success': True, 'message': 'Template cache cleared'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# CSP заголовки и кэш-контроль
import secrets

@app.after_request
def add_csp_headers(response):
    # Генерируем уникальный nonce для каждого запроса
    nonce = secrets.token_urlsafe(16)
    response.set_cookie('nonce', nonce)
    
    # Устанавливаем разрешительный CSP для тестирования
    response.headers['Content-Security-Policy'] = "default-src 'self' *; script-src 'self' 'unsafe-inline' 'unsafe-eval' *; style-src 'self' 'unsafe-inline' *; img-src 'self' data: *; font-src 'self' *;"
    if 'X-Content-Type-Options' in response.headers:
        del response.headers['X-Content-Type-Options']
    
    # Кэш-контроль для предотвращения кэширования
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    # CORS заголовки
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    
    # Устанавливаем правильную кодировку UTF-8
    if response.content_type and 'text/html' in response.content_type:
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    
    return response

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Отдельный файл для ошибок веб-приложения (traceback 500 и других необработанных исключений)
try:
    logs_dir = Path(__file__).resolve().parents[2] / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    web_errors_log = logs_dir / "web_app_errors.log"
    web_errors_log.touch(exist_ok=True)

    file_handler = logging.FileHandler(web_errors_log, encoding="utf-8")
    file_handler.setLevel(logging.ERROR)
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    if not any(
        isinstance(h, logging.FileHandler) and getattr(h, "baseFilename", "") == str(web_errors_log)
        for h in logger.handlers
    ):
        logger.addHandler(file_handler)
except Exception:
    pass


@got_request_exception.connect_via(app)
def log_unhandled_exception(sender, exception, **extra):
    """Логирует необработанные исключения Flask с traceback в файл логов."""
    try:
        logger.exception(
            "Необработанное исключение в Flask: %s | path=%s | method=%s | args=%s",
            exception,
            request.path,
            request.method,
            dict(request.args),
        )
    except Exception:
        logger.exception("Необработанное исключение в Flask: %s", exception)


@app.errorhandler(Exception)
def handle_all_exceptions(e):
    """Последний рубеж: логируем любую необработанную ошибку запроса."""
    try:
        logger.exception(
            "Flask errorhandler captured exception: %s | path=%s | method=%s | args=%s",
            e,
            request.path,
            request.method,
            dict(request.args),
        )
    except Exception:
        logger.exception("Flask errorhandler captured exception: %s", e)
    return "Internal Server Error", 500

# Добавляем кастомные фильтры для Jinja2
@app.template_filter('format_number')
def format_number(value):
    """Форматирует число с разделителями тысяч"""
    if value is None:
        return '0'
    try:
        return f"{int(value):,}".replace(',', ' ')
    except (ValueError, TypeError):
        return str(value)

# Добавляем глобальные функции в Jinja2
app.jinja_env.globals.update(max=max, min=min, range=range)

def extract_titul_from_iso_string(iso_string):
    """
    Извлекает данные титула из строки ISO
    Пример: GCC-NAG-DDD-12470-13-1400-TK-ISO-00008 -> 12470-13
    """
    if not iso_string:
        return None
    
    # Паттерн для извлечения номера титула
    # Ищем паттерн: число-число (например, 12470-13)
    pattern = r'(\d{5}-\d{2})'
    match = re.search(pattern, iso_string)
    
    if match:
        return match.group(1)
    
    # Альтернативный паттерн для других форматов
    pattern2 = r'(\d{4,5}-\d{1,2})'
    match2 = re.search(pattern2, iso_string)
    
    if match2:
        return match2.group(1)
    
    return None

def create_id_folders(id_number):
    """
    Создает папки для ИД в указанной директории
    """
    try:
        # Основная папка для ИД
        base_path = r"D:\МК_Кран\МК_Кран_Кингесеп\ПТО\ИД"
        
        # Проверяем существование основной папки
        if not os.path.exists(base_path):
            logger.warning(f"Основная папка не существует: {base_path}")
            return False, f"Основная папка не существует: {base_path}"
        
        # Очищаем номер ИД от недопустимых символов для имени папки
        safe_id = re.sub(r'[<>:"/\\|?*]', '_', str(id_number))
        safe_id = safe_id.strip()
        
        if not safe_id:
            return False, "Номер ИД пустой или содержит только недопустимые символы"
        
        # Путь к папке ИД
        id_folder_path = os.path.join(base_path, safe_id)
        
        # Проверяем существование основной папки ИД
        if os.path.exists(id_folder_path):
            logger.info(f"Папка ИД уже существует: {id_folder_path}")
            return True, f"Папка для ИД {safe_id} уже существует"
        
        # Создаем основную папку ИД
        os.makedirs(id_folder_path)
        logger.info(f"Создана основная папка ИД: {id_folder_path}")
        
        # Подпапки для ИД
        subfolders = [
            "Заключение ВИК",
            "Заключение РК", 
            "Сварочные материалы",
            "Сварщики документы"
        ]
        
        created_folders = []
        for subfolder in subfolders:
            subfolder_path = os.path.join(id_folder_path, subfolder)
            if not os.path.exists(subfolder_path):
                os.makedirs(subfolder_path)
                created_folders.append(subfolder)
                logger.info(f"Создана подпапка: {subfolder_path}")
            else:
                logger.info(f"Подпапка уже существует: {subfolder_path}")
        
        if created_folders:
            logger.info(f"Созданы папки для ИД {safe_id}: {', '.join(created_folders)}")
        
        return True, f"Папки для ИД {safe_id} созданы успешно"
        
    except Exception as e:
        error_msg = f"Ошибка создания папок для ИД {id_number}: {str(e)}"
        logger.error(error_msg)
        return False, error_msg

def safe_encode_value(value):
    """Безопасно кодирует значение для отображения в веб-интерфейсе"""
    if value is None:
        return ''
    
    try:
        # Преобразуем в строку
        str_value = str(value)
        
        # Заменяем проблемные Unicode символы на их HTML-эквиваленты
        replacements = {
            '\u221a': '√',  # символ квадратного корня
            '\u2713': '✓',  # галочка
            '\u2714': '✔',  # жирная галочка
            '\u2717': '✗',  # крестик
            '\u2718': '✘',  # жирный крестик
            '\u00b0': '°',  # градус
            '\u00b1': '±',  # плюс-минус
            '\u00b2': '²',  # квадрат
            '\u00b3': '³',  # куб
            '\u00bc': '¼',  # четверть
            '\u00bd': '½',  # половина
            '\u00be': '¾',  # три четверти
            '\u0394': 'Δ',  # греческая дельта
            '\u03b4': 'δ',  # греческая дельта (строчная)
        }
        
        # Применяем замены
        for unicode_char, replacement in replacements.items():
            str_value = str_value.replace(unicode_char, replacement)
        
        # Дополнительно обрабатываем другие проблемные символы
        # Заменяем непечатаемые символы на пробелы
        str_value = ''.join(char if ord(char) >= 32 and ord(char) != 127 else ' ' for char in str_value)
        
        return str_value
        
    except Exception as e:
        print(f"[WARNING] Ошибка кодирования значения '{value}': {e}")
        # В случае ошибки возвращаем безопасную версию
        try:
            # Пытаемся сначала с UTF-8, затем с ASCII
            if isinstance(value, str):
                return value.encode('utf-8', 'replace').decode('utf-8')
            else:
                return str(value).encode('utf-8', 'replace').decode('utf-8')
        except:
            try:
                return str(value).encode('ascii', 'replace').decode('ascii')
            except:
                return '[ОШИБКА КОДИРОВКИ]'


def get_filename_timestamp():
    """
    Возвращает временную метку для имени файла в формате день.месяц.год_часы-минуты-секунды
    (двоеточия заменяются на дефисы для совместимости с Windows)
    """
    return datetime.now().strftime('%d.%m.%Y_%H:%M:%S').replace(':', '-')

def clean_data_for_excel(df):
    """Очищает данные для безопасного экспорта в Excel"""
    cleaned_df = df.copy()
    
    # Обрабатываем NaN значения - заменяем на пустые строки
    cleaned_df = cleaned_df.fillna('')
    
    for col in cleaned_df.columns:
        cleaned_df[col] = cleaned_df[col].apply(lambda x: 
            '' if x is None or x == '' else 
            str(x).replace('\x00', '').replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
            if isinstance(x, str) else str(x)
        )
    
    return cleaned_df

# Добавляем функцию в Jinja2
app.jinja_env.globals.update(safe_encode_value=safe_encode_value)

# Конфигурация
class Config:
    PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    DB_PATH = os.path.join(PROJECT_ROOT, 'database', 'BD_Kingisepp', 'M_Kran_Kingesepp.db')
    SCRIPTS_DIR = os.path.join(PROJECT_ROOT, 'scripts')
    UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
    MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB max file size

app.config.from_object(Config)

# Создаем папку для загрузок
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Инициализируем ScriptRunner
init_script_runner(Config.PROJECT_ROOT)

def safe_filename(filename):
    """
    Безопасное сохранение имени файла с поддержкой русских символов
    """
    if not filename:
        return 'unnamed_file'
    
    # Разделяем имя файла и расширение
    name, ext = os.path.splitext(filename)
    
    # Нормализуем Unicode символы
    name = unicodedata.normalize('NFKD', name)
    
    # Удаляем опасные символы, но сохраняем русские буквы, цифры, пробелы, дефисы и подчеркивания
    # Разрешенные символы: буквы (любые), цифры, пробелы, дефисы, подчеркивания, скобки
    safe_name = re.sub(r'[^\w\s\-_()№.]', '', name, flags=re.UNICODE)
    
    # Заменяем множественные пробелы на одиночные
    safe_name = re.sub(r'\s+', ' ', safe_name).strip()
    
    # Ограничиваем длину имени файла (максимум 100 символов)
    if len(safe_name) > 100:
        safe_name = safe_name[:100].strip()
    
    # Если имя файла стало пустым, используем дефолтное
    if not safe_name:
        safe_name = 'uploaded_file'
    
    # Возвращаем безопасное имя с расширением
    return safe_name + ext

def get_unique_filename(directory, filename):
    """
    Получить уникальное имя файла, добавив суффикс если файл уже существует
    """
    if not os.path.exists(os.path.join(directory, filename)):
        return filename
    
    name, ext = os.path.splitext(filename)
    counter = 1
    
    while True:
        new_filename = f"{name} ({counter}){ext}"
        if not os.path.exists(os.path.join(directory, new_filename)):
            return new_filename
        counter += 1

# ETL процессы с детальной информацией о файлах и таблицах
ETL_CATEGORIES = {
    'extract': {
        'name': '📥 Extract - Извлечение данных',
        'description': 'Загрузка исходных данных из различных источников в систему',
        'type': 'extract',
        'scripts': {
            'load_lnk_data.py': {
                'description': '📖 Журнал НК НГС - результаты контроля качества',
                'source_files': ['Журнал ЛНК_*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/НК/Журнал',
                'target_table': 'logs_lnk',
                'data_type': 'Результаты контроля качества (ВИК, РК), статусы дефектов'
            },
            'load_lnk_nk_aks.py': {
                'description': 'Журнал НК АКС - результаты контроля качества',
                'source_files': ['LOG_М-КРАН_RT_ТТ*.xlsx', '*.xlsb', '*.csv'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/НК_АКС',
                'target_table': 'logs_lnk',
                'data_type': 'Журнал М-Кран RT/ТТ — вставка строк (после журнала НГС)'
            },
            'run_full_logs_lnk_update.py': {
                'description': 'Полное обновление журнала НК',
                'source_files': ['Журнал ЛНК_*.xlsx', 'LOG_М-КРАН_RT_ТТ* (.xlsx/.xls/.xlsb/.csv)'],
                'source_folder': 'НК/Журнал и НК_АКС',
                'target_table': 'logs_lnk',
                'data_type': 'Цепочка: load_lnk_data.py → load_lnk_nk_aks.py',
                'exclude_from_bulk_daily': True,
            },
            'load_staff_titles_M_Kran.py': {
                'description': '👥 Расстановка персонала М_Кран по участкам',
                'source_files': ['Сварка *.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/СМР/отчет_площадка/растановка 12460',
                'target_table': 'Staff_Titles_M_Kran',
                'data_type': 'Расстановка персонала М_Кран по участкам'
            },
            'load_ndt_findings_transmission_register.py': {
                'description': '📋 Реестр заключений НГС Эксперт',
                'source_files': ['Реестр передачи заключений*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/НК/Реестр_передачи_заключений',
                'target_table': 'NDT_Findings_Transmission_Register',
                'data_type': 'Реестр переданных заключений по неразрушающему контролю от НГС Эксперт'
            },
            'load_wl_report_smr_web.py': {
                'description': '📋 Отчеты мастеров СМР (площадка)',
                'source_files': ['Сварка *.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/СМР/отчет_площадка/сварка',
                'target_table': 'wl_report_smr',
                'data_type': 'Отчеты мастеров по сварочно-монтажным работам'
            },
            'load_work_order_log_NDT.py': {
                'description': '📝 Заявки на НК от М_Кран - планирование работ',
                'source_files': ['Заявки_НК_*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/НК/Заявки_НК/Заявки_excel',
                'target_table': 'Work_Order_Log_NDT',
                'data_type': 'Заявки на неразрушающий контроль, планирование работ'
            },
            'load_wl_china.py': {
                'description': '🇨🇳 Данные китайских подрядчиков WELDLOG',
                'source_files': ['Журнал сварочных работ.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ОГС/Журналы',
                'target_table': 'wl_china',
                'data_type': 'Данные сварки от китайских подрядчиков, номера швов'
            },
            'load_Piping_Log_PTO.py': {
                'description': '📋 Извлечение перечня ISO от ПТО',
                'source_files': ['Свод_ISO-Линия.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ПТО',
                'target_table': 'Log_Piping_PTO',
                'data_type': 'Проектная документация, ISO чертежи, линии трубопроводов'
            },
            'load_lst_data.py': {
                'description': '📋 Извлечение реестра ТТ',
                'source_files': ['Реестор_ТТ_категория_контроль.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ОГС',
                'target_table': 'lst_data',
                'data_type': 'Реестр технологического трубопровода, схемы и спецификации'
            },
            'load_dl_data.py': {
                'description': '📊 Извлечение реестра ДЛ',
                'source_files': ['Реестор ДЛ.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ОГС/ДЛ',
                'target_table': 'dl_data',
                'data_type': 'Реестр допускных листов на сварщиков, квалификационные данные'
            },
            'load_tks_data.py': {
                'description': '📊 Извлечение данных ТКС',
                'source_files': ['Реестор_ТКС.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ОГС/ТКС',
                'target_table': 'tks_data',
                'data_type': 'Технологические карты сварки, процедуры и параметры'
            },
            'load_Pipeline_Test_Package.py': {
                'description': '📦 Извлечение тест-пакетов ТТ',
                'source_files': ['Тест-пакет трубопроводов*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ОГС/Испытания',
                'target_table': 'Pipeline_Test_Package',
                'data_type': 'Испытания трубопроводов, гидравлические тесты'
            },
            'create_ndt_reports_table.py': {
                'description': '📁 Загрузить перечень заключений НК',
                'source_files': ['*.pdf', '*.doc', '*.docx', '*.xls', '*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/НК/Заключения',
                'target_table': 'folder_NDT_Report',
                'data_type': 'Файловый реестр заключений неразрушающего контроля'
            },
            'load_pipeline_weld_joint_iso.py': {
                'description': '🔗 Сварные соединения ISO - номерация стыков',
                'source_files': ['номерация_стыков_по_iso_*.xlsx'],
                'source_folder': 'D:/МК_Кран/МК_Кран_Кингесеп/ПТО/номерация стыков по iso',
                'target_table': 'pipeline_weld_joint_iso',
                'data_type': 'Номерация стыков трубопровода по ISO стандартам'
            },
        }
    },
    'transform': {
        'name': '🔄 Transform - Преобразование данных', 
        'description': 'Очистка, нормализация и преобразование загруженных данных',
        'type': 'transform',
        'scripts': {
            'clean_journal_lnk_data.py': {
                'description': '🧹 Очистка и нормализация журнала ЛНК',
                'source_table': 'logs_lnk',
                'target_table': 'logs_lnk (обновленная)',
                'operations': 'Удаление дубликатов, нормализация статусов, очистка пробелов'
            },
            'sync_pipeline_wl_china.py': {
                'description': '🔄 Синхронизация pipeline_weld_joint_iso и wl_china',
                'source_table': 'wl_china',
                'target_table': 'pipeline_weld_joint_iso',
                'operations': 'Проверка соответствия записей по ISO и стык, вставка недостающих данных'
            },
            'merge_duplicates_lnk_logs.py': {
                'description': '🔗 Объединение дубликатов в ЛНК',
                'source_table': 'logs_lnk',
                'target_table': 'logs_lnk_merged',
                'operations': 'Поиск и объединение дублирующихся записей по ключевым полям'
            },
            'create_pipeline_weld_joint.py': {
                'description': '🔗 Создание связей стыков и трубопроводов',
                'source_table': 'Log_Piping_PTO',
                'target_table': 'pipeline_weld_joint',
                'operations': 'Генерация реестра стыков из объемов трубопроводов'
            },
            'clean_weld_repair_log.py': {
                'description': '🔧 Очистка журнала ремонта сварных швов',
                'source_table': 'weld_repair_log',
                'target_table': 'weld_repair_log (очищенная)',
                'operations': 'Очистка и нормализация данных о ремонте швов'
            },
            'sync_weld_repair_log.py': {
                'description': '🔄 Синхронизация журнала ремонта швов',
                'source_table': 'weld_repair_log',
                'target_table': 'weld_repair_log (синхронизированная)',
                'operations': 'Синхронизация данных о ремонте сварных швов'
            },
            'update_wl_china_from_osnovnaya_nk.py': {
                'description': '🔄 Обновление wl_china из основнаяНК',
                'source_table': 'основнаяНК',
                'target_table': 'wl_china',
                'operations': 'Перенос значения "РК (Радиографический контроль) / RT" в поля Проектный_контроля и проектный_объем_РК по ключам титул + номер линии'
            },
            'create_condition_weld_table.py': {
                'description': '🏗️ Создание последнее состояния сварных швов',
                'source_table': 'pipeline_weld_joint_iso, logs_lnk, wl_china',
                'target_table': 'condition_weld',
                'operations': 'Объединение данных RT, VT и заключений по сварным швам'
            },
            'create_svarenno_svarshchikom_table.py': {
                'description': '👷 Создание таблицы сварено сварщиком',
                'source_table': 'wl_china',
                'target_table': 'сварено_сварщиком',
                'operations': 'Группировка по линиям и чертежам, объединение клейм сварщиков, методов сварки и типов швов'
            },
        }
    },
    'load': {
        'name': '📤 Load - Загрузка в хранилище',
        'description': 'Финальная загрузка обработанных данных в целевые таблицы',
        'type': 'load',
        'scripts': {
            'load_pipeline_weld_joint_iso.py': {
                'description': '🔗 Загрузка связей стыков с ISO в БД',
                'source_table': 'pipeline_weld_joint (временная)',
                'target_table': 'pipeline_weld_joint_iso (финальная)',
                'operations': 'Финальная загрузка с индексами и проверкой целостности'
            },
        }
    },
    'additional_tools': {
        'name': '🛠️ Дополнительные инструменты',
        'description': 'Вспомогательные инструменты для анализа и экспорта данных',
        'type': 'tools',
        'scripts': {
            'export_condition_weld_to_excel.py': {
                'description': '📊 Экспорт состояния сварных швов в Excel',
                'source_table': 'condition_weld',
                'target_file': 'condition_weld_report_YYYYMMDD_HHMMSS.xlsx',
                'operations': 'Создание Excel отчета с листами RT, VT, WL_China и статистикой'
            },
            'view_condition_weld.py': {
                'description': '👁️ Просмотр данных состояния сварных швов',
                'source_table': 'condition_weld',
                'target_output': 'Консольный вывод',
                'operations': 'Интерактивный просмотр данных с фильтрами и статистикой'
            }
        }
    }
}

def get_db_connection():
    """Создает соединение с базой данных"""
    logger.info(f"[DB] Попытка подключения к БД: {app.config['DB_PATH']}")
    try:
        # Добавляем timeout и другие параметры для лучшей обработки блокировок
        conn = sqlite3.connect(
            app.config['DB_PATH'],
            timeout=30.0,  # 30 секунд timeout
            check_same_thread=False
        )
        conn.row_factory = sqlite3.Row
        
        # Включаем WAL режим для лучшей производительности и меньших блокировок
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size=10000")
        conn.execute("PRAGMA temp_store=MEMORY")
        
        # Устанавливаем кодировку UTF-8
        conn.execute("PRAGMA encoding='UTF-8'")
        
        logger.info("[DB] Подключение к БД успешно")
        return conn
    except sqlite3.OperationalError as e:
        logger.error(f"Ошибка подключения к БД (OperationalError): {e}")
        if "database is locked" in str(e).lower():
            logger.error("База данных заблокирована другим процессом")
        return None
    except Exception as e:
        logger.error(f"Ошибка подключения к БД: {e}")
        return None

def get_db_stats():
    """Получает статистику базы данных"""
    conn = get_db_connection()
    if not conn:
        return {
            'total_records': 0,
            'table_count': 0,
            'db_size': '0 МБ'
        }
    
    try:
        cursor = conn.cursor()
        
        # Получаем список всех таблиц
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row['name'] for row in cursor.fetchall()]
        
        total_records = 0
        # Считаем общее количество записей во всех таблицах
        for table in tables:
            try:
                cursor.execute(f"SELECT COUNT(*) as count FROM [{table}]")
                count = cursor.fetchone()['count']
                total_records += count
            except Exception as e:
                logger.warning(f"Ошибка при подсчете записей в таблице {table}: {e}")
                continue
        
        # Получаем размер базы данных
        db_size = 0
        if os.path.exists(app.config['DB_PATH']):
            db_size = os.path.getsize(app.config['DB_PATH'])
            if db_size < 1024 * 1024:
                db_size_str = f"{db_size // 1024} КБ"
            else:
                db_size_str = f"{db_size // (1024 * 1024)} МБ"
        else:
            db_size_str = "0 МБ"
        
        conn.close()
        
        return {
            'total_records': total_records,
            'table_count': len(tables),
            'db_size': db_size_str
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики БД: {e}")
        conn.close()
        return {
            'total_records': 0,
            'table_count': 0,
            'db_size': '0 МБ'
        }

def get_logs_lnk_stats():
    """Получает статистику из таблицы logs_lnk"""
    conn = get_db_connection()
    if not conn:
        return {
            'total_records': 0,
            'vik_good': 0,
            'rk_good': 0,
            'rk_defects': 0,
            'rk_np': 0,
            'rk_pending': 0,
            'vik_pending': 0,
            'rt_requested': 0,
            'vt_requested': 0,
            'pmi_requested': 0,
            'unique_drawings': 0,
            'unique_joints': 0,
            'unique_zones': 0
        }
    
    try:
        cursor = conn.cursor()
        
        # Всего записей
        cursor.execute('SELECT COUNT(*) FROM logs_lnk')
        total_records = cursor.fetchone()[0]
        
        # ВИК Годен
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_ВИК" = "Годен"')
        vik_good = cursor.fetchone()[0]
        
        # РК Годен
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Годен"')
        rk_good = cursor.fetchone()[0]
        
        # РК дефекты: Не годен (статусы требующие исправления)
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" LIKE "%Не годен%"')
        rk_defects = cursor.fetchone()[0]
        
        # РК Н/П (неофициальный ремонт или вырез)
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Н/П"')
        rk_np = cursor.fetchone()[0]
        
        # Всего негодных (сумма официальных и неофициальных ремонтов)
        total_defects = rk_defects + rk_np
        
        # РК Заявлен (отставание в контроле) - по столбцу РК
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE LOWER("РК") LIKE "%явлен%"')
        rk_pending = cursor.fetchone()[0]
        
        # ВИК Заявлен (отставание в контроле) - по столбцу ВИК
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE LOWER("ВИК") LIKE "%явлен%"')
        vik_pending = cursor.fetchone()[0]
        
        # Заявки на РК (содержащие RT)
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%RT%"')
        rt_requested = cursor.fetchone()[0]
        
        # Заявки на ВИК (содержащие VT)
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%VT%"')
        vt_requested = cursor.fetchone()[0]
        
        # Заявки на PMI
        cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%PMI%"')
        pmi_requested = cursor.fetchone()[0]
        
        # Уникальные чертежи
        cursor.execute('SELECT COUNT(DISTINCT "Чертеж") FROM logs_lnk WHERE "Чертеж" IS NOT NULL AND "Чертеж" != ""')
        unique_drawings = cursor.fetchone()[0]
        
        # Уникальные стыки
        cursor.execute('SELECT COUNT(DISTINCT "Номер_стыка") FROM logs_lnk WHERE "Номер_стыка" IS NOT NULL AND "Номер_стыка" != ""')
        unique_joints = cursor.fetchone()[0]
        
        # Уникальные зоны
        cursor.execute('SELECT COUNT(DISTINCT "Зона") FROM logs_lnk WHERE "Зона" IS NOT NULL AND "Зона" != ""')
        unique_zones = cursor.fetchone()[0]
        
        # Всего негодных теперь считается выше через точный запрос по статусам
        
        # Расчет процентов брака (округление в большую сторону)
        import math
        
        if rt_requested > 0:
            # % брака ОФИЦИАЛЬНЫЙ
            official_defect_percent = math.ceil((rk_defects / rt_requested) * 100)
            # % брака НЕ ОФИЦИАЛЬНЫЙ
            unofficial_defect_percent = math.ceil((rk_np / rt_requested) * 100)
            # % Брака общий
            total_defect_percent = math.ceil((total_defects / rt_requested) * 100)
        else:
            official_defect_percent = 0
            unofficial_defect_percent = 0
            total_defect_percent = 0
        
        conn.close()
        
        return {
            'total_records': total_records,
            'vik_good': vik_good,
            'rk_good': rk_good,
            'rk_defects': rk_defects,
            'rk_np': rk_np,
            'rk_pending': rk_pending,
            'vik_pending': vik_pending,
            'rt_requested': rt_requested,
            'vt_requested': vt_requested,
            'pmi_requested': pmi_requested,
            'unique_drawings': unique_drawings,
            'unique_joints': unique_joints,
            'unique_zones': unique_zones,
            'total_defects': total_defects,
            'official_defect_percent': official_defect_percent,
            'unofficial_defect_percent': unofficial_defect_percent,
            'total_defect_percent': total_defect_percent
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики logs_lnk: {e}")
        conn.close()
        return {
            'total_records': 0,
            'vik_good': 0,
            'rk_good': 0,
            'rk_defects': 0,
            'rk_np': 0,
            'rk_pending': 0,
            'vik_pending': 0,
            'rt_requested': 0,
            'vt_requested': 0,
            'pmi_requested': 0,
            'unique_drawings': 0,
            'unique_joints': 0,
            'unique_zones': 0,
            'total_defects': 0,
            'official_defect_percent': 0,
            'unofficial_defect_percent': 0,
            'total_defect_percent': 0
        }

def _format_last_date(date_value):
    """Форматирует дату и рассчитывает возраст/статус свежести."""
    if not date_value:
        return {
            'formatted': None,
            'ago': None,
            'status': ''
        }

    parsed = None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            parsed = datetime.strptime(str(date_value), fmt)
            break
        except (ValueError, TypeError):
            continue

    if not parsed:
        return {
            'formatted': str(date_value),
            'ago': None,
            'status': ''
        }

    now = datetime.now()
    delta = now - parsed
    days = max(delta.days, 0)

    if days <= 1:
        status = 'fresh'
    elif days <= 7:
        status = 'old'
    else:
        status = 'very_old'

    if days == 0:
        ago = 'сегодня'
    elif days == 1:
        ago = '1 день назад'
    elif 2 <= days <= 4:
        ago = f'{days} дня назад'
    else:
        ago = f'{days} дней назад'

    return {
        'formatted': parsed.strftime('%d.%m.%Y'),
        'ago': ago,
        'status': status
    }

def get_last_update_dates(title='all'):
    """Получает последние даты обновления для dashboard."""
    conn = get_db_connection()
    if not conn:
        return {'logs_lnk': {}, 'wl_china': {}}

    try:
        cursor = conn.cursor()

        logs_where = ""
        logs_params = []
        wl_where = ""
        wl_params = []
        if title and title != 'all':
            logs_where = ' WHERE "Титул" LIKE ? '
            wl_where = ' WHERE "Титул" LIKE ? '
            like_title = f'%{title}%'
            logs_params.append(like_title)
            wl_params.append(like_title)

        cursor.execute(
            f'SELECT MAX("Дата_РК"), MAX("Дата_контроля_ВИК"), MAX("Дата_заявки") FROM logs_lnk {logs_where}',
            logs_params
        )
        logs_row = cursor.fetchone() or (None, None, None)

        cursor.execute(
            f'SELECT MAX("Дата_сварки") FROM wl_china {wl_where}',
            wl_params
        )
        wl_row = cursor.fetchone() or (None,)

        rk_info = _format_last_date(logs_row[0])
        vik_info = _format_last_date(logs_row[1])
        req_info = _format_last_date(logs_row[2])
        weld_info = _format_last_date(wl_row[0])

        return {
            'logs_lnk': {
                'date_rk_formatted': rk_info['formatted'],
                'date_rk_ago': rk_info['ago'],
                'date_rk_status': rk_info['status'],
                'date_vik_formatted': vik_info['formatted'],
                'date_vik_ago': vik_info['ago'],
                'date_vik_status': vik_info['status'],
                'date_request_formatted': req_info['formatted'],
                'date_request_ago': req_info['ago'],
                'date_request_status': req_info['status']
            },
            'wl_china': {
                'date_welding_formatted': weld_info['formatted'],
                'date_welding_ago': weld_info['ago'],
                'date_welding_status': weld_info['status']
            }
        }
    except Exception as e:
        logger.error(f"Ошибка получения последних дат обновления: {e}")
        return {'logs_lnk': {}, 'wl_china': {}}
    finally:
        conn.close()

def get_status_statistics():
    """Получает детальную статистику по статусам РК и ВИК"""
    return get_status_statistics_filtered('all', 'all', 'all')

def get_results_statistics(status_filter='all', title_filter='all', line_filter='all'):
    """Получает статистику по результатам заключений РК и ВИК с фильтром Код_удаления и дополнительными фильтрами"""
    conn = get_db_connection()
    if not conn:
        return {
            'rk_results_stats': [],
            'vik_results_stats': [],
            'total_records': 0
        }
    
    try:
        cursor = conn.cursor()
        
        # Формируем условия фильтрации
        conditions = []
        params = []
        
        if status_filter and status_filter != 'all':
            conditions.append('"Статус_РК" = ?')
            params.append(status_filter)
        
        if title_filter and title_filter != 'all':
            conditions.append('"Титул" LIKE ?')
            params.append(f'%{title_filter}%')
        
        if line_filter and line_filter != 'all':
            conditions.append('"Линия" = ?')
            params.append(line_filter)
        
        # Добавляем условие исключения записей где Код_удаления содержит 'R' (оставляем NULL и пустые)
        conditions.append('("Код_удаления" IS NULL OR "Код_удаления" = "" OR "Код_удаления" = "None" OR "Код_удаления" NOT LIKE "%R%")')
        
        # Формируем WHERE условие
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        # Получаем общее количество записей с фильтром Код_удаления и дополнительными фильтрами
        cursor.execute(f'''
            SELECT COUNT(*) FROM condition_weld 
            WHERE {where_clause}
        ''', params)
        total_records = cursor.fetchone()[0]
        
        # Статистика по Результаты_Заключения_РК с фильтром (только непустые значения)
        cursor.execute(f'''
            SELECT "Результаты_Заключения_РК", COUNT(*) as count
            FROM condition_weld 
            WHERE {where_clause} AND "Результаты_Заключения_РК" IS NOT NULL AND "Результаты_Заключения_РК" != '' AND "Результаты_Заключения_РК" != 'NULL'
            GROUP BY "Результаты_Заключения_РК"
            ORDER BY count DESC
        ''', params)
        rk_results_stats = []
        rk_total = 0
        for result, count in cursor.fetchall():
            rk_results_stats.append({
                'result': result,
                'count': count,
                'percentage': 0  # Временно, будет рассчитано ниже
            })
            rk_total += count
        
        # Рассчитываем проценты от общего количества непустых значений РК
        for stat in rk_results_stats:
            stat['percentage'] = round((stat['count'] / rk_total) * 100, 1) if rk_total > 0 else 0
        
        # Статистика по Результаты_АКТ_ВИК с фильтром (включая пустые как "Не поданно")
        cursor.execute(f'''
            SELECT 
                CASE 
                    WHEN "Результаты_АКТ_ВИК" IS NULL OR "Результаты_АКТ_ВИК" = '' OR "Результаты_АКТ_ВИК" = 'NULL' 
                    THEN 'Не поданно' 
                    ELSE "Результаты_АКТ_ВИК" 
                END as result,
                COUNT(*) as count
            FROM condition_weld 
            WHERE {where_clause}
            GROUP BY 
                CASE 
                    WHEN "Результаты_АКТ_ВИК" IS NULL OR "Результаты_АКТ_ВИК" = '' OR "Результаты_АКТ_ВИК" = 'NULL' 
                    THEN 'Не поданно' 
                    ELSE "Результаты_АКТ_ВИК" 
                END
            ORDER BY count DESC
        ''', params)
        vik_results_stats = []
        for result, count in cursor.fetchall():
            percentage = (count / total_records) * 100 if total_records > 0 else 0
            vik_results_stats.append({
                'result': result,
                'count': count,
                'percentage': round(percentage, 1)
            })
        
        # Записи с пустыми результатами с фильтром
        cursor.execute(f'''
            SELECT COUNT(*) FROM condition_weld 
            WHERE {where_clause} AND ("Результаты_Заключения_РК" IS NULL OR "Результаты_Заключения_РК" = '' OR "Результаты_Заключения_РК" = 'NULL')
        ''', params)
        empty_rk_results_count = cursor.fetchone()[0]
        
        cursor.execute(f'''
            SELECT COUNT(*) FROM condition_weld 
            WHERE {where_clause} AND ("Результаты_АКТ_ВИК" IS NULL OR "Результаты_АКТ_ВИК" = '' OR "Результаты_АКТ_ВИК" = 'NULL')
        ''', params)
        empty_vik_results_count = cursor.fetchone()[0]
        
        conn.close()
        
        return {
            'rk_results_stats': rk_results_stats,
            'vik_results_stats': vik_results_stats,
            'total_records': total_records,
            'empty_rk_results_count': empty_rk_results_count,
            'empty_vik_results_count': empty_vik_results_count,
            'empty_rk_results_percentage': round((empty_rk_results_count / total_records) * 100, 1) if total_records > 0 else 0,
            'empty_vik_results_percentage': round((empty_vik_results_count / total_records) * 100, 1) if total_records > 0 else 0
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики результатов: {e}")
        conn.close()
        return {
            'rk_results_stats': [],
            'vik_results_stats': [],
            'total_records': 0,
            'empty_rk_results_count': 0,
            'empty_vik_results_count': 0,
            'empty_rk_results_percentage': 0,
            'empty_vik_results_percentage': 0
        }

def get_status_statistics_filtered(status_filter='all', title_filter='all', line_filter='all'):
    """Получает детальную статистику по статусам РК и ВИК с учетом фильтров"""
    conn = get_db_connection()
    if not conn:
        return {
            'rk_status_stats': [],
            'vik_status_stats': [],
            'total_records': 0
        }
    
    try:
        cursor = conn.cursor()
        
        # Формируем условия фильтрации
        conditions = []
        params = []
        
        if status_filter and status_filter != 'all':
            conditions.append('"Статус_РК" = ?')
            params.append(status_filter)
        
        if title_filter and title_filter != 'all':
            conditions.append('"Титул" LIKE ?')
            params.append(f'%{title_filter}%')
        
        if line_filter and line_filter != 'all':
            conditions.append('"Линия" = ?')
            params.append(line_filter)
        
        # Добавляем условие исключения записей где Код_удаления содержит 'R' (оставляем NULL и пустые)
        conditions.append('("Код_удаления" IS NULL OR "Код_удаления" = "" OR "Код_удаления" = "None" OR "Код_удаления" NOT LIKE "%R%")')
        
        # Формируем WHERE условие
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        # Получаем общее количество записей с учетом фильтров
        cursor.execute(f'SELECT COUNT(*) FROM condition_weld WHERE {where_clause}', params)
        total_records = cursor.fetchone()[0]
        
        # Статистика по Статус_РК с учетом фильтров (только непустые значения)
        cursor.execute(f'''
            SELECT "Статус_РК", COUNT(*) as count
            FROM condition_weld 
            WHERE {where_clause} AND "Статус_РК" IS NOT NULL AND "Статус_РК" != '' AND "Статус_РК" != 'NULL'
            GROUP BY "Статус_РК"
            ORDER BY count DESC
        ''', params)
        rk_status_stats = []
        rk_total = 0
        for status, count in cursor.fetchall():
            rk_status_stats.append({
                'status': status,
                'count': count,
                'percentage': 0  # Временно, будет рассчитано ниже
            })
            rk_total += count
        
        # Рассчитываем проценты от общего количества непустых значений РК
        for stat in rk_status_stats:
            stat['percentage'] = round((stat['count'] / rk_total) * 100, 1) if rk_total > 0 else 0
        
        # Статистика по Статус_ВИК с учетом фильтров (включая пустые как "Не поданно")
        cursor.execute(f'''
            SELECT 
                CASE 
                    WHEN "Статус_ВИК" IS NULL OR "Статус_ВИК" = '' OR "Статус_ВИК" = 'NULL' 
                    THEN 'Не поданно' 
                    ELSE "Статус_ВИК" 
                END as status,
                COUNT(*) as count
            FROM condition_weld 
            WHERE {where_clause}
            GROUP BY 
                CASE 
                    WHEN "Статус_ВИК" IS NULL OR "Статус_ВИК" = '' OR "Статус_ВИК" = 'NULL' 
                    THEN 'Не поданно' 
                    ELSE "Статус_ВИК" 
                END
            ORDER BY count DESC
        ''', params)
        vik_status_stats = []
        for status, count in cursor.fetchall():
            percentage = (count / total_records) * 100 if total_records > 0 else 0
            vik_status_stats.append({
                'status': status,
                'count': count,
                'percentage': round(percentage, 1)
            })
        
        # Записи с пустыми статусами с учетом фильтров
        cursor.execute(f'''
            SELECT COUNT(*) FROM logs_lnk 
            WHERE {where_clause} AND ("Статус_РК" IS NULL OR "Статус_РК" = '' OR "Статус_РК" = 'NULL')
        ''', params)
        empty_rk_count = cursor.fetchone()[0]
        
        # Для ВИК пустые записи теперь считаются как "Не поданно" в основной статистике
        cursor.execute(f'''
            SELECT COUNT(*) FROM logs_lnk 
            WHERE {where_clause} AND ("Статус_ВИК" IS NULL OR "Статус_ВИК" = '' OR "Статус_ВИК" = 'NULL')
        ''', params)
        empty_vik_count = cursor.fetchone()[0]
        
        conn.close()
        
        return {
            'rk_status_stats': rk_status_stats,
            'vik_status_stats': vik_status_stats,
            'total_records': total_records,
            'empty_rk_count': empty_rk_count,
            'empty_vik_count': empty_vik_count,
            'empty_rk_percentage': round((empty_rk_count / total_records) * 100, 1) if total_records > 0 else 0,
            'empty_vik_percentage': round((empty_vik_count / total_records) * 100, 1) if total_records > 0 else 0
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики статусов: {e}")
        conn.close()
        return {
            'rk_status_stats': [],
            'vik_status_stats': [],
            'total_records': 0,
            'empty_rk_count': 0,
            'empty_vik_count': 0,
            'empty_rk_percentage': 0,
            'empty_vik_percentage': 0
        }

def get_titles_list():
    """Получает список всех титулов из таблицы logs_lnk с TRIM для удаления пробелов и правильной сортировкой"""
    conn = get_db_connection()
    if not conn:
        logger.error("Не удалось подключиться к базе данных для получения списка титулов")
        return []
    
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT TRIM("Титул") as clean_title FROM condition_weld WHERE "Титул" IS NOT NULL AND "Титул" != "" ORDER BY clean_title')
        titles = [row[0] for row in cursor.fetchall()]
        
        # Сортируем титулы в нужном порядке: сначала основные, потом подтитулы
        def custom_sort(title):
            # Основные титулы (без дефиса) идут первыми
            if '-' not in title:
                return (0, title)  # 0 - приоритет для основных титулов
            else:
                return (1, title)  # 1 - приоритет для подтитулов
        
        # Сортируем список с помощью кастомной функции
        sorted_titles = sorted(titles, key=custom_sort)
        
        logger.info(f"Получен список титулов (с TRIM и сортировкой): {sorted_titles}")
        conn.close()
        return sorted_titles
    except Exception as e:
        logger.error(f"Ошибка получения списка титулов: {e}")
        conn.close()
        return []

def get_title_parts_list():
    """Получает список уникальных частей титулов для фильтрации 'содержит'"""
    conn = get_db_connection()
    if not conn:
        logger.error("Не удалось подключиться к базе данных для получения частей титулов")
        return []
    
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT TRIM("Титул") as clean_title FROM condition_weld WHERE "Титул" IS NOT NULL AND "Титул" != "" ORDER BY clean_title')
        titles = [row[0] for row in cursor.fetchall()]
        
        # Извлекаем уникальные части титулов
        title_parts = set()
        for title in titles:
            # Добавляем полный титул
            title_parts.add(title)
            
            # Добавляем основную часть (до дефиса)
            if '-' in title:
                main_part = title.split('-')[0].strip()
                if main_part:
                    title_parts.add(main_part)
            
            # Добавляем части по дефисам
            parts = title.split('-')
            for part in parts:
                part = part.strip()
                if part and len(part) >= 4:  # Минимум 4 символа для значимых частей
                    title_parts.add(part)
        
        # Сортируем части титулов
        sorted_parts = sorted(list(title_parts), key=lambda x: (len(x), x))
        
        logger.info(f"Получен список частей титулов: {sorted_parts}")
        conn.close()
        return sorted_parts
        
    except Exception as e:
        logger.error(f"Ошибка получения частей титулов: {e}")
        conn.close()
        return []

def get_logs_lnk_stats_by_titles(titles=None):
    """Получает статистику из таблицы logs_lnk для конкретных титулов или общую"""
    conn = get_db_connection()
    if not conn:
        return {
            'total_records': 0,
            'vik_good': 0,
            'rk_good': 0,
            'rk_defects': 0,
            'rk_np': 0,
            'rk_pending': 0,
            'vik_pending': 0,
            'rt_requested': 0,
            'vt_requested': 0,
            'pmi_requested': 0,
            'unique_drawings': 0,
            'unique_joints': 0,
            'unique_zones': 0,
            'total_defects': 0,
            'official_defect_percent': 0,
            'unofficial_defect_percent': 0,
            'total_defect_percent': 0
        }
    
    try:
        cursor = conn.cursor()
        
        # Условие для фильтрации по титулам (умная фильтрация)
        if titles and len(titles) > 0:
            # Создаем условие для каждого титула
            title_conditions = []
            title_and_conditions = []
            for title in titles:
                # Для основных титулов (12460, 12470) используем частичное совпадение
                # Для подтитулов (12460-12, 12470-12 и т.д.) используем точное совпадение с учетом пробелов
                if '-' in title:
                    # Подтитул - используем точное совпадение с TRIM для удаления пробелов
                    title_conditions.append(f'TRIM("Титул") = "{title}"')
                    title_and_conditions.append(f'TRIM("Титул") = "{title}"')
                else:
                    # Основной титул - используем частичное совпадение
                    title_conditions.append(f'"Титул" LIKE "%{title}%"')
                    title_and_conditions.append(f'"Титул" LIKE "%{title}%"')
            
            if len(title_conditions) == 1:
                title_condition = f'WHERE {title_conditions[0]}'
                title_and = f'AND {title_and_conditions[0]}'
            else:
                title_condition = f'WHERE ({" OR ".join(title_conditions)})'
                title_and = f'AND ({" OR ".join(title_and_conditions)})'
        else:
            title_condition = ""
            title_and = ""
            titles = []
        
        # Всего записей
        if title_condition:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk {title_condition}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk')
        total_records = cursor.fetchone()[0]
        
        # ВИК Годен
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Статус_ВИК" = "Годен" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_ВИК" = "Годен"')
        vik_good = cursor.fetchone()[0]
        
        # РК Годен
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Годен" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Годен"')
        rk_good = cursor.fetchone()[0]
        
        # РК дефекты: Не годен (статусы требующие исправления)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" LIKE "%Не годен%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" LIKE "%Не годен%"')
        rk_defects = cursor.fetchone()[0]
        
        # РК Н/П (неофициальный ремонт или вырез)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Н/П" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Н/П"')
        rk_np = cursor.fetchone()[0]
        
        # Всего негодных (сумма официальных и неофициальных ремонтов)
        total_defects = rk_defects + rk_np
        
        # РК заявлено не проконтролировано (отставание)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE LOWER("РК") LIKE "%явлен%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE LOWER("РК") LIKE "%явлен%"')
        rk_pending = cursor.fetchone()[0]
        
        # ВИК заявлено не проконтролировано (отставание)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE LOWER("ВИК") LIKE "%явлен%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE LOWER("ВИК") LIKE "%явлен%"')
        vik_pending = cursor.fetchone()[0]
        
        # Заявки на РК (RT)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%RT%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%RT%"')
        rt_requested = cursor.fetchone()[0]
        
        # Заявки на ВИК (VT)
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%VT%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%VT%"')
        vt_requested = cursor.fetchone()[0]
        
        # Заявки на PMI
        if title_and:
            cursor.execute(f'SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%PMI%" {title_and}')
        else:
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%PMI%"')
        pmi_requested = cursor.fetchone()[0]
        
        # Уникальные чертежи
        if title_and:
            cursor.execute(f'SELECT COUNT(DISTINCT "Чертеж") FROM logs_lnk WHERE "Чертеж" IS NOT NULL AND "Чертеж" != "" {title_and}')
        else:
            cursor.execute('SELECT COUNT(DISTINCT "Чертеж") FROM logs_lnk WHERE "Чертеж" IS NOT NULL AND "Чертеж" != ""')
        unique_drawings = cursor.fetchone()[0]
        
        # Уникальные стыки
        if title_and:
            cursor.execute(f'SELECT COUNT(DISTINCT "Номер_стыка") FROM logs_lnk WHERE "Номер_стыка" IS NOT NULL AND "Номер_стыка" != "" {title_and}')
        else:
            cursor.execute('SELECT COUNT(DISTINCT "Номер_стыка") FROM logs_lnk WHERE "Номер_стыка" IS NOT NULL AND "Номер_стыка" != ""')
        unique_joints = cursor.fetchone()[0]
        
        # Уникальные зоны
        if title_and:
            cursor.execute(f'SELECT COUNT(DISTINCT "Зона") FROM logs_lnk WHERE "Зона" IS NOT NULL AND "Зона" != "" {title_and}')
        else:
            cursor.execute('SELECT COUNT(DISTINCT "Зона") FROM logs_lnk WHERE "Зона" IS NOT NULL AND "Зона" != ""')
        unique_zones = cursor.fetchone()[0]
        
        # Всего негодных теперь считается выше через точный запрос по статусам
        
        # Расчет процентов брака (округление в большую сторону)
        import math
        
        if rt_requested > 0:
            # % брака ОФИЦИАЛЬНЫЙ
            official_defect_percent = math.ceil((rk_defects / rt_requested) * 100)
            # % брака НЕ ОФИЦИАЛЬНЫЙ
            unofficial_defect_percent = math.ceil((rk_np / rt_requested) * 100)
            # % Брака общий
            total_defect_percent = math.ceil((total_defects / rt_requested) * 100)
        else:
            official_defect_percent = 0
            unofficial_defect_percent = 0
            total_defect_percent = 0
        
        conn.close()
        
        return {
            'total_records': total_records,
            'vik_good': vik_good,
            'rk_good': rk_good,
            'rk_defects': rk_defects,
            'rk_np': rk_np,
            'rk_pending': rk_pending,
            'vik_pending': vik_pending,
            'rt_requested': rt_requested,
            'vt_requested': vt_requested,
            'pmi_requested': pmi_requested,
            'unique_drawings': unique_drawings,
            'unique_joints': unique_joints,
            'unique_zones': unique_zones,
            'total_defects': total_defects,
            'official_defect_percent': official_defect_percent,
            'unofficial_defect_percent': unofficial_defect_percent,
            'total_defect_percent': total_defect_percent
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики logs_lnk по титулам: {e}")
        conn.close()
        return {
            'total_records': 0,
            'vik_good': 0,
            'rk_good': 0,
            'rk_defects': 0,
            'rk_np': 0,
            'rk_pending': 0,
            'vik_pending': 0,
            'rt_requested': 0,
            'vt_requested': 0,
            'pmi_requested': 0,
            'unique_drawings': 0,
            'unique_joints': 0,
            'unique_zones': 0,
            'total_defects': 0,
            'official_defect_percent': 0,
            'unofficial_defect_percent': 0,
            'total_defect_percent': 0
        }

def get_recent_activities():
    """Получает список последних активностей из логов и базы данных"""
    activities = []
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Получаем последние записи из logs_lnk по дате загрузки
            cursor.execute('''
                SELECT "Дата_загрузки", "Чертеж", "Номер_стыка", "Статус_ВИК", "Статус_РК"
                FROM logs_lnk 
                WHERE "Дата_загрузки" IS NOT NULL 
                ORDER BY "Дата_загрузки" DESC 
                LIMIT 10
            ''')
            recent_records = cursor.fetchall()
            
            for record in recent_records[:3]:  # Показываем только 3 последние
                date_str = record[0] if record[0] else 'Недавно'
                drawing = record[1] if record[1] else 'N/A'
                joint = record[2] if record[2] else 'N/A'
                vik_status = record[3] if record[3] else 'N/A'
                rk_status = record[4] if record[4] else 'N/A'
                
                # Формируем более подробное описание
                status_info = []
                if vik_status and vik_status != 'N/A':
                    status_info.append(f'ВИК: {vik_status}')
                if rk_status and rk_status != 'N/A':
                    status_info.append(f'РК: {rk_status}')
                
                description = f'Чертеж: {drawing}, Стык: {joint}'
                if status_info:
                    description += f' ({", ".join(status_info)})'
                
                activities.append({
                    'icon': 'plus-circle',
                    'action': 'Добавлена запись в logs_lnk',
                    'description': description,
                    'time': date_str
                })
            
            # Получаем информацию о последних операциях с файлами
            cursor.execute('''
                SELECT "Дата_загрузки", "Титул", "Количество_записей"
                FROM (
                    SELECT "Дата_загрузки", "Титул", COUNT(*) as "Количество_записей"
                    FROM logs_lnk 
                    WHERE "Дата_загрузки" IS NOT NULL 
                    GROUP BY "Дата_загрузки", "Титул"
                    ORDER BY "Дата_загрузки" DESC 
                    LIMIT 5
                )
            ''')
            file_operations = cursor.fetchall()
            
            for op in file_operations[:2]:  # Показываем 2 последние операции с файлами
                date_str = op[0] if op[0] else 'Недавно'
                title = op[1] if op[1] else 'Неизвестный титул'
                count = op[2] if op[2] else 0
                
                activities.append({
                    'icon': 'file-upload',
                    'action': 'Загружен файл данных',
                    'description': f'Титул: {title}, Записей: {count}',
                    'time': date_str
                })
            
            # Получаем логированные активности системы
            cursor.execute('''
                SELECT action, description, icon, timestamp
                FROM system_activities 
                ORDER BY timestamp DESC 
                LIMIT 5
            ''')
            system_activities = cursor.fetchall()
            
            for activity in system_activities:
                activities.append({
                    'icon': activity[2] if activity[2] else 'info-circle',
                    'action': activity[0],
                    'description': activity[1] if activity[1] else '',
                    'time': activity[3] if activity[3] else 'Недавно'
                })
            
            conn.close()
            
        except Exception as e:
            logger.error(f"Ошибка получения активностей: {e}")
            conn.close()
    
    # Добавляем статические активности системы, если нет данных в БД
    if not activities:
        system_activities = [
            {
                'icon': 'database',
                'action': 'Система запущена',
                'description': 'Веб-интерфейс M_Kran активен',
                'time': 'При запуске'
            },
            {
                'icon': 'chart-line',
                'action': 'Обновление статистики',
                'description': 'Данные dashboard обновлены',
                'time': 'Сейчас'
            }
        ]
        activities.extend(system_activities)
    
    return activities[:8]  # Возвращаем максимум 8 активностей

def get_duplicates_count():
    """Получает количество найденных дубликатов"""
    conn = get_db_connection()
    if not conn:
        return 0
    
    try:
        cursor = conn.cursor()
        
        # Проверяем наличие таблиц с дубликатами
        tables_to_check = ['wl_china_duplicates', 'wl_report_smr_duplicates']
        total_duplicates = 0
        
        for table in tables_to_check:
            try:
                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
                if cursor.fetchone():
                    cursor.execute(f"SELECT COUNT(*) as count FROM [{table}]")
                    count = cursor.fetchone()['count']
                    total_duplicates += count
            except Exception as e:
                logger.warning(f"Ошибка при проверке дубликатов в таблице {table}: {e}")
                continue
        
        conn.close()
        return total_duplicates
        
    except Exception as e:
        logger.error(f"Ошибка получения количества дубликатов: {e}")
        conn.close()
        return 0

def log_activity(action, description, icon='info-circle'):
    """Логирует активность в базу данных"""
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Создаем таблицу для активностей, если её нет
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_activities (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    action TEXT NOT NULL,
                    description TEXT,
                    icon TEXT DEFAULT 'info-circle',
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Добавляем запись об активности
            cursor.execute('''
                INSERT INTO system_activities (action, description, icon)
                VALUES (?, ?, ?)
            ''', (action, description, icon))
            
            conn.commit()
            conn.close()
            logger.info(f"Активность залогирована: {action} - {description}")
            
        except Exception as e:
            logger.error(f"Ошибка логирования активности: {e}")
            conn.close()

def get_uptime():
    """Получает время работы приложения"""
    # Простая реализация - можно улучшить
    return "24 часа"

def get_etl_scripts():
    """Получает ETL скрипты, организованные по этапам"""
    print("DEBUG: Starting get_etl_scripts()")
    logger.info("Starting get_etl_scripts()")
    etl_scripts = {}
    
    # Папки со скриптами
    scripts_dirs = [
        app.config['SCRIPTS_DIR'],
        os.path.join(app.config['SCRIPTS_DIR'], 'data_loaders'),
        os.path.join(app.config['SCRIPTS_DIR'], 'data_cleaners'),
        os.path.join(app.config['SCRIPTS_DIR'], 'utilities'),
        app.config['PROJECT_ROOT']  # Добавляем корневую папку проекта
    ]
    
    # Собираем все доступные файлы скриптов
    available_scripts = {}
    print(f"DEBUG: Checking scripts directories: {scripts_dirs}")
    for scripts_dir in scripts_dirs:
        if os.path.exists(scripts_dir):
            print(f"DEBUG: Found directory: {scripts_dir}")
            for file in os.listdir(scripts_dir):
                if file.endswith('.py'):
                    script_path = os.path.join(scripts_dir, file).replace('\\', '/')
                    available_scripts[file] = script_path
                    print(f"DEBUG: Found script: {file} -> {script_path}")
        else:
            print(f"DEBUG: Directory not found: {scripts_dir}")
    
    print(f"DEBUG: Total available scripts: {len(available_scripts)}")
    print(f"DEBUG: Available scripts: {list(available_scripts.keys())}")
    
    # Организуем по ETL этапам
    for category_id, category_info in ETL_CATEGORIES.items():
        category_scripts = []
        
        # Обрабатываем скрипты категории
        for script_file, script_info in category_info['scripts'].items():
            if script_file in available_scripts:
                if isinstance(script_info, dict):
                    script_data = {
                        'name': script_file,
                        'path': available_scripts[script_file],
                        'description': script_info['description'],
                        'priority': get_etl_priority(script_file),
                        'etl_type': category_info['type'],
                        'source_files': script_info.get('source_files', []),
                        'source_folder': script_info.get('source_folder', ''),
                        'target_table': script_info.get('target_table', ''),
                        'source_table': script_info.get('source_table', ''),
                        'data_type': script_info.get('data_type', ''),
                        'operations': script_info.get('operations', ''),
                        'exclude_from_bulk_daily': script_info.get('exclude_from_bulk_daily', False),
                    }
                else:
                    script_data = {
                        'name': script_file,
                        'path': available_scripts[script_file],
                        'description': script_info,
                        'priority': get_etl_priority(script_file),
                        'etl_type': category_info['type']
                    }
                
                category_scripts.append(script_data)
        
        # Сортируем по приоритету
        category_scripts.sort(key=lambda x: x['priority'])
        
        if category_scripts:  # Добавляем категорию только если есть скрипты
            etl_scripts[category_id] = {
                'name': category_info['name'],
                'description': category_info['description'],
                'type': category_info['type'],
                'scripts': category_scripts
            }
    
    return etl_scripts

def get_etl_priority(script_name):
    """Определяет приоритет ETL скрипта для правильной последовательности выполнения"""
    priority_map = {
        # EXTRACT этап - ЕЖЕДНЕВНЫЕ скрипты (приоритет 1-10)
        'load_lnk_data.py': 1,                    # Журнал НК НГС - результаты контроля качества
        'run_full_logs_lnk_update.py': 1.25,      # Полное обновление журнала НК (НГС + АКС)
        'load_lnk_nk_aks.py': 1.5,               # Журнал НК АКС → logs_lnk
        'load_staff_titles_M_Kran.py': 2,         # Расстановка персонала М_Кран по участкам
        'load_ndt_findings_transmission_register.py': 3,  # Реестр заключений НГС Эксперт
        'load_wl_report_smr_web.py': 4,           # Отчеты мастеров СМР (оптимизированная версия)
        'load_work_order_log_NDT.py': 5,          # Заявки на НК от М_Кран - планирование работ
        'load_wl_china.py': 6,                    # Данные китайских подрядчиков WELDLOG
        
        # EXTRACT этап - ПЕРИОДИЧЕСКИЕ скрипты (приоритет 11-20)
        'load_Piping_Log_PTO.py': 11,             # Основа - ISO от ПТО
        'load_lst_data.py': 12,                   # Технологический трубопровод
        'load_dl_data.py': 13,                    # Допускные листы на сварщиков
        'load_tks_data.py': 14,                   # Технологические карты
        'load_Pipeline_Test_Package.py': 15,      # Тест-пакеты
        'create_ndt_reports_table.py': 16,        # Загрузить перечень заключений НК
        'load_pipeline_weld_joint_iso.py': 17,    # Сварные соединения ISO - номерация стыков (периодический)
        
        # TRANSFORM этап - преобразование данных (приоритет 21-40)
        'clean_journal_lnk_data.py': 21,          # Очистка основной таблицы
        'sync_pipeline_wl_china.py': 22,          # Синхронизация pipeline_weld_joint_iso и wl_china
        'create_pipeline_weld_joint.py': 23,      # Создание связей стыков
        'merge_duplicates_lnk_logs.py': 23,       # Объединение дубликатов
        'clean_weld_repair_log.py': 24,           # Очистка журнала ремонта швов
        'sync_weld_repair_log.py': 25,            # Синхронизация журнала ремонта
        
        # LOAD этап - финальная загрузка (приоритет 41-60)
        # 'load_pipeline_weld_joint_iso.py': 41,    # Загрузка связей в БД (удалено - теперь в EXTRACT)
    }
    return priority_map.get(script_name, 99)

@app.route('/')
def index():
    """Главная страница - ОБНОВЛЕННАЯ ВЕРСИЯ 2025"""
    # Получаем статистику базы данных
    db_stats = get_db_stats()
    
    # Получаем статистику из logs_lnk
    logs_stats = get_logs_lnk_stats()
    
    # Получаем последние активности
    recent_activities = get_recent_activities()
    
    # Получаем количество дубликатов
    duplicates_count = get_duplicates_count()
    
    # Получаем время работы системы
    uptime = get_uptime()
    
    # Получаем список титулов для выпадающего списка
    titles_list = get_titles_list()
    # Получаем последние даты обновления для блока на главной
    last_update_dates = get_last_update_dates()
    
    # Логируем активность обновления dashboard
    log_activity('Обновление dashboard', f'Загружена главная страница, титулов: {len(titles_list)}', 'chart-line')
    
    # Логируем информацию для отладки
    logger.info(f"Главная страница загружена. Титулов в списке: {len(titles_list)}")
    logger.info(f"Список титулов: {titles_list}")
    
    return render_template('index.html', 
                         db_stats=db_stats,
                         logs_stats=logs_stats,
                         recent_activities=recent_activities,
                         duplicates_count=duplicates_count,
                         uptime=uptime,
                         titles_list=titles_list,
                         last_update_dates=last_update_dates)

@app.route('/api/title_stats')
def api_title_stats():
    """API для получения статистики по выбранному титулу"""
    title_param = request.args.get('titles')
    
    if title_param and title_param != 'all':
        # Один титул
        stats = get_logs_lnk_stats_by_titles([title_param])
    else:
        stats = get_logs_lnk_stats()
    
    return jsonify(stats)

@app.route('/api/etl_scripts')
def api_etl_scripts():
    """API для получения ETL скриптов"""
    try:
        etl_scripts = get_etl_scripts()
        return jsonify(etl_scripts)
    except Exception as e:
        app.logger.error(f"Ошибка при получении ETL скриптов: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/update_stats')
def api_update_stats():
    """API для обновления статистики по титулу"""
    try:
        title = request.args.get('title', 'all')
        logger.info(f"API update_stats вызван с title: {title}")
        
        if title and title != 'all':
            # Получаем статистику для конкретного титула
            logger.info(f"Получаем статистику для титула: {title}")
            stats = get_logs_lnk_stats_by_titles([title])
        else:
            # Получаем общую статистику
            logger.info("Получаем общую статистику")
            stats = get_logs_lnk_stats()
        
        logger.info(f"Статистика получена: {stats}")
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Ошибка обновления статистики: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/update_dates')
def api_update_dates():
    """API для обновления последних дат по выбранному титулу."""
    try:
        title = request.args.get('title', 'all')
        dates = get_last_update_dates(title=title)
        return jsonify({
            'success': True,
            'last_update_dates': dates
        })
    except Exception as e:
        logger.error(f"Ошибка обновления дат: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })



@app.route('/database')
def database():
    """Страница работы с базой данных"""
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return render_template('database.html', tables=[])
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row['name'] for row in cursor.fetchall()]
        conn.close()
        return render_template('database.html', tables=tables)
    except Exception as e:
        flash(f'Ошибка получения списка таблиц: {e}', 'error')
        return render_template('database.html', tables=[])

# Словарь для маппинга технических имен таблиц на понятные названия
TABLE_NAME_MAPPING = {
    'wl_china': 'Журнал сварки',
    'logs_lnk': 'Журнал НГС Эксперт',
    'weld_repair_log': 'Журнал ремонта',
    'nk_results': 'Отчеты НК',
    'condition_weld': 'Последнее состояние стыка',
    'work_order_log_NDT': 'Журнал заказов НК',
    'pipeline_weld_joint_iso': 'Сварные соединения трубопровода ISO',
    'pto_ndt_volume_register': 'Реестр объемов НК ПТО',
    'ndt_findings_transmission_register': 'Реестр результатов НК передачи',
    'pipeline_test_package': 'Пакет испытаний трубопровода',
    'piping_log_pto': 'Журнал трубопроводов ПТО',
    'wl_report_smr': 'Отчет сварки СМР'
}

@app.route('/table/<table_name>')
def view_table(table_name):
    """Просмотр содержимого таблицы с улучшенной фильтрацией"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # Ограничиваем per_page разумными значениями
    if per_page < 10:
        per_page = 10
    elif per_page > 1000:
        per_page = 1000
    
    offset = (page - 1) * per_page
    search_term = request.args.get('search', '')
    selected_columns = request.args.get('columns', '')
    
    # Получаем фильтры по столбцам
    column_filters = {}
    column_filter_types = {}
    for key, value in request.args.items():
        if key.startswith('filter_') and not key.startswith('filter_type_') and value.strip():
            column_name = key[7:]  # Убираем префикс 'filter_'
            column_filters[column_name] = value.strip()
            
            # Получаем тип фильтра
            filter_type_key = f'filter_type_{column_name}'
            filter_type = request.args.get(filter_type_key, 'contains')
            column_filter_types[column_name] = filter_type
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('database'))
    
    try:
        cursor = conn.cursor()
        
        # Получаем все названия столбцов
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        all_columns = [col['name'] for col in cursor.fetchall()]
        
        # Определяем какие столбцы показывать (как в logs_lnk_table)
        if selected_columns and selected_columns.strip():
            # Очищаем строку от лишних символов
            cleaned_columns = selected_columns.strip()
            if cleaned_columns.startswith('[') and cleaned_columns.endswith(']'):
                # Это JSON-подобная строка, извлекаем содержимое
                import ast
                try:
                    selected_columns_list = ast.literal_eval(cleaned_columns)
                except:
                    # Если не удалось распарсить, используем split
                    selected_columns_list = [col.strip().strip("'\"") for col in cleaned_columns[1:-1].split(',')]
            else:
                # Обычная строка через запятую
                selected_columns_list = [col.strip() for col in selected_columns.split(',')]
            
            # Фильтруем только существующие столбцы и убираем пустые значения
            
            # Создаем словарь для сопоставления столбцов (убираем лишние пробелы)
            column_mapping = {}
            for db_col in all_columns:
                # Создаем варианты с разным количеством пробелов
                normalized_col = db_col.strip()
                column_mapping[normalized_col] = db_col
                # Также добавляем вариант с одним пробелом в конце
                if normalized_col.endswith('Material'):
                    column_mapping[normalized_col + ' '] = db_col
            
            columns_to_show = []
            for col in selected_columns_list:
                col_stripped = col.strip()
                if col_stripped in column_mapping:
                    columns_to_show.append(column_mapping[col_stripped])
                elif col in all_columns:
                    columns_to_show.append(col)
            
            # Если после фильтрации не осталось столбцов, показываем все
            if not columns_to_show:
                columns_to_show = all_columns
        else:
            # По умолчанию показываем ВСЕ столбцы
            columns_to_show = all_columns
        
        # Формируем условие поиска и фильтрации
        where_conditions = []
        search_params = []
        
        # Добавляем общий поиск
        if search_term:
            # Поиск по всем видимым столбцам (регистронезависимый)
            search_conditions = []
            for col in columns_to_show:
                # Используем двойной поиск: оригинальный термин + нижний регистр
                search_conditions.append(f'(CAST(`{col}` AS TEXT) LIKE ? OR LOWER(CAST(`{col}` AS TEXT)) LIKE ?)')
                search_params.append(f'%{search_term}%')
                search_params.append(f'%{search_term.lower()}%')
            where_conditions.append(f"({' OR '.join(search_conditions)})")
        
        # Добавляем фильтры по столбцам
        for column_name, filter_value in column_filters.items():
            if column_name in all_columns:  # Проверяем, что столбец существует
                filter_type = column_filter_types.get(column_name, 'contains')
                
                # Специальная обработка для столбца ФИО в таблице слов_клейм_факт
                if table_name == 'слов_клейм_факт' and column_name == 'ФИО':
                    if filter_type == 'contains':
                        # Ищем в полном ФИО из таблицы ФИО_свар
                        where_conditions.append(f'(CAST(f.ФИО AS TEXT) LIKE ? OR LOWER(CAST(f.ФИО AS TEXT)) LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'not_contains':
                        # Исключение записей, содержащих значение в полном ФИО
                        where_conditions.append(f'(CAST(f.ФИО AS TEXT) NOT LIKE ? AND LOWER(CAST(f.ФИО AS TEXT)) NOT LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'empty':
                        # Показывать только пустые значения ФИО
                        where_conditions.append(f'(f.ФИО IS NULL OR CAST(f.ФИО AS TEXT) = \'\' OR TRIM(CAST(f.ФИО AS TEXT)) = \'\' OR CAST(f.ФИО AS TEXT) = \'-\' OR CAST(f.ФИО AS TEXT) = \'None\')')
                    elif filter_type == 'not_empty':
                        # Показывать только непустые значения ФИО
                        where_conditions.append(f'(f.ФИО IS NOT NULL AND CAST(f.ФИО AS TEXT) != \'\' AND TRIM(CAST(f.ФИО AS TEXT)) != \'\' AND CAST(f.ФИО AS TEXT) != \'-\' AND CAST(f.ФИО AS TEXT) != \'None\')')
                else:
                    # Обычная обработка для других столбцов
                    # Пропускаем если фильтр 'special' (для empty/not_empty)
                    if filter_value == 'special':
                        if filter_type == 'empty':
                            # Показывать только пустые значения (NULL, пустая строка, только пробелы, "-", "None", "null")
                            where_conditions.append(f'(`{column_name}` IS NULL OR CAST(`{column_name}` AS TEXT) = \'\' OR TRIM(CAST(`{column_name}` AS TEXT)) = \'\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'-\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'NONE\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'NULL\')')
                        elif filter_type == 'not_empty':
                            # Показывать только непустые значения - исключаем NULL, пустые строки, пробелы, "-", "None", "null"
                            where_conditions.append(f'(`{column_name}` IS NOT NULL AND CAST(`{column_name}` AS TEXT) != \'\' AND TRIM(CAST(`{column_name}` AS TEXT)) != \'\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'-\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'NONE\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'NULL\' AND LENGTH(TRIM(CAST(`{column_name}` AS TEXT))) > 0)')
                    elif filter_type == 'contains':
                        # Регистронезависимый поиск по конкретному столбцу
                        where_conditions.append(f'(CAST(`{column_name}` AS TEXT) LIKE ? OR LOWER(CAST(`{column_name}` AS TEXT)) LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'not_contains':
                        # Исключение записей, содержащих значение
                        where_conditions.append(f'(CAST(`{column_name}` AS TEXT) NOT LIKE ? AND LOWER(CAST(`{column_name}` AS TEXT)) NOT LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'empty':
                        # Показывать только пустые значения (NULL, пустая строка, только пробелы, "-", "None", "null")
                        where_conditions.append(f'(`{column_name}` IS NULL OR CAST(`{column_name}` AS TEXT) = \'\' OR TRIM(CAST(`{column_name}` AS TEXT)) = \'\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'-\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'NONE\' OR UPPER(TRIM(CAST(`{column_name}` AS TEXT))) = \'NULL\')')
                    elif filter_type == 'not_empty':
                        # Показывать только непустые значения - исключаем NULL, пустые строки, пробелы, "-", "None", "null"
                        where_conditions.append(f'(`{column_name}` IS NOT NULL AND CAST(`{column_name}` AS TEXT) != \'\' AND TRIM(CAST(`{column_name}` AS TEXT)) != \'\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'-\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'NONE\' AND UPPER(TRIM(CAST(`{column_name}` AS TEXT))) != \'NULL\' AND LENGTH(TRIM(CAST(`{column_name}` AS TEXT))) > 0)')
        
        # Формируем итоговое условие WHERE
        if where_conditions:
            search_condition = f"WHERE {' AND '.join(where_conditions)}"
        else:
            search_condition = ""
        
        # Получаем общее количество записей
        # Специальная обработка для таблицы слов_клейм_факт
        if table_name == 'слов_клейм_факт':
            # Правильно заменяем только названия столбцов в WHERE условии
            modified_search_condition = search_condition
            if modified_search_condition:
                # Заменяем только названия столбцов на s.название_столбца
                for col in all_columns:
                    if col != 'ФИО':  # ФИО уже обрабатывается как f.ФИО в фильтрах
                        modified_search_condition = modified_search_condition.replace(f'`{col}`', f's.`{col}`')
            
            count_query = f"SELECT COUNT(*) as count FROM `{table_name}` s LEFT JOIN `ФИО_свар` f ON s.ФИО = f.id_fio {modified_search_condition}"
        else:
            count_query = f"SELECT COUNT(*) as count FROM `{table_name}` {search_condition}"
        
        cursor.execute(count_query, search_params)
        total_records = cursor.fetchone()['count']
        
        # Получаем данные с пагинацией
        # Специальная обработка для таблицы слов_клейм_факт - делаем JOIN с ФИО_свар
        if table_name == 'слов_клейм_факт':
            # Заменяем столбец ФИО на полное ФИО из таблицы ФИО_свар
            modified_columns = []
            for col in columns_to_show:
                if col == 'ФИО':
                    modified_columns.append('COALESCE(f.ФИО, s.ФИО) as ФИО')
                else:
                    modified_columns.append(f's.`{col}`')
            
            columns_str = ', '.join(modified_columns)
            
            # Правильно заменяем только названия столбцов в WHERE условии
            modified_search_condition = search_condition
            if modified_search_condition:
                # Заменяем только названия столбцов на s.название_столбца
                for col in all_columns:
                    if col != 'ФИО':  # ФИО уже обрабатывается как f.ФИО в фильтрах
                        modified_search_condition = modified_search_condition.replace(f'`{col}`', f's.`{col}`')
            
            data_query = f"""
                SELECT {columns_str}
                FROM `{table_name}` s
                LEFT JOIN `ФИО_свар` f ON s.ФИО = f.id_fio
                {modified_search_condition}
                LIMIT {per_page} OFFSET {offset}
            """
        else:
            columns_str = ', '.join([f'`{col}`' for col in columns_to_show])
            data_query = f"""
                SELECT {columns_str}
                FROM `{table_name}` 
                {search_condition}
                LIMIT {per_page} OFFSET {offset}
            """
        
        cursor.execute(data_query, search_params)
        records = cursor.fetchall()
        
        # Преобразуем в список словарей (как в logs_lnk_table)
        records_list = []
        for record in records:
            record_dict = {}
            for i, col in enumerate(columns_to_show):
                try:
                    value = record[i]
                    # Применяем безопасное кодирование
                    record_dict[col] = safe_encode_value(value)
                except Exception as col_error:
                    record_dict[col] = '[ОШИБКА КОДИРОВКИ]'
            records_list.append(record_dict)
        
        if records_list:
            # Безопасно выводим первую запись, избегая проблем с кодировкой
            first_record = records_list[0]
            safe_first_record = {}
            for key, value in first_record.items():
                try:
                    safe_first_record[key] = str(value).encode('ascii', 'replace').decode('ascii')
                except:
                    safe_first_record[key] = '[ОШИБКА КОДИРОВКИ]'
            
            # Безопасно выводим значения
            safe_values = []
            for value in first_record.values():
                try:
                    safe_values.append(str(value).encode('ascii', 'replace').decode('ascii'))
                except:
                    safe_values.append('[ОШИБКА КОДИРОВКИ]')
        
        total_pages = (total_records + per_page - 1) // per_page
        
        conn.close()
        
        # Получаем понятное название таблицы
        table_title = TABLE_NAME_MAPPING.get(table_name, table_name)
        
        logger.debug(
            f"[DEBUG] Передаем в шаблон: all_columns={len(all_columns) if 'all_columns' in locals() else 'НЕ ОПРЕДЕЛЕНА'}, "
            f"selected_columns={len(columns_to_show) if 'columns_to_show' in locals() else 'НЕ ОПРЕДЕЛЕНА'}"
        )
        if 'all_columns' in locals():
            logger.debug(f"[DEBUG] all_columns: {all_columns}")
        else:
            logger.debug("[DEBUG] all_columns НЕ ОПРЕДЕЛЕНА")
        if 'columns_to_show' in locals():
            logger.debug(f"[DEBUG] columns_to_show: {columns_to_show}")
        else:
            logger.debug("[DEBUG] columns_to_show НЕ ОПРЕДЕЛЕНА")
        
        return render_template('table_view.html', 
                             table_name=table_name,
                             table_title=table_title,
                             records=records_list,
                             columns=columns_to_show,
                             all_columns=all_columns,
                             selected_columns=columns_to_show,
                             page=page,
                             per_page=per_page,
                             total_pages=total_pages,
                             total_records=total_records,
                             search_term=search_term,
                             column_filters=column_filters,
                             column_filter_types=column_filter_types)
        
    except Exception as e:
        logger.error(f"Ошибка в view_table: {e}", exc_info=True)
        
        # Безопасно обрабатываем сообщение об ошибке
        try:
            error_msg = str(e)
            safe_error_msg = ''.join(char if ord(char) < 128 else '?' for char in error_msg)
            flash(f'Ошибка получения данных таблицы: {safe_error_msg}', 'error')
        except:
            flash('Ошибка получения данных таблицы: Неизвестная ошибка кодировки', 'error')
        
        return redirect(url_for('database'))

@app.route('/debug_table/<table_name>')
def debug_table_view(table_name):
    """Отладочный просмотр таблицы"""
    try:
        conn = get_db_connection()
        if not conn:
            flash('Ошибка подключения к базе данных', 'error')
            return redirect(url_for('database'))
        
        cursor = conn.cursor()
        
        # Получаем информацию о столбцах
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        table_info = cursor.fetchall()
        all_columns = [col['name'] for col in table_info]
        
        # Получаем первые 5 записей для отладки
        columns_str = ', '.join([f'`{col}`' for col in all_columns])
        query = f"SELECT {columns_str} FROM `{table_name}` LIMIT 5"
        cursor.execute(query)
        records = cursor.fetchall()
        
        # Преобразуем в список словарей
        records_list = []
        for record in records:
            record_dict = {}
            for i, col in enumerate(all_columns):
                try:
                    value = record[i]
                    record_dict[col] = safe_encode_value(value)
                except Exception:
                    record_dict[col] = '[ОШИБКА]'
            records_list.append(record_dict)
        
        conn.close()
        
        # Получаем понятное название таблицы
        table_title = TABLE_NAME_MAPPING.get(table_name, table_name)
        
        return render_template('debug_table.html',
                             table_name=table_name,
                             table_title=table_title,
                             records=records_list,
                             columns=all_columns)
        
    except Exception as e:
        flash(f'Ошибка отладки таблицы: {str(e)}', 'error')
        return redirect(url_for('database'))

@app.route('/api/table/<table_name>/search')
def search_table(table_name):
    """API для поиска в таблице"""
    search_query = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    selected_columns = request.args.get('columns', '')
    per_page = 10
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к базе данных'})
    
    try:
        cursor = conn.cursor()
        
        # Получаем все названия столбцов
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        all_columns = [col['name'] for col in cursor.fetchall()]
        
        # Определяем какие столбцы показывать
        if selected_columns and selected_columns.strip():
            selected_columns_list = selected_columns.split(',')
            # Фильтруем только существующие столбцы и убираем пустые значения
            columns_to_show = [col for col in selected_columns_list if col in all_columns and col.strip()]
            if not columns_to_show:
                columns_to_show = all_columns
        else:
            columns_to_show = all_columns
        
        # Формируем SQL запрос с фильтрацией
        columns_sql = ', '.join([f'`{col}`' for col in columns_to_show])
        
        # Добавляем условие поиска (регистронезависимый)
        where_clause = ""
        params = []
        if search_query:
            search_conditions = []
            for col in columns_to_show:
                # Используем двойной поиск: оригинальный термин + нижний регистр
                search_conditions.append(f'(CAST(`{col}` AS TEXT) LIKE ? OR LOWER(CAST(`{col}` AS TEXT)) LIKE ?)')
                params.append(f'%{search_query}%')
                params.append(f'%{search_query.lower()}%')
            where_clause = f"WHERE {' OR '.join(search_conditions)}"
        
        # Получаем общее количество записей с учетом фильтра
        count_query = f"SELECT COUNT(*) as count FROM `{table_name}` {where_clause}"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()['count']
        
        # Получаем данные с пагинацией и фильтрацией
        data_query = f"SELECT {columns_sql} FROM `{table_name}` {where_clause} LIMIT {per_page} OFFSET {offset}"
        cursor.execute(data_query, params)
        records = cursor.fetchall()
        
        # Преобразуем записи в список словарей
        records_list = []
        for record in records:
            record_dict = {}
            for i, col in enumerate(columns_to_show):
                record_dict[col] = record[i]
            records_list.append(record_dict)
        
        conn.close()
        
        total_pages = (total_records + per_page - 1) // per_page
        
        return jsonify({
            'success': True,
            'records': records_list,
            'total_records': total_records,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'columns': columns_to_show
        })
        
    except Exception as e:
        return jsonify({'error': f'Ошибка поиска: {str(e)}'})

@app.route('/scripts')
def scripts():
    """Страница ETL процессов - Extract, Transform, Load"""
    try:
        print("DEBUG: Начинаем загрузку ETL скриптов...")
        logger.info("Начинаем загрузку ETL скриптов...")
        
        etl_scripts = get_etl_scripts()
        print(f"DEBUG: Получено {len(etl_scripts)} категорий ETL скриптов")
        logger.info(f"Получено {len(etl_scripts)} категорий ETL скриптов")
        
        for stage_id, stage in etl_scripts.items():
            try:
                print(f"DEBUG: {stage_id}: {stage['name']} - {len(stage['scripts'])} скриптов")
                logger.info(f"{stage_id}: {stage['name']} - {len(stage['scripts'])} скриптов")
            except UnicodeEncodeError:
                # Если есть проблемы с кодировкой, используем ASCII
                safe_name = stage['name'].encode('ascii', 'ignore').decode('ascii')
                print(f"DEBUG: {stage_id}: {safe_name} - {len(stage['scripts'])} скриптов")
                logger.info(f"{stage_id}: {safe_name} - {len(stage['scripts'])} скриптов")
        
        full_logs_lnk_update_script = os.path.normpath(
            os.path.join(Config.SCRIPTS_DIR, 'data_loaders', 'run_full_logs_lnk_update.py')
        ).replace('\\', '/')
        logger.info("Рендерим шаблон scripts.html")
        return render_template(
            'scripts.html',
            etl_scripts=etl_scripts,
            full_logs_lnk_update_script=full_logs_lnk_update_script,
        )
    except Exception as e:
        print(f"DEBUG: Ошибка в функции scripts(): {e}")
        logger.error(f"Ошибка в функции scripts(): {e}")
        import traceback
        traceback.print_exc()
        return f"Ошибка: {e}", 500

@app.route('/run_script', methods=['POST'])
def run_script():
    """Запуск скрипта с использованием нового ScriptRunner"""
    # Проверяем, пришли ли данные как JSON или как форма
    if request.is_json:
        data = request.get_json()
        script_path = data.get('script_path')
        script_args = data.get('script_args', [])
    else:
        script_path = request.form.get('script_path')
        script_args = []
    
    # Детальное логирование для отладки
    logger.info(f"Получен запрос на запуск скрипта")
    logger.info(f"Полученный путь: {repr(script_path)}")
    logger.info(f"Текущая рабочая директория: {os.getcwd()}")
    
    if not script_path:
        logger.error("Путь к скрипту не указан")
        return jsonify({'success': False, 'message': 'Путь к скрипту не указан'})
    

    
    # Нормализуем путь и исправляем возможные проблемы с разделителями
    script_path = script_path.replace('/', os.sep)
    script_path = os.path.normpath(script_path)
    logger.info(f"Нормализованный путь: {repr(script_path)}")
    logger.info(f"Существует ли файл: {os.path.exists(script_path)}")
    
    # Если путь все еще неправильный, попробуем восстановить его
    if not os.path.exists(script_path) and '\\' not in script_path and '/' not in script_path:
        logger.info("Пытаемся восстановить путь с потерянными разделителями")
        if script_path.startswith('D:МК_Кран'):
            restored_path = script_path.replace('D:МК_Кран', 'D:\\МК_Кран\\')
            restored_path = restored_path.replace('script_M_Kran', 'script_M_Kran\\')
            restored_path = restored_path.replace('loud_M_Kran_Kingesepp', 'loud_M_Kran_Kingesepp\\')
            restored_path = restored_path.replace('SQLite_data_cleansing', 'SQLite_data_cleansing\\')
            restored_path = restored_path.replace('\\\\', '\\')
            if restored_path.endswith('\\'):
                restored_path = restored_path[:-1]
            
            logger.info(f"Восстановленный путь: {repr(restored_path)}")
            if os.path.exists(restored_path):
                script_path = restored_path
                logger.info("Путь успешно восстановлен!")
            else:
                logger.error("Не удалось восстановить путь")
    
    if not os.path.exists(script_path):
        logger.error(f"Скрипт не найден по пути: {script_path}")
        
        # Расширенный поиск скриптов в подпапках
        script_name = os.path.basename(script_path)
        possible_paths = [
            script_path,
            os.path.join(os.getcwd(), script_path),
            os.path.join(Config.PROJECT_ROOT, script_path),
            os.path.join(Config.SCRIPTS_DIR, script_name)
        ]
        
        # Поиск в подпапках scripts
        if os.path.exists(Config.SCRIPTS_DIR):
            for root, dirs, files in os.walk(Config.SCRIPTS_DIR):
                if script_name in files:
                    found_path = os.path.join(root, script_name)
                    possible_paths.append(found_path)
                    logger.info(f"Найден скрипт в подпапке: {found_path}")
        
        logger.info("Проверяем возможные пути:")
        for path in possible_paths:
            exists = os.path.exists(path)
            logger.info(f"  {path}: {exists}")
            if exists:
                script_path = path
                logger.info(f"Используем найденный путь: {script_path}")
                break
        
        if not os.path.exists(script_path):
            logger.error(f"Скрипт не найден ни по одному из путей")
            return jsonify({'success': False, 'message': f'Скрипт не найден: {script_path}'})
    
    try:
        # Используем новый ScriptRunner
        script_runner = get_script_runner()
        script_id = script_runner.run_script_async(script_path, script_args)
        
        # Логируем активность запуска скрипта
        script_name = os.path.basename(script_path)
        log_activity('Запуск скрипта', f'Запущен скрипт: {script_name}', 'play-circle')
        
        return jsonify({
            'success': True, 
            'message': 'Скрипт запущен', 
            'script_id': script_id
        })
    except Exception as e:
        logger.error(f"Ошибка запуска скрипта: {e}")
        # Логируем ошибку
        script_name = os.path.basename(script_path) if script_path else 'Неизвестный скрипт'
        log_activity('Ошибка запуска', f'Ошибка при запуске скрипта: {script_name}', 'exclamation-triangle')
        return jsonify({'success': False, 'message': f'Ошибка запуска: {e}'})

@app.route('/run_etl_stage', methods=['POST'])
def run_etl_stage():
    """Запуск всех скриптов определенного этапа ETL"""
    try:
        data = request.get_json()
        stage = data.get('stage')
        
        if not stage:
            return jsonify({'success': False, 'error': 'Этап не указан'})
        
        # Получаем скрипты для указанного этапа
        etl_scripts = get_etl_scripts()
        if stage not in etl_scripts:
            return jsonify({'success': False, 'error': f'Неизвестный этап: {stage}'})
        
        stage_scripts = etl_scripts[stage].get('scripts', [])
        if not stage_scripts:
            return jsonify({'success': False, 'error': f'Нет скриптов для этапа {stage}'})
        
        # Запускаем все скрипты этапа
        script_runner = get_script_runner()
        results = []
        
        for script in stage_scripts:
            try:
                script_id = script_runner.run_script_async(script['path'])
                results.append({
                    'script': script['name'],
                    'script_id': script_id,
                    'status': 'started'
                })
            except Exception as e:
                results.append({
                    'script': script['name'],
                    'error': str(e),
                    'status': 'failed'
                })
        
        return jsonify({
            'success': True,
            'message': f'Запущено {len(results)} скриптов для этапа {stage}',
            'results': results
        })
        
    except Exception as e:
        logger.error(f"Ошибка запуска этапа ETL: {e}")
        return jsonify({'success': False, 'message': f'Ошибка запуска этапа: {e}'})

@app.route('/run_all_etl', methods=['POST'])
def run_all_etl():
    """Запуск всех скриптов ETL в правильном порядке"""
    try:
        etl_scripts = get_etl_scripts()
        script_runner = get_script_runner()
        results = []
        
        # Запускаем скрипты в правильном порядке
        stages = ['extract', 'transform', 'load']
        
        for stage in stages:
            if stage in etl_scripts:
                stage_scripts = etl_scripts[stage].get('scripts', [])
                for script in stage_scripts:
                    try:
                        script_id = script_runner.run_script_async(script['path'])
                        results.append({
                            'stage': stage,
                            'script': script['name'],
                            'script_id': script_id,
                            'status': 'started'
                        })
                    except Exception as e:
                        results.append({
                            'stage': stage,
                            'script': script['name'],
                            'error': str(e),
                            'status': 'failed'
                        })
        
        return jsonify({
            'success': True,
            'message': f'Запущено {len(results)} скриптов ETL',
            'results': results
        })
        
    except Exception as e:
        logger.error(f"Ошибка запуска ETL: {e}")
        return jsonify({'success': False, 'message': f'Ошибка запуска ETL: {e}'})

@app.route('/run_etl_pipeline', methods=['POST'])
def run_etl_pipeline():
    """Запуск полного ETL pipeline (Extract -> Transform -> Load)"""
    try:
        # Получаем все скрипты
        etl_scripts = get_etl_scripts()
        script_runner = get_script_runner()
        results = []
        
        # Запускаем скрипты в правильном порядке
        stages = ['extract', 'transform', 'load']
        
        for stage in stages:
            if stage in etl_scripts:
                stage_scripts = etl_scripts[stage].get('scripts', [])
                for script in stage_scripts:
                    try:
                        script_id = script_runner.run_script_async(script['path'])
                        results.append({
                            'stage': stage,
                            'script': script['name'],
                            'script_id': script_id,
                            'status': 'started'
                        })
                    except Exception as e:
                        results.append({
                            'stage': stage,
                            'script': script['name'],
                            'error': str(e),
                            'status': 'failed'
                        })
        
        return jsonify({
            'success': True,
            'message': f'Запущен полный ETL pipeline',
            'results': results
        })
        
    except Exception as e:
        logger.error(f"Ошибка запуска ETL pipeline: {e}")
        return jsonify({'success': False, 'error': f'Ошибка запуска: {e}'})



@app.route('/script_status/<script_id>')
def script_status(script_id):
    """Получает статус выполнения скрипта"""
    try:
        log_offset = request.args.get('log_offset', 0, type=int) or 0
        script_runner = get_script_runner()
        status = script_runner.get_script_status(script_id)
        
        if status is None:
            return jsonify({'success': False, 'message': 'Скрипт не найден'})

        live_log = status.get('live_log', []) or []
        if log_offset < 0:
            log_offset = 0
        if log_offset > len(live_log):
            log_offset = len(live_log)
        new_log_lines = live_log[log_offset:]
        
        return jsonify({
            'success': True,
            'status': status,
            'log_lines': new_log_lines,
            'next_log_offset': len(live_log)
        })
    except Exception as e:
        logger.error(f"Ошибка получения статуса скрипта: {e}")
        return jsonify({'success': False, 'message': f'Ошибка получения статуса: {e}'})

@app.route('/duplicates')
def duplicates():
    """Страница работы с дубликатами"""
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return render_template('duplicates.html', tables=[])
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'duplicates_%'")
        duplicate_tables = [row['name'] for row in cursor.fetchall()]
        conn.close()
        
        return render_template('duplicates.html', tables=duplicate_tables)
    except Exception as e:
        flash(f'Ошибка получения таблиц дубликатов: {e}', 'error')
        return render_template('duplicates.html', tables=[])

@app.route('/duplicates/<table_name>')
def view_duplicates(table_name):
    """Просмотр дубликатов"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('duplicates'))
    
    try:
        cursor = conn.cursor()
        
        # Получаем количество записей
        cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
        total_records = cursor.fetchone()['count']
        
        # Получаем данные с пагинацией
        cursor.execute(f"SELECT * FROM `{table_name}` LIMIT {per_page} OFFSET {offset}")
        records = cursor.fetchall()
        
        # Получаем названия столбцов
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        columns = [col['name'] for col in cursor.fetchall()]
        
        conn.close()
        
        total_pages = (total_records + per_page - 1) // per_page
        
        # Получаем понятное название таблицы
        table_title = TABLE_NAME_MAPPING.get(table_name, table_name)
        
        return render_template('duplicates_view.html', 
                             table_name=table_name,
                             table_title=table_title,
                             records=records,
                             columns=columns,
                             page=page,
                             total_pages=total_pages,
                             total_records=total_records)
    except Exception as e:
        flash(f'Ошибка получения дубликатов: {e}', 'error')
        return redirect(url_for('duplicates'))

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """Страница загрузки файлов"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран', 'error')
            return redirect(request.url)
        
        if file and file.filename:
            # Используем нашу безопасную функцию вместо secure_filename
            filename = safe_filename(file.filename)
            # Получаем уникальное имя файла, если такой уже существует
            unique_filename = get_unique_filename(app.config['UPLOAD_FOLDER'], filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            
            try:
                file.save(filepath)
                logger.info(f"Файл сохранен: {file.filename} -> {unique_filename}")
                
                # Логируем активность загрузки файла
                file_size = os.path.getsize(filepath)
                log_activity('Загрузка файла', f'Загружен файл: {unique_filename} ({file_size} байт)', 'upload')
                
                flash(f'Файл "{unique_filename}" успешно загружен', 'success')
            except Exception as e:
                logger.error(f"Ошибка сохранения файла {file.filename}: {e}")
                
                # Логируем ошибку загрузки
                log_activity('Ошибка загрузки', f'Ошибка при загрузке файла: {file.filename}', 'exclamation-triangle')
                
                flash(f'Ошибка загрузки файла: {e}', 'error')
            
            return redirect(url_for('upload'))
    
    # Показываем список загруженных файлов
    uploaded_files = []
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                uploaded_files.append({
                    'name': filename,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime)
                })
    
    return render_template('upload.html', files=uploaded_files)

@app.route('/export/<table_name>')
def export_table(table_name):
    """Экспорт таблицы в Excel"""
    selected_columns = request.args.get('columns', '')
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('database'))
    
    try:
        # Определяем какие столбцы экспортировать
        if selected_columns:
            selected_columns_list = selected_columns.split(',')
            # Получаем все столбцы для проверки
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info(`{table_name}`)")
            all_columns = [col['name'] for col in cursor.fetchall()]
            
            # Фильтруем только существующие столбцы
            columns_to_export = [col for col in selected_columns_list if col in all_columns]
            if not columns_to_export:
                columns_to_export = all_columns
            
            # Специальная обработка для таблицы слов_клейм_факт
            if table_name == 'слов_клейм_факт':
                # Заменяем столбец ФИО на полное ФИО из таблицы ФИО_свар
                modified_columns = []
                for col in columns_to_export:
                    if col == 'ФИО':
                        modified_columns.append('COALESCE(f.ФИО, s.ФИО) as ФИО')
                    else:
                        modified_columns.append(f's.`{col}`')
                
                columns_sql = ', '.join(modified_columns)
                query = f"""
                    SELECT {columns_sql}
                    FROM `{table_name}` s
                    LEFT JOIN `ФИО_свар` f ON s.ФИО = f.id_fio
                """
            else:
                # Формируем SQL запрос с выбранными столбцами
                columns_sql = ', '.join([f'`{col}`' for col in columns_to_export])
                query = f"SELECT {columns_sql} FROM `{table_name}`"
            
            df = pd.read_sql_query(query, conn)
        else:
            # Специальная обработка для таблицы слов_клейм_факт
            if table_name == 'слов_клейм_факт':
                query = """
                    SELECT s.id, s.Фактическое_Клеймо, COALESCE(f.ФИО, s.ФИО) as ФИО, s.Примечание
                    FROM слов_клейм_факт s
                    LEFT JOIN ФИО_свар f ON s.ФИО = f.id_fio
                """
            else:
                # Экспортируем все столбцы
                query = f"SELECT * FROM `{table_name}`"
            
            df = pd.read_sql_query(query, conn)
        
        conn.close()
        
        # Очищаем данные от проблемных символов
        df = clean_data_for_excel(df)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Ограничиваем длину имени листа
                sheet_name = table_name[:31] if len(table_name) > 31 else table_name
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Получаем рабочий лист для форматирования
                worksheet = writer.sheets[sheet_name]
                
                # Автоматически подгоняем ширину столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Проверяем заголовок столбца
                    if column[0].value:
                        max_length = len(str(column[0].value))
                    
                    # Проверяем данные в столбце
                    for cell in column[1:]:
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except (TypeError, AttributeError):
                            pass
                    
                    # Устанавливаем ширину с ограничениями
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Устанавливаем свойства документа
                workbook = writer.book
                workbook.properties.title = f"Экспорт таблицы {table_name}"
                workbook.properties.creator = "M_Kran System"
                workbook.properties.created = datetime.now()
                
        except Exception as excel_error:
            # Если произошла ошибка при создании Excel, создаем CSV
            print(f"Ошибка создания Excel: {excel_error}")
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            timestamp = get_filename_timestamp()
            filename = f"{table_name}_{timestamp}.csv"
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{table_name}_{get_filename_timestamp()}.xlsx"
        )
    except Exception as e:
        flash(f'Ошибка экспорта: {e}', 'error')
        return redirect(url_for('database'))

@app.route('/export_table_structure/<table_name>')
def export_table_structure(table_name):
    """Экспорт структуры таблицы (имя таблицы, столбцы, типы) в Excel"""
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('database'))

    try:
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        pragma_rows = cursor.fetchall()
        conn.close()

        # Подготовим DataFrame со структурой
        columns = []
        types = []
        table_names = []
        for row in pragma_rows:
            # row по ключам: cid, name, type, notnull, dflt_value, pk
            columns.append(row['name'])
            types.append(row['type'])
            table_names.append(table_name)

        df = pd.DataFrame({
            'Таблица': table_names,
            'Столбец': columns,
            'Тип данных': types
        })

        # Очистка на всякий случай
        df = clean_data_for_excel(df)

        # Формируем Excel в памяти
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sheet_name = (f"{table_name}_struct")[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    if column[0].value:
                        max_length = len(str(column[0].value))
                    for cell in column[1:]:
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except (TypeError, AttributeError):
                            pass
                    adjusted_width = min(max(max_length + 2, 10), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                workbook = writer.book
                workbook.properties.title = f"Структура таблицы {table_name}"
                workbook.properties.creator = "M_Kran System"
                workbook.properties.created = datetime.now()
        except Exception as excel_error:
            # fallback в CSV
            print(f"Ошибка создания Excel (структура): {excel_error}")
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f"{table_name}_structure_{get_filename_timestamp()}.csv"
            )

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{table_name}_structure_{get_filename_timestamp()}.xlsx"
        )
    except Exception as e:
        try:
            conn.close()
        except:
            pass
        flash(f'Ошибка экспорта структуры: {e}', 'error')
        return redirect(url_for('database'))

@app.route('/export_filtered/<table_name>')
def export_filtered_table(table_name):
    """Экспорт отфильтрованных данных таблицы в Excel"""
    search_query = request.args.get('search', '')
    selected_columns = request.args.get('columns', '')
    
    # Дополнительные параметры фильтрации для специфичных таблиц
    drawing_filter = request.args.get('drawing', '')
    date_filter = request.args.get('date_filter', '')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к базе данных'}), 500
    
    try:
        cursor = conn.cursor()
        
        # Получаем все названия столбцов
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        all_columns = [col['name'] for col in cursor.fetchall()]
        
        # Определяем какие столбцы экспортировать
        if selected_columns and selected_columns.strip():
            selected_columns_list = selected_columns.split(',')
            # Фильтруем только существующие столбцы и убираем пустые значения
            columns_to_export = [col for col in selected_columns_list if col in all_columns and col.strip()]
            if not columns_to_export:
                columns_to_export = all_columns
        else:
            columns_to_export = all_columns
        
        # Специальная обработка для таблицы слов_клейм_факт
        if table_name == 'слов_клейм_факт':
            # Заменяем столбец ФИО на полное ФИО из таблицы ФИО_свар
            modified_columns = []
            for col in columns_to_export:
                if col == 'ФИО':
                    modified_columns.append('COALESCE(f.ФИО, s.ФИО) as ФИО')
                else:
                    modified_columns.append(f's.`{col}`')
            
            columns_sql = ', '.join(modified_columns)
        else:
            # Формируем SQL запрос с фильтрацией
            columns_sql = ', '.join([f'`{col}`' for col in columns_to_export])
        
        # Добавляем условия фильтрации
        where_conditions = []
        params = []
        
        # Поиск по тексту (регистронезависимый)
        if search_query and search_query.strip():
            search_conditions = []
            for col in columns_to_export:
                # Используем двойной поиск: оригинальный термин + нижний регистр
                search_conditions.append(f'(CAST(`{col}` AS TEXT) LIKE ? OR LOWER(CAST(`{col}` AS TEXT)) LIKE ?)')
                params.append(f'%{search_query}%')
                params.append(f'%{search_query.lower()}%')
            where_conditions.append(f"({' OR '.join(search_conditions)})")
        
        # Фильтры по отдельным столбцам
        column_filters = {}
        column_filter_types = {}
        
        # Собираем все фильтры и их типы
        for key, value in request.args.items():
            if key.startswith('filter_') and not key.startswith('filter_type_'):
                if value and value.strip():  # Проверяем, что значение не None и не пустое
                    column_name = key[7:]  # Убираем префикс 'filter_'
                    
                    # Получаем тип фильтра
                    filter_type_key = f'filter_type_{column_name}'
                    filter_type = request.args.get(filter_type_key, 'contains')
                    
                    # Добавляем фильтр только если это не специальное значение или если тип не empty/not_empty
                    if value.strip() != 'special' or filter_type in ['empty', 'not_empty']:
                        column_filters[column_name] = value.strip()
                        column_filter_types[column_name] = filter_type
        
        # Применяем фильтры с учетом их типов
        if column_filters:
            for column_name, filter_value in column_filters.items():
                if column_name in all_columns:  # Проверяем, что столбец существует
                    filter_type = column_filter_types.get(column_name, 'contains')
                    
                    # Специальная обработка для столбца ФИО в таблице слов_клейм_факт
                    if table_name == 'слов_клейм_факт' and column_name == 'ФИО':
                        if filter_type == 'contains':
                            # Ищем в полном ФИО из таблицы ФИО_свар
                            where_conditions.append(f'(CAST(f.ФИО AS TEXT) LIKE ? OR LOWER(CAST(f.ФИО AS TEXT)) LIKE ?)')
                            params.append(f'%{filter_value}%')
                            params.append(f'%{filter_value.lower()}%')
                        elif filter_type == 'not_contains':
                            # Исключение записей, содержащих значение в полном ФИО
                            where_conditions.append(f'(CAST(f.ФИО AS TEXT) NOT LIKE ? AND LOWER(CAST(f.ФИО AS TEXT)) NOT LIKE ?)')
                            params.append(f'%{filter_value}%')
                            params.append(f'%{filter_value.lower()}%')
                        elif filter_type == 'empty':
                            # Показывать только пустые значения ФИО
                            where_conditions.append(f'(f.ФИО IS NULL OR CAST(f.ФИО AS TEXT) = \'\' OR TRIM(CAST(f.ФИО AS TEXT)) = \'\' OR CAST(f.ФИО AS TEXT) = \'-\' OR CAST(f.ФИО AS TEXT) = \'None\')')
                        elif filter_type == 'not_empty':
                            # Показывать только непустые значения ФИО
                            where_conditions.append(f'(f.ФИО IS NOT NULL AND CAST(f.ФИО AS TEXT) != \'\' AND TRIM(CAST(f.ФИО AS TEXT)) != \'\' AND CAST(f.ФИО AS TEXT) != \'-\' AND CAST(f.ФИО AS TEXT) != \'None\')')
                        elif filter_type == 'equals':
                            # Точное совпадение с полным ФИО
                            where_conditions.append(f'CAST(f.ФИО AS TEXT) = ?')
                            params.append(filter_value)
                        elif filter_type == 'not_equals':
                            # Не равно полному ФИО
                            where_conditions.append(f'CAST(f.ФИО AS TEXT) != ?')
                            params.append(filter_value)
                    else:
                        # Обычная обработка для других столбцов
                        if filter_type == 'contains':
                            # Регистронезависимый поиск по конкретному столбцу
                            where_conditions.append(f'(CAST(`{column_name}` AS TEXT) LIKE ? OR LOWER(CAST(`{column_name}` AS TEXT)) LIKE ?)')
                            params.append(f'%{filter_value}%')
                            params.append(f'%{filter_value.lower()}%')
                        elif filter_type == 'not_contains':
                            # Исключение записей, содержащих значение
                            where_conditions.append(f'(CAST(`{column_name}` AS TEXT) NOT LIKE ? AND LOWER(CAST(`{column_name}` AS TEXT)) NOT LIKE ?)')
                            params.append(f'%{filter_value}%')
                            params.append(f'%{filter_value.lower()}%')
                        elif filter_type == 'empty':
                            # Показывать только пустые значения (NULL, пустая строка, только пробелы, "-", "None")
                            where_conditions.append(f'(`{column_name}` IS NULL OR CAST(`{column_name}` AS TEXT) = \'\' OR TRIM(CAST(`{column_name}` AS TEXT)) = \'\' OR CAST(`{column_name}` AS TEXT) = \'-\' OR CAST(`{column_name}` AS TEXT) = \'None\')')
                        elif filter_type == 'not_empty':
                            # Показывать только непустые значения (не NULL, не пустая строка, не только пробелы, не "-", не "None")
                            where_conditions.append(f'(`{column_name}` IS NOT NULL AND CAST(`{column_name}` AS TEXT) != \'\' AND TRIM(CAST(`{column_name}` AS TEXT)) != \'\' AND CAST(`{column_name}` AS TEXT) != \'-\' AND CAST(`{column_name}` AS TEXT) != \'None\')')
                        elif filter_type == 'equals':
                            # Точное совпадение
                            where_conditions.append(f'CAST(`{column_name}` AS TEXT) = ?')
                            params.append(filter_value)
                        elif filter_type == 'not_equals':
                            # Не равно
                            where_conditions.append(f'CAST(`{column_name}` AS TEXT) != ?')
                            params.append(filter_value)
                    
        
        # Фильтр по чертежу (для таблиц с полем "Чертеж")
        if drawing_filter and drawing_filter.strip() and 'Чертеж' in columns_to_export:
            where_conditions.append('`Чертеж` = ?')
            params.append(drawing_filter)
        
        # Фильтр по дате (для таблиц с полем "Дата сварки")
        if date_filter and date_filter.strip() and 'Дата сварки' in columns_to_export:
            if date_filter == 'today':
                where_conditions.append('DATE(`Дата сварки`) = DATE("now")')
            elif date_filter == 'week':
                where_conditions.append('DATE(`Дата сварки`) >= DATE("now", "-7 days")')
            elif date_filter == 'month':
                where_conditions.append('DATE(`Дата сварки`) >= DATE("now", "-1 month")')
        
        # Формируем WHERE clause
        where_clause = ""
        if where_conditions:
            where_clause = f"WHERE {' AND '.join(where_conditions)}"
        
        # Получаем отфильтрованные данные
        if table_name == 'слов_клейм_факт':
            # Для таблицы слов_клейм_факт используем JOIN
            data_query = f"""
                SELECT {columns_sql}
                FROM `{table_name}` s
                LEFT JOIN `ФИО_свар` f ON s.ФИО = f.id_fio
                {where_clause}
            """
        else:
            data_query = f"SELECT {columns_sql} FROM `{table_name}` {where_clause}"
        
        cursor.execute(data_query, params)
        records = cursor.fetchall()
        
        # Преобразуем записи в DataFrame с очисткой данных
        df = pd.DataFrame(records, columns=columns_to_export)
        
        # Очищаем данные от проблемных символов и значений
        df = clean_data_for_excel(df)
        
        conn.close()
        
        # Создаем Excel файл в памяти с улучшенными настройками
        output = io.BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Ограничиваем длину имени листа (Excel ограничение - 31 символ)
                sheet_name = table_name[:31] if len(table_name) > 31 else table_name
                
                # Записываем данные в Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Получаем рабочий лист для форматирования
                worksheet = writer.sheets[sheet_name]
                
                # Устанавливаем защиту от ошибок форматирования
                worksheet.protection.sheet = False
                
                # Автоматически подгоняем ширину столбцов с ограничениями
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Проверяем заголовок столбца
                    if column[0].value:
                        max_length = len(str(column[0].value))
                    
                    # Проверяем данные в столбце
                    for cell in column[1:]:  # Пропускаем заголовок
                        try:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except (TypeError, AttributeError):
                            pass
                    
                    # Устанавливаем ширину с разумными ограничениями
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Устанавливаем свойства документа
                workbook = writer.book
                workbook.properties.title = f"Экспорт таблицы {table_name}"
                workbook.properties.creator = "M_Kran System"
                workbook.properties.created = datetime.now()
                
        except Exception as excel_error:
            # Если произошла ошибка при создании Excel, создаем простой CSV
            print(f"Ошибка создания Excel: {excel_error}")
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            # Возвращаем CSV вместо Excel
            timestamp = get_filename_timestamp()
            filename = f"{table_name}_filtered_{timestamp}.csv"
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
        
        output.seek(0)
        
        # Формируем имя файла с информацией о фильтрах
        timestamp = get_filename_timestamp()
        filename = f"{table_name}_filtered_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Ошибка экспорта: {str(e)}'}), 500



@app.route('/api/table_info')
def get_all_tables():
    """API для получения списка всех таблиц"""
    # Логируем в файл
    try:
        with open('debug.log', 'a', encoding='utf-8') as f:
            f.write(f"\n=== ПОЛУЧЕНИЕ СПИСКА ТАБЛИЦ === {datetime.now()}\n")
    except:
        pass
    
    try:
        conn = get_db_connection()
        if not conn:
            try:
                with open('debug.log', 'a', encoding='utf-8') as f:
                    f.write("Ошибка подключения к БД\n")
            except:
                pass
            return jsonify({'error': 'Ошибка подключения к БД'})
        
        cursor = conn.cursor()
        
        # Получаем список всех таблиц
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = []
        
        for row in cursor.fetchall():
            table_name = row['name']
            
            # Получаем количество записей для каждой таблицы
            try:
                cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
                record_count = cursor.fetchone()['count']
            except Exception as e:
                print(f"Ошибка при подсчете записей в таблице {table_name}: {e}")
                record_count = 0
            
            tables.append({
                'name': table_name,
                'record_count': record_count
            })
        
        conn.close()
        
        print(f"Найдено таблиц: {len(tables)}")
        for table in tables:
            try:
                print(f"  - {table['name']}: {table['record_count']} записей")
            except UnicodeEncodeError:
                print(f"  - [table]: {table['record_count']} записей")
        
        return jsonify({
            'tables': tables
        })
        
    except Exception as e:
        print(f"Ошибка в get_all_tables: {e}")
        return jsonify({'error': str(e)})

@app.route('/api/table_info/<table_name>')
def table_info(table_name):
    """API для получения информации о таблице"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Получаем количество записей
        cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
        total_records = cursor.fetchone()['count']
        
        # Получаем информацию о столбцах
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        columns = [{'name': col['name'], 'type': col['type']} for col in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'table_name': table_name,
            'total_records': total_records,
            'columns': columns
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/extract_numbers', methods=['POST'])
def extract_numbers():
    """API для извлечения чисел из номера"""
    try:
        data = request.get_json()
        text = data.get('text', '')
        mode = data.get('mode', 'simple')  # simple, clean, combined
        
        if not text:
            return jsonify({'error': 'Текст не предоставлен'})
        
        # Импортируем локальные функции извлечения
        from extract_utils import extract_joint_number, clean_joint_number, extract_and_clean_joint_number
        
        result = {}
        
        if mode == 'simple':
            # Простое извлечение чисел
            extracted_number = extract_joint_number(text)
            result = {
                'success': True,
                'original_text': text,
                'extracted_number': extracted_number,
                'mode': 'simple'
            }
        elif mode == 'clean':
            # Очистка от префиксов S/F
            cleaned_number = clean_joint_number(text)
            result = {
                'success': True,
                'original_text': text,
                'cleaned_number': cleaned_number,
                'mode': 'clean'
            }
        elif mode == 'combined':
            # Комбинированный режим
            combined_number = extract_and_clean_joint_number(text)
            result = {
                'success': True,
                'original_text': text,
                'combined_number': combined_number,
                'mode': 'combined'
            }
        else:
            # Все режимы
            extracted_number = extract_joint_number(text)
            cleaned_number = clean_joint_number(text)
            combined_number = extract_and_clean_joint_number(text)
            result = {
                'success': True,
                'original_text': text,
                'extracted_number': extracted_number,
                'cleaned_number': cleaned_number,
                'combined_number': combined_number,
                'mode': 'all'
            }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/process_wl_china_numbers', methods=['POST'])
def process_wl_china_numbers():
    """API: извлечение номера стыка из «Номер_сварного_шва» в столбец «_Номер_сварного_шва_без_S_F_»."""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы wl_china
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='wl_china'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица wl_china не найдена'})
        
        # Целевой столбец для извлечённого номера шва (без S/F) — как в JOIN по журналу
        try:
            cursor.execute('ALTER TABLE wl_china ADD COLUMN "_Номер_сварного_шва_без_S_F_" TEXT')
            print('✅ Добавлен столбец "_Номер_сварного_шва_без_S_F_" в wl_china')
        except Exception as e:
            if "duplicate column name" in str(e).lower():
                print('ℹ️ Столбец "_Номер_сварного_шва_без_S_F_" уже есть в wl_china')
            else:
                print(f"⚠️ Ошибка при добавлении столбца: {e}")

        try:
            cursor.execute('ALTER TABLE wl_china DROP COLUMN "_Номер_сварного_шва"')
            print('🗑️ Удалён устаревший столбец "_Номер_сварного_шва"')
        except Exception:
            pass
        
        # Получаем данные для обработки
        cursor.execute('SELECT rowid, "Номер_сварного_шва" FROM wl_china WHERE "Номер_сварного_шва" IS NOT NULL')
        records = cursor.fetchall()
        
        if not records:
            return jsonify({'error': 'Нет данных для обработки в столбце Номер_сварного_шва'})
        
        # Импортируем функцию извлечения
        from extract_utils import extract_joint_number
        
        # Обрабатываем записи
        updated_count = 0
        for record in records:
            rowid, joint_text = record
            joint_number = extract_joint_number(joint_text)
            if joint_number is not None:
                cursor.execute(
                    'UPDATE wl_china SET "_Номер_сварного_шва_без_S_F_" = ? WHERE rowid = ?',
                    (joint_number, rowid),
                )
                updated_count += 1
        
        # Сохраняем изменения
        conn.commit()
        conn.close()
        
        # Логируем активность обработки данных
        log_activity('Обработка данных', f'Обработано записей в wl_china: {updated_count} из {len(records)}', 'cogs')
        
        return jsonify({
            'success': True,
            'message': f'Обработано записей: {updated_count} из {len(records)}',
            'total_records': len(records),
            'updated_records': updated_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/process_table_numbers', methods=['POST'])
def process_table_numbers():
    """API для обработки любой таблицы и извлечения чисел с удалением префиксов S/F"""
    try:
        data = request.get_json()
        table_name = data.get('table_name')
        source_column = data.get('source_column')
        target_column = data.get('target_column')
        mode = data.get('mode', 'combined')  # simple, clean, combined
        
        # Если target_column не указан, создаем автоматическое название
        if not target_column:
            target_column = f'_{source_column}_без_S_F_'
        
        if not all([table_name, source_column]):
            return jsonify({'error': 'Необходимо указать table_name и source_column'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            return jsonify({'error': f'Таблица {table_name} не найдена'})
        
        # Проверяем существование исходного столбца
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        columns = [row[1] for row in cursor.fetchall()]
        if source_column not in columns:
            return jsonify({'error': f'Столбец {source_column} не найден в таблице {table_name}'})
        
        # Добавляем целевой столбец, если его нет
        try:
            cursor.execute(f'ALTER TABLE `{table_name}` ADD COLUMN `{target_column}` TEXT')
        except Exception as e:
            if "duplicate column name" not in str(e).lower():
                return jsonify({'error': f'Ошибка при добавлении столбца: {e}'})
        
        # Получаем данные для обработки
        cursor.execute(f'SELECT rowid, `{source_column}` FROM `{table_name}` WHERE `{source_column}` IS NOT NULL')
        records = cursor.fetchall()
        
        if not records:
            return jsonify({'error': f'Нет данных для обработки в столбце {source_column}'})
        
        # Импортируем функции извлечения
        from extract_utils import extract_joint_number, clean_joint_number, extract_and_clean_joint_number
        
        # Выбираем функцию в зависимости от режима
        if mode == 'simple':
            process_function = extract_joint_number
        elif mode == 'clean':
            process_function = clean_joint_number
        else:  # combined
            process_function = extract_and_clean_joint_number
        
        # Обрабатываем записи
        updated_count = 0
        
        for record in records:
            rowid, source_text = record
            
            try:
                # Безопасное преобразование в строку
                if source_text is not None:
                    try:
                        source_text_str = str(source_text)
                    except UnicodeEncodeError:
                        source_text_str = str(source_text).encode('utf-8', errors='ignore').decode('utf-8')
                else:
                    source_text_str = ""
                
                processed_number = process_function(source_text_str)
                if processed_number is not None:
                    cursor.execute(f'UPDATE `{table_name}` SET `{target_column}` = ? WHERE rowid = ?', (processed_number, rowid))
                    updated_count += 1
            except Exception as e:
                # Игнорируем ошибки и продолжаем обработку
                continue
        
        # Сохраняем изменения
        conn.commit()
        conn.close()
        
        message = f'Обработано записей: {updated_count} из {len(records)}'
        
        # Логируем активность обработки данных
        log_activity('Обработка таблицы', f'Обработана таблица {table_name}: {updated_count} из {len(records)} записей', 'table')
        
        return jsonify({
            'success': True,
            'message': message,
            'table_name': table_name,
            'source_column': source_column,
            'target_column': target_column,
            'mode': mode,
            'total_records': len(records),
            'updated_records': updated_count
        })
        
    except Exception as e:
        return jsonify({'error': 'Произошла ошибка при обработке данных'})

@app.route('/download/<filename>')
def download_file(filename):
    """Скачивание загруженного файла"""
    try:
        return send_file(
            os.path.join(app.config['UPLOAD_FOLDER'], filename),
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Ошибка скачивания файла: {e}', 'error')
        return redirect(url_for('upload'))

@app.route('/delete_file/<filename>', methods=['POST'])
def delete_file(filename):
    """Удаление загруженного файла"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            
            # Логируем активность удаления файла
            log_activity('Удаление файла', f'Удален файл: {filename}', 'trash')
            
            return jsonify({'success': True, 'message': 'Файл успешно удален'})
        else:
            return jsonify({'success': False, 'message': 'Файл не найден'})
    except Exception as e:
        # Логируем ошибку удаления
        log_activity('Ошибка удаления', f'Ошибка при удалении файла: {filename}', 'exclamation-triangle')
        return jsonify({'success': False, 'message': f'Ошибка удаления: {e}'})

@app.route('/joint_manager')
def joint_manager():
    """Страница управления стыками"""
    return render_template('joint_manager.html')

@app.route('/api/joint_manager/export', methods=['GET', 'POST'])
def api_joint_manager_export():
    """Экспорт стыков в Excel.

    Режимы:
    - GET или POST без JSON: экспорт всех записей из pipeline_weld_joint_iso
    - POST с JSON { scope: 'filtered', rows: [...] }: экспорт только переданных строк
    Структура rows элементов (от клиента):
      {
        id: str,               # опционально
        joint_number: str,     # стык
        weld_type: str,        # тип шва (краткий код)
        sheet: str|int,        # лист
        deletion_code: str,    # код удаления
        comment: str,          # комментарий
        ISO: str,              # обязательный для экспорта
        Линия: str             # обязательный для экспорта
      }
    """
    try:
        # Попытка понять, не пришла ли выборка отфильтрованных строк
        request_json = None
        try:
            request_json = request.get_json(silent=True)
        except Exception:
            request_json = None

        if request.method == 'POST' and request_json and request_json.get('scope') == 'filtered':
            client_rows = request_json.get('rows') or []
            if not client_rows:
                return jsonify({'success': False, 'message': 'Нет переданных строк для экспорта'}), 400

            # Преобразуем клиентские поля к экспортируемым колонкам
            export_rows = []
            for r in client_rows:
                export_rows.append({
                    'id': r.get('id', ''),
                    'Титул': '',
                    'ISO': r.get('ISO', ''),
                    'Линия': r.get('Линия', ''),
                    'ключь_жср_смр': '',
                    'Линия2': '',
                    'стык': r.get('joint_number', ''),
                    'Тип_соединения_российский_стандарт': r.get('weld_type', ''),
                    'Код_удаления': r.get('deletion_code', ''),
                    'лист': r.get('sheet', ''),
                    'Комментарий': r.get('comment', ''),
                })

            df = pd.DataFrame(export_rows)
        else:
            # Экспорт всех записей из БД
            conn = get_db_connection()
            if not conn:
                return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
            cursor = conn.cursor()

            # Проверяем существование таблицы
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
            if not cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': 'Таблица pipeline_weld_joint_iso не найдена'})

            # Простая серверная фильтрация по параметрам запроса (если заданы)
            drawing = request.args.get('drawing')
            line = request.args.get('line')
            where_clauses = []
            params = []
            if drawing:
                where_clauses.append('"ISO" = ?')
                params.append(drawing)
            if line:
                where_clauses.append('"Линия" = ?')
                params.append(line)

            base_sql = '''
                SELECT id, "Титул", "ISO", "Линия", "ключь_жср_смр", "Линия2", "стык",
                       "Тип_соединения_российский_стандарт", "Код_удаления", "лист", "Комментарий"
                FROM pipeline_weld_joint_iso
            '''
            if where_clauses:
                base_sql += ' WHERE ' + ' AND '.join(where_clauses)
            base_sql += ' ORDER BY "ISO", "Линия", "стык"'

            cursor.execute(base_sql, params)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            conn.close()

            if not rows:
                return jsonify({'success': False, 'message': 'Нет данных для экспорта'})

            df = pd.DataFrame(rows, columns=columns)

        # Сформируем в Excel отдельную колонку "Тип шва" из БД-колонки "Тип_соединения_российский_стандарт"
        # с принудительной нормализацией, а исходную колонку удалим из экспорта, чтобы не было дублей.
        def _normalize_weld_type(val):
            if val is None:
                return ''
            s = str(val).strip().strip('"').strip("'")
            # Нормализуем латинские буквы к кириллице (визуально похожие)
            latin_to_cyr = {'C': 'С', 'T': 'Т', 'H': 'Н', 'Y': 'У'}
            if len(s) == 1 and s.upper() in latin_to_cyr:
                return latin_to_cyr[s.upper()]
            # Уже корректный код
            if s in ('С', 'Т', 'Н', 'У'):
                return s
            # Убираем служебные значения и варианты с регистрами/пробелами
            s_lower = s.lower().replace('_', ' ').replace('\u00a0', ' ').strip()
            if s_lower in ('тип шва', 'тип_шва', 'типшва') or 'тип' in s_lower:
                return ''
            # Первая буква как код (с учетом латиницы)
            if s:
                ch = s[0].upper()
                if ch in latin_to_cyr:
                    return latin_to_cyr[ch]
                if ch in ('С', 'Т', 'Н', 'У'):
                    return ch
            return ''

        if 'Тип_соединения_российский_стандарт' in df.columns:
            df['Тип шва'] = df['Тип_соединения_российский_стандарт'].apply(_normalize_weld_type)
            df = df.drop(columns=['Тип_соединения_российский_стандарт'])
        elif 'Тип шва' in df.columns:
            df['Тип шва'] = df['Тип шва'].apply(_normalize_weld_type)


        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Joints', index=False)
            ws = writer.sheets['Joints']
            # Автоширина колонок (ограничим)
            for column in ws.columns:
                max_length = 0
                letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[letter].width = min(max_length + 2, 50)

            # Добавим выпадающий список для колонки "Тип шва" (значения: С, Т, Н)
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                # Найдём номер колонки по заголовку
                header_to_col = {cell.value: cell.column for cell in ws[1]}
                weld_col_idx = header_to_col.get('Тип шва')
                if weld_col_idx:
                    # Преобразуем индекс в букву
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(weld_col_idx)
                    # Диапазон со 2-й строки до разумного предела (например, 100000)
                    data_range = f"{col_letter}2:{col_letter}100000"
                    dv = DataValidation(type="list", formula1='"С,Т,Н,У"', allow_blank=True, showDropDown=True)
                    dv.error = 'Пожалуйста, выберите одно из значений: С, Т, Н или У'
                    dv.errorTitle = 'Неверный тип шва'
                    ws.add_data_validation(dv)
                    dv.add(data_range)
            except Exception as _e:
                # Если не получилось добавить валидацию, просто пропускаем
                pass

        output.seek(0)
        filename = f'joints_{get_filename_timestamp()}.xlsx'
        try:
            count_exported = len(df)
        except Exception:
            count_exported = 0
        log_activity('Экспорт данных', f'Экспортировано {count_exported} стыков', 'download')
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f'Ошибка экспорта стыков: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/joint_manager/export_current', methods=['POST'])
def api_joint_manager_export_current():
    """Экспорт по текущим фильтрам шапки (Номер чертежа/Номер линии)."""
    try:
        payload = request.get_json(silent=True) or {}
        drawing = (payload.get('drawing') or '').strip()
        line = (payload.get('line') or '').strip()

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        cursor = conn.cursor()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Таблица pipeline_weld_joint_iso не найдена'})

        base_sql = '''
            SELECT id, "Титул", "ISO", "Линия", "ключь_жср_смр", "Линия2", "стык",
                   "Тип_соединения_российский_стандарт", "Код_удаления", "лист", "Комментарий"
            FROM pipeline_weld_joint_iso
        '''
        params = []
        where = []
        if drawing:
            where.append('"ISO" = ?')
            params.append(drawing)
        if line:
            where.append('"Линия" = ?')
            params.append(line)
        if where:
            base_sql += ' WHERE ' + ' AND '.join(where)
        base_sql += ' ORDER BY "ISO", "Линия", "стык"'

        cursor.execute(base_sql, params)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()

        if not rows:
            return jsonify({'success': False, 'message': 'Нет данных для экспорта по заданным фильтрам'})

        df = pd.DataFrame(rows, columns=columns)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Joints', index=False)
        output.seek(0)
        filename = f'joints_filtered_{get_filename_timestamp()}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f'Ошибка экспорта (текущий фильтр): {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/joint_manager/import', methods=['POST'])
def api_joint_manager_import():
    """Импорт Excel/CSV для массового обновления pipeline_weld_joint_iso.
    Матч по id, либо по паре (ISO, стык), если id нет.
    Обновляются только переданные поля.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Файл не передан'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Файл не выбран'}), 400

        filename_lower = file.filename.lower()
        if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls') or filename_lower.endswith('.csv')):
            return jsonify({'success': False, 'message': 'Поддерживаются только .xlsx, .xls, .csv'}), 400

        stream = io.BytesIO(file.read())
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(stream)
        else:
            df = pd.read_excel(stream)
        if df.empty:
            return jsonify({'success': False, 'message': 'Файл пустой'}), 400

        # Нормализация названий столбцов
        df.columns = [str(c).strip() for c in df.columns]

        # Возможные ключи
        has_id = any(c.lower() == 'id' for c in df.columns)
        has_iso = any(c.lower() in ['iso', 'чертеж', 'чертеж_iso'] for c in df.columns)
        has_joint = any(c.strip().lower() in ['стык', 'joint', 'joint_number', '№ стыка', 'номер стыка'] for c in df.columns)

        if not has_id and not (has_iso and has_joint):
            return jsonify({'success': False, 'message': 'Нет ключей для сопоставления (id или ISO+стык)'}), 400

        # Разрешенные к обновлению поля (если есть в файле)
        updatable_candidates = [
            'Титул', 'ISO', 'Линия', 'ключь_жср_смр', 'Линия2', 'стык',
            'Тип_шва', 'Тип шва', 'Тип_соединения_российский_стандарт', 'Код_удаления', 'лист', 'Комментарий'
        ]
        present_fields = [c for c in updatable_candidates if c in df.columns]

        # Если пользователь прислал "Тип шва" (читаемая форма), конвертируем к базе "Тип_шва"
        # Если пользователь прислал "Тип шва" (удобочитаемо) — конвертируем в оба поля БД,
        # приоритет отдаем "Тип_соединения_российский_стандарт" (как вы используете в БД)
        if 'Тип шва' in present_fields and 'Тип_соединения_российский_стандарт' not in present_fields:
            # Приводим к корректным кодам перед записью в БД
            def _norm_import(val):
                if val is None:
                    return ''
                s = str(val).strip().strip('"').strip("'")
                latin_to_cyr = {'C': 'С', 'T': 'Т', 'H': 'Н', 'Y': 'У'}
                if len(s) == 1 and s.upper() in latin_to_cyr:
                    return latin_to_cyr[s.upper()]
                if s in ('С', 'Т', 'Н', 'У'):
                    return s
                s_lower = s.lower().replace('_', ' ').replace('\u00a0', ' ').strip()
                if s_lower in ('тип шва', 'тип_шва', 'типшва') or 'тип' in s_lower:
                    return ''
                if s:
                    ch = s[0].upper()
                    if ch in latin_to_cyr:
                        return latin_to_cyr[ch]
                    if ch in ('С', 'Т', 'Н', 'У'):
                        return ch
                return ''
            normalized_series = df['Тип шва'].apply(_norm_import)
            df['Тип_соединения_российский_стандарт'] = normalized_series
            # Для совместимости можно продублировать и в Тип_шва, если столбец существует в таблице
            df['Тип_шва'] = normalized_series
            # Меняем имя поля в списке обновляемых на то, что в БД
            present_fields = [c if c != 'Тип шва' else 'Тип_соединения_российский_стандарт' for c in present_fields]
        if not present_fields:
            return jsonify({'success': False, 'message': 'В файле нет столбцов для обновления'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'}), 500
        cursor = conn.cursor()

        processed = 0
        updated = 0
        created = 0

        # Подготовим индексы для ускорения сопоставления по ISO+стык при необходимости
        if not has_id and has_iso and has_joint:
            cursor.execute('SELECT id, "ISO", "стык" FROM pipeline_weld_joint_iso')
            existing = cursor.fetchall()
            # Ключ: (ISO, стык)
            iso_joint_to_id = {}
            for row in existing:
                row_id, iso_val, joint_val = row
                iso_joint_to_id[(str(iso_val).strip(), str(joint_val).strip())] = row_id

        for _, r in df.iterrows():
            processed += 1
            record_id = None

            if has_id:
                # Найдем имя столбца id в любом регистре
                id_col = next((c for c in df.columns if c.lower() == 'id'), None)
                val = r.get(id_col)
                if pd.isna(val) or val == '' or str(val).strip() == '':
                    record_id = None
                else:
                    try:
                        record_id = int(val)
                    except Exception:
                        record_id = None
            if record_id is None and has_iso and has_joint:
                iso_col = next((c for c in df.columns if c.lower() in ['iso', 'чертеж', 'чертеж_iso']), None)
                joint_col = next((c for c in df.columns if c.strip().lower() in ['стык', 'joint', 'joint_number', '№ стыка', 'номер стыка']), None)
                iso_val = str(r.get(iso_col) or '').strip()
                joint_val = str(r.get(joint_col) or '').strip()
                if iso_val and joint_val:
                    if has_id:
                        # если и id был в файле, но невалидный — пробуем найти по ISO+стык на лету
                        cursor.execute('SELECT id FROM pipeline_weld_joint_iso WHERE "ISO" = ? AND "стык" = ?', (iso_val, joint_val))
                        found = cursor.fetchone()
                        if found:
                            record_id = int(found[0])
                    else:
                        record_id = iso_joint_to_id.get((iso_val, joint_val))
                
                # Если запись не найдена, создаем новую
                if not record_id:
                    # Создаем новую запись
                    if iso_val and joint_val:
                        # Создаем новую запись
                        insert_fields = ['ISO', 'стык']
                        insert_values = [iso_val, joint_val]
                        
                        # Добавляем остальные поля если они есть
                        for field in present_fields:
                            if field not in ['ISO', 'стык']:
                                value = r.get(field)
                                if isinstance(value, float) and pd.isna(value):
                                    value = ''
                                insert_fields.append(field)
                                insert_values.append(value)
                        
                        # Добавляем обязательные поля с значениями по умолчанию
                        insert_fields.extend(['Титул', 'Линия', 'ключь_жср_смр', 'Линия2', 'Код_удаления', 'лист', 'повтор', 'открыть', 'Дата_загрузки', 'Комментарий'])
                        insert_values.extend(['', '', '', '', '', '', 0, 0, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''])
                        
                        # Создаем SQL для вставки
                        placeholders = ', '.join(['?' for _ in insert_values])
                        field_names = ', '.join([f'"{f}"' for f in insert_fields])
                        insert_sql = f'INSERT INTO pipeline_weld_joint_iso ({field_names}) VALUES ({placeholders})'
                        
                        cursor.execute(insert_sql, tuple(insert_values))
                        created += 1
                        continue

            # Проверим, что запись существует
            cursor.execute('SELECT COUNT(*) FROM pipeline_weld_joint_iso WHERE id = ?', (record_id,))
            if cursor.fetchone()[0] == 0:
                continue

            set_parts = []
            values = []
            for field in present_fields:
                value = r.get(field)
                if isinstance(value, float) and pd.isna(value):
                    value = ''
                set_parts.append(f'"{field}" = ?')
                values.append(value)

            if set_parts:
                values.append(record_id)
                sql = 'UPDATE pipeline_weld_joint_iso SET ' + ', '.join(set_parts) + ' WHERE id = ?'
                cursor.execute(sql, tuple(values))
                updated += 1

        conn.commit()
        conn.close()

        log_activity('Импорт данных', f'Импорт стыков: обработано {processed}, обновлено {updated}, создано {created}', 'upload')
        return jsonify({'success': True, 'processed': int(processed), 'updated': int(updated), 'created': int(created)})
    except Exception as e:
        logger.error(f'Ошибка импорта стыков: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================
# Резервные копии: список и восстановление
# ==========================

@app.route('/backups')
def backups_page():
    """Страница: список архивов в папке backups/ и запуск восстановления"""
    try:
        backups = []
        backups_dir = os.path.join(os.getcwd(), 'backups')
        if os.path.exists(backups_dir):
            for file in os.listdir(backups_dir):
                if file.lower().endswith('.zip'):
                    full_path = os.path.join(backups_dir, file)
                    meta_path = os.path.splitext(full_path)[0] + '.json'
                    metadata = {}
                    if os.path.exists(meta_path):
                        try:
                            with open(meta_path, 'r', encoding='utf-8') as f:
                                metadata = json.load(f)
                        except Exception:
                            pass
                    backups.append({
                        'file': file,
                        'size': os.path.getsize(full_path),
                        'metadata': metadata
                    })
        backups.sort(key=lambda x: x['file'], reverse=True)
        return render_template('restore_backups.html', backups=backups)
    except Exception as e:
        return f"Ошибка загрузки резервных копий: {e}", 500


def _extract_zip_to_target(zip_path: str, target_dir: str, overwrite: bool = True) -> dict:
    import shutil
    os.makedirs(target_dir, exist_ok=True)
    extracted = 0
    skipped = 0
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for member in zf.infolist():
            name = member.filename
            if name.endswith('/'):
                continue
            dst_path = os.path.join(target_dir, name)
            dst_dir = os.path.dirname(dst_path)
            os.makedirs(dst_dir, exist_ok=True)
            if os.path.exists(dst_path) and not overwrite:
                skipped += 1
                continue
            with zf.open(member) as src, open(dst_path, 'wb') as dst:
                shutil.copyfileobj(src, dst)
            extracted += 1
    return {'extracted': extracted, 'skipped': skipped}


@app.route('/api/backups/list')
def api_backups_list():
    try:
        backups = []
        backups_dir = os.path.join(os.getcwd(), 'backups')
        if os.path.exists(backups_dir):
            for file in os.listdir(backups_dir):
                if file.lower().endswith('.zip'):
                    full_path = os.path.join(backups_dir, file)
                    meta_path = os.path.splitext(full_path)[0] + '.json'
                    metadata = {}
                    if os.path.exists(meta_path):
                        try:
                            with open(meta_path, 'r', encoding='utf-8') as f:
                                metadata = json.load(f)
                        except Exception:
                            pass
                    backups.append({
                        'file': file,
                        'size': os.path.getsize(full_path),
                        'metadata': metadata
                    })
        backups.sort(key=lambda x: x['file'], reverse=True)
        return jsonify({'success': True, 'backups': backups})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/backups/restore', methods=['POST'])
def api_backups_restore():
    try:
        data = request.get_json() or {}
        filename = data.get('file')
        overwrite = bool(data.get('overwrite', True))
        target = data.get('target') or '.'

        if not filename or not filename.lower().endswith('.zip'):
            return jsonify({'success': False, 'message': 'Некорректное имя файла'}), 400

        backups_dir = os.path.join(os.getcwd(), 'backups')
        zip_path = os.path.join(backups_dir, filename)
        if not os.path.exists(zip_path):
            return jsonify({'success': False, 'message': 'Архив не найден'}), 404

        # Цель: корень проекта (cwd), если target пустой или '.'
        if target in ['', '.', './']:
            target_dir = os.getcwd()
        else:
            target_dir = os.path.join(os.getcwd(), target)
        stats = _extract_zip_to_target(zip_path, target_dir, overwrite=overwrite)
        log_activity('Восстановление', f'Архив {filename} → {target_dir} (extracted={stats["extracted"]}, skipped={stats["skipped"]})', 'upload')
        return jsonify({'success': True, 'target': target_dir, 'stats': stats})
    except Exception as e:
        logger.error(f'Ошибка восстановления из backups: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/debug_table_structure')
def debug_table_structure():
    """API для отладки структуры таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Получаем структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        
        # Получаем количество записей
        cursor.execute("SELECT COUNT(*) FROM Log_Piping_PTO")
        total_records = cursor.fetchone()[0]
        
        # Получаем первые 3 записи для примера
        cursor.execute("SELECT * FROM Log_Piping_PTO LIMIT 3")
        sample_rows = cursor.fetchall()
        
        # Преобразуем Row объекты в обычные списки
        sample_data = []
        for row in sample_rows:
            sample_data.append(list(row))
        
        conn.close()
        
        return jsonify({
            'table_exists': True,
            'total_records': total_records,
            'columns': [{'name': col[1], 'type': col[2]} for col in columns_info],
            'sample_data': sample_data
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_tables')
def get_tables():
    """API для получения списка всех таблиц в базе данных"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'tables': tables})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_columns', methods=['POST'])
def get_columns():
    """API для получения списка столбцов таблицы"""
    try:
        data = request.get_json()
        table_name = data.get('table_name')
        
        if not table_name:
            return jsonify({'error': 'Не указано имя таблицы'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            return jsonify({'error': f'Таблица {table_name} не найдена'})
        
        # Получаем список столбцов
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        columns = [row[1] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'columns': columns})
    except Exception as e:
        return jsonify({'error': str(e)})





@app.route('/api/get_drawings')
def get_drawings():
    """API для получения списка чертежей (ISO) из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбец с ISO - возможные варианты названий
        iso_column = None
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
                break
        
        if not iso_column:
            return jsonify({'error': f'Столбец с ISO не найден. Доступные столбцы: {column_names}'})
        
        # Получаем уникальные ISO (чертежи)
        cursor.execute(f"SELECT DISTINCT `{iso_column}` FROM `Log_Piping_PTO` WHERE `{iso_column}` IS NOT NULL AND `{iso_column}` != '' ORDER BY `{iso_column}`")
        drawings = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({'drawings': drawings})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_lines')
def get_lines():
    """API для получения списка линий из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбец с линиями - используем ТОЛЬКО 'Линия'
        line_column = None
        for col in column_names:
            if col == 'Линия':
                line_column = col
                break
        
        if not line_column:
            return jsonify({'error': f'Столбец с линиями не найден. Доступные столбцы: {column_names}'})
        
        # Получаем уникальные линии
        cursor.execute(f"SELECT DISTINCT `{line_column}` FROM `Log_Piping_PTO` WHERE `{line_column}` IS NOT NULL AND `{line_column}` != '' ORDER BY `{line_column}`")
        lines = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({'lines': lines})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_lines_by_drawing', methods=['POST'])
def get_lines_by_drawing():
    """API для получения линий по выбранному чертежу (ISO) из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        iso = data.get('iso')
        logger.info(f"Получен запрос на линии для ISO: {iso}")
        
        cursor = conn.cursor()
        
        # Получаем структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        logger.info(f"Доступные столбцы в таблице: {column_names}")
        
        # Ищем столбцы ISO и Линия (используем ТОЛЬКО 'Линия', игнорируем 'Линия2')
        iso_column = None
        line_column = None
        
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
            # Используем ТОЛЬКО столбец 'Линия'
            if col == 'Линия':
                line_column = col
        

        logger.info(f"Найденные столбцы - ISO: {iso_column}, Линия: {line_column}")
        
        if not iso_column or not line_column:
            error_msg = f'Необходимые столбцы не найдены. Доступные: {column_names}'
            logger.error(error_msg)
            return jsonify({'error': error_msg})
        
        if iso:
            # Получаем линии для конкретного ISO
            query = f"SELECT DISTINCT `{line_column}` FROM `Log_Piping_PTO` WHERE `{iso_column}` = ? AND `{line_column}` IS NOT NULL AND `{line_column}` != '' ORDER BY `{line_column}`"
            logger.info(f"Выполняем запрос: {query} с параметром: {iso}")
            cursor.execute(query, (iso,))
        else:
            # Получаем все линии
            query = f"SELECT DISTINCT `{line_column}` FROM `Log_Piping_PTO` WHERE `{line_column}` IS NOT NULL AND `{line_column}` != '' ORDER BY `{line_column}`"
            logger.info(f"Выполняем запрос для всех линий: {query}")
            cursor.execute(query)
        
        lines = [row[0] for row in cursor.fetchall()]
        logger.info(f"Найдено линий: {len(lines)}")
        logger.info(f"Первые 5 линий: {lines[:5]}")
        
        conn.close()
        
        return jsonify({'lines': lines})
    except Exception as e:
        error_msg = f"Ошибка в get_lines_by_drawing: {str(e)}"
        logger.error(error_msg)
        return jsonify({'error': error_msg})

@app.route('/api/get_drawings_by_line', methods=['POST'])
def get_drawings_by_line():
    """API для получения чертежей (ISO) по выбранной линии из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        line = data.get('line')

        logger.info(f"Получен запрос на чертежи для линии: {line}")
        
        cursor = conn.cursor()
        
        # Получаем структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        logger.info(f"Доступные столбцы в таблице: {column_names}")
        
        # Ищем столбцы ISO и Линия (используем ТОЛЬКО 'Линия', игнорируем 'Линия2')
        iso_column = None
        line_column = None
        
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
            # Используем ТОЛЬКО столбец 'Линия'
            if col == 'Линия':
                line_column = col
        

        logger.info(f"Найденные столбцы - ISO: {iso_column}, Линия: {line_column}")
        
        if not iso_column or not line_column:
            error_msg = f'Необходимые столбцы не найдены. Доступные: {column_names}'
            logger.error(error_msg)
            return jsonify({'error': error_msg})
        
        if line:
            # Получаем ISO для конкретной линии
            query = f"SELECT DISTINCT `{iso_column}` FROM `Log_Piping_PTO` WHERE `{line_column}` = ? AND `{iso_column}` IS NOT NULL AND `{iso_column}` != '' ORDER BY `{iso_column}`"
            logger.info(f"Выполняем запрос: {query} с параметром: {line}")
            cursor.execute(query, (line,))
        else:
            # Получаем все ISO
            query = f"SELECT DISTINCT `{iso_column}` FROM `Log_Piping_PTO` WHERE `{iso_column}` IS NOT NULL AND `{iso_column}` != '' ORDER BY `{iso_column}`"
            logger.info(f"Выполняем запрос для всех чертежей: {query}")
            cursor.execute(query)
        
        drawings = [row[0] for row in cursor.fetchall()]

        logger.info(f"Найдено чертежей: {len(drawings)}")
        logger.info(f"Первые 5 чертежей: {drawings[:5]}")
        
        conn.close()
        
        return jsonify({'drawings': drawings})
    except Exception as e:
        error_msg = f"Ошибка в get_drawings_by_line: {str(e)}"
        logger.error(error_msg)
        return jsonify({'error': error_msg})

@app.route('/api/get_weld_types')
def get_weld_types():
    """API для получения типов швов из таблицы type_weld, столбец 'Русское обозначение'"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы type_weld
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='type_weld'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица type_weld не найдена'})
        
        # Получаем структуру таблицы
        cursor.execute("PRAGMA table_info(type_weld)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        # Проверяем наличие столбца 'Русское обозначение'
        if 'Русское обозначение' not in column_names:
            return jsonify({'error': f'Столбец "Русское обозначение" не найден. Доступные столбцы: {column_names}'})
        
        # Ищем столбец с описанием
        description_column = None
        for col in column_names:
            if 'тип шва руское' in col.lower() or 'описание' in col.lower() or 'наименование' in col.lower() or 'название' in col.lower():
                description_column = col
                break
        

        
        # Получаем типы швов с описанием
        if description_column:
            cursor.execute(f"SELECT DISTINCT `Русское обозначение`, `{description_column}` FROM `type_weld` WHERE `Русское обозначение` IS NOT NULL AND `Русское обозначение` != '' ORDER BY `Русское обозначение`")
            weld_types = [{'value': row[0], 'description': row[1] or ''} for row in cursor.fetchall()]
        else:
            # Если нет столбца с описанием, возвращаем только значения
            cursor.execute("SELECT DISTINCT `Русское обозначение` FROM `type_weld` WHERE `Русское обозначение` IS NOT NULL AND `Русское обозначение` != '' ORDER BY `Русское обозначение`")
            weld_types = [{'value': row[0], 'description': ''} for row in cursor.fetchall()]
        

        
        conn.close()
        
        return jsonify({'weld_types': weld_types})
    except Exception as e:

        return jsonify({'error': str(e)})


@app.route('/api/get_deletion_codes')
def get_deletion_codes():
    """API для получения кодов удаления из таблицы код_для_удаления, столбец 'Код'"""
    print("DEBUG: API get_deletion_codes вызван")
    conn = get_db_connection()
    if not conn:
        print("DEBUG: Ошибка подключения к БД")
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы код_для_удаления
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='код_для_удаления'")
        if not cursor.fetchone():
            print("DEBUG: Таблица код_для_удаления не найдена")
            return jsonify({'error': 'Таблица код_для_удаления не найдена'})
        
        # Получаем структуру таблицы
        cursor.execute("PRAGMA table_info(код_для_удаления)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        print(f"DEBUG: Столбцы таблицы: {column_names}")
        
        # Проверяем наличие столбца 'Код'
        if 'Код' not in column_names:
            print(f"DEBUG: Столбец 'Код' не найден. Доступные столбцы: {column_names}")
            return jsonify({'error': f'Столбец "Код" не найден. Доступные столбцы: {column_names}'})
        
        # Получаем коды удаления с описанием и расшифровкой
        cursor.execute("SELECT DISTINCT `Код`, `Расшифровка`, `Пример_применения` FROM `код_для_удаления` WHERE `Код` IS NOT NULL AND `Код` != '' ORDER BY `Код`")
        codes = cursor.fetchall()
        print(f"DEBUG: Найдено кодов: {len(codes)}")
        print(f"DEBUG: Коды с расшифровкой и описанием: {[(row[0], row[1], row[2]) for row in codes]}")
        
        deletion_codes = []
        for row in codes:
            code, rasshifrovka, primen_prim = row
            # Формируем описание: если есть пример применения, используем его, иначе расшифровка
            description = primen_prim if primen_prim else (rasshifrovka if rasshifrovka else code)
            deletion_codes.append({
                'value': code, 
                'description': description,
                'rasshifrovka': rasshifrovka,
                'primen_prim': primen_prim
            })
        
        conn.close()
        
        print(f"DEBUG: Возвращаем {len(deletion_codes)} кодов удаления")
        return jsonify({'deletion_codes': deletion_codes})
    except Exception as e:
        print(f"DEBUG: Ошибка в API get_deletion_codes: {e}")
        return jsonify({'error': str(e)})


@app.route('/api/get_joints_data', methods=['POST'])
def get_joints_data():
    """API для получения данных о стыках из таблицы pipeline_weld_joint_iso"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        drawing_number = data.get('drawing_number', '').strip()  # ISO
        line_number = data.get('line_number', '').strip()  # Линия
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица pipeline_weld_joint_iso не найдена'})
        
        # Проверяем существование столбца Тип_соединения_российский_стандарт
        cursor.execute("PRAGMA table_info(pipeline_weld_joint_iso)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Определяем правильное название столбца для типа соединения
        weld_type_column = None
        for col in column_names:
            if 'тип' in col.lower() and ('соединен' in col.lower() or 'российск' in col.lower() or 'стандарт' in col.lower()):
                weld_type_column = col
                break
        
        # Получаем информацию о файле распуловки из Log_Piping_PTO
        raspulovka_file = None
        if drawing_number and line_number:
            try:
                cursor.execute("""
                    SELECT "распуловка" 
                    FROM Log_Piping_PTO 
                    WHERE "ISO" = ? AND "Линия" = ? 
                    LIMIT 1
                """, (drawing_number, line_number))
                raspulovka_result = cursor.fetchone()
                if raspulovka_result and raspulovka_result[0]:
                    raspulovka_file = raspulovka_result[0]
            except Exception as e:
                # Если таблица Log_Piping_PTO не существует или нет столбца распуловка, игнорируем ошибку
                pass
        
        # Базовый запрос
        if weld_type_column:
            query = f"""
            SELECT id, "Титул", "ISO", "Линия", "ключь_жср_смр", "Линия2", "стык", 
                   "Код_удаления", "лист", "повтор", "открыть", "Дата_загрузки", "{weld_type_column}", "Комментарий"
            FROM pipeline_weld_joint_iso 
            WHERE 1=1
            """
        else:
            query = """
            SELECT id, "Титул", "ISO", "Линия", "ключь_жср_смр", "Линия2", "стык", 
                   "Код_удаления", "лист", "повтор", "открыть", "Дата_загрузки", "Комментарий"
            FROM pipeline_weld_joint_iso 
            WHERE 1=1
            """
        
        params = []
        
        # Добавляем фильтры
        if drawing_number:
            query += ' AND "ISO" = ?'
            params.append(drawing_number)
            
        if line_number:
            query += ' AND "Линия" = ?'
            params.append(line_number)
            
        query += ' ORDER BY "ISO", "Линия", "стык"'
        
        cursor.execute(query, params)
        joints = []
        for row in cursor.fetchall():
            joint_data = {
                'id': row[0],
                'drawing_number': row[2] or '',  # ISO
                'line_number': row[3] or '',     # Линия
                'joint_number': row[6] or '',    # стык
                'sheet': row[8] or '1',          # лист
                'weld_type': row[12] if weld_type_column and len(row) > 12 else (row[7] or ''),  # Тип_соединения_российский_стандарт или Код_удаления
                'deletion_code': row[7] or '',   # Код_удаления
                'comment': row[13] or ''         # Комментарий (столбец 13)
            }
            print(f"[DEBUG] Загружена запись: ID={joint_data['id']}, стык={joint_data['joint_number']}, тип_шва={joint_data['weld_type']}")
            joints.append(joint_data)
        
        conn.close()
        
        return jsonify({
            'joints': joints,
            'raspulovka_file': raspulovka_file
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/save_joints_data', methods=['POST'])
def save_joints_data():
    """API для сохранения данных о стыках в таблицу pipeline_weld_joint_iso"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        if data is None:
            return jsonify({'error': 'Не удалось получить данные запроса'})
        
        joints_data = data.get('joints', [])
        deleted_joints = data.get('deleted_joints', [])
        drawing_number = data.get('drawing_number', '').strip()
        line_number = data.get('line_number', '').strip()
        
        if not joints_data:
            return jsonify({'error': 'Нет данных для сохранения'})
        
        if not drawing_number:
            return jsonify({'error': 'Номер чертежа не может быть пустым'})
        
        if not line_number:
            return jsonify({'error': 'Номер линии не может быть пустым'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица pipeline_weld_joint_iso не найдена'})
        
        # Проверяем существование столбца Тип_соединения_российский_стандарт
        cursor.execute("PRAGMA table_info(pipeline_weld_joint_iso)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Определяем правильное название столбца для типа соединения
        weld_type_column = None
        for col in column_names:
            if 'тип' in col.lower() and ('соединен' in col.lower() or 'российск' in col.lower() or 'стандарт' in col.lower()):
                weld_type_column = col
                break
        
        if not weld_type_column:
            # Если столбец не найден, создаем его
            cursor.execute('ALTER TABLE pipeline_weld_joint_iso ADD COLUMN "Тип_соединения_российский_стандарт" TEXT')
            weld_type_column = "Тип_соединения_российский_стандарт"
        
        print(f"[DEBUG] Используем столбец для типа шва: {weld_type_column}")
        
        saved_count = 0
        updated_count = 0
        deleted_count = 0
        errors = []
        joint_updates = []  # Для отслеживания обновлений ID
        
        for joint in joints_data:
            try:
                # Подготавливаем данные из формы
                joint_id = joint.get('id')
                joint_number = joint.get('joint_number', '').strip()
                weld_type = joint.get('weld_type', '').strip()
                deletion_code = joint.get('deletion_code', '').strip()
                sheet = joint.get('sheet', '1').strip()
                comment = joint.get('comment', '').strip()
                
                # Проверяем обязательные поля
                if not joint_number:
                    errors.append(f'Стык {joint_number}: номер стыка не может быть пустым')
                    continue
                
                if not weld_type:
                    errors.append(f'Стык {joint_number}: тип шва не может быть пустым')
                    continue
                
                if not sheet or sheet == '0':
                    errors.append(f'Стык {joint_number}: лист не может быть пустым и должен быть больше 0')
                    continue
                
                # Проверяем дубликаты по ISO + стык
                cursor.execute('SELECT id FROM pipeline_weld_joint_iso WHERE "ISO" = ? AND "стык" = ?', 
                             (drawing_number, joint_number))
                existing_record = cursor.fetchone()
                
                if existing_record and (not joint_id or joint_id == ''):
                    errors.append(f'Дубликат: стык {joint_number} для чертежа {drawing_number} уже существует')
                    continue
                
                # Если есть ID - обновляем существующую запись
                if joint_id and joint_id != '':
                    print(f"[DEBUG] Обновляем существующую запись ID: {joint_id}")
                    print(f"[DEBUG] Новые данные: стык={joint_number}, лист={sheet}, тип_шва={weld_type}, комментарий={comment}")
                    
                    # Извлекаем титул из номера чертежа
                    titul_value = extract_titul_from_iso_string(drawing_number)
                    
                    update_query = f"""
                    UPDATE pipeline_weld_joint_iso 
                    SET "стык" = ?, "лист" = ?, "Комментарий" = ?, "{weld_type_column}" = ?, "Код_удаления" = ?, "Титул" = ?
                    WHERE id = ?
                    """
                    print(f"[DEBUG] SQL запрос: {update_query}")
                    print(f"[DEBUG] Параметры: {joint_number}, {sheet}, {comment}, {weld_type}, {deletion_code}, {titul_value or ''}, {joint_id}")
                    
                    cursor.execute(update_query, (
                        joint_number, sheet, comment, weld_type, deletion_code, titul_value or '', joint_id
                    ))
                    
                    # Проверяем, сколько строк было обновлено
                    rows_affected = cursor.rowcount
                    print(f"[DEBUG] Строк обновлено: {rows_affected}")
                    
                    if rows_affected == 0:
                        print(f"[WARNING] Запись с ID {joint_id} не найдена для обновления")
                        errors.append(f'Запись с ID {joint_id} не найдена для обновления')
                    else:
                        updated_count += 1
                    
                    # Добавляем информацию об обновлении
                    joint_updates.append({
                        'old_id': joint_id,
                        'new_id': joint_id,  # ID остается тем же
                        'joint_number': joint_number
                    })
                else:
                    # Извлекаем титул из номера чертежа
                    titul_value = extract_titul_from_iso_string(drawing_number)
                    
                    # Создаем новую запись
                    insert_query = f"""
                    INSERT INTO pipeline_weld_joint_iso 
                    ("Титул", "ISO", "Линия", "ключь_жср_смр", "Линия2", "стык", 
                     "Код_удаления", "лист", "повтор", "открыть", "Дата_загрузки", "{weld_type_column}", "Комментарий")
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute(insert_query, (
                        titul_value or '',  # Титул (извлеченный из ISO)
                        drawing_number,     # ISO
                        line_number,        # Линия
                        '',                 # ключь_жср_смр (оставляем пустым)
                        '',                 # Линия2 (пустая)
                        joint_number,       # стык
                        deletion_code,      # Код_удаления
                        sheet,              # лист
                        '',                 # повтор (пустой)
                        '',                 # открыть (пустой)
                        current_date,       # Дата_загрузки
                        weld_type,          # Тип_соединения_российский_стандарт
                        comment             # Комментарий
                    ))
                    
                    # Получаем ID новой записи
                    new_id = cursor.lastrowid
                    saved_count += 1
                    
                    # Добавляем информацию о новой записи
                    joint_updates.append({
                        'old_id': '',  # Пустой ID для новых записей
                        'new_id': str(new_id),
                        'joint_number': joint_number
                    })
                    
            except Exception as e:
                errors.append(f'Ошибка сохранения стыка {joint.get("joint_number", "неизвестный")}: {str(e)}')
        
        # Удаляем помеченные записи
        for deleted_id in deleted_joints:
            try:
                cursor.execute("DELETE FROM pipeline_weld_joint_iso WHERE id = ?", (deleted_id,))
                deleted_count += 1
                print(f"Удалена запись с ID: {deleted_id}")
            except Exception as e:
                errors.append(f'Ошибка удаления записи с ID {deleted_id}: {str(e)}')
        
        # Фиксируем изменения
        conn.commit()
        conn.close()
        
        print(f"[DEBUG] Отправляем joint_updates: {joint_updates}")
        
        result = {
            'success': True,
            'saved': saved_count,
            'updated': updated_count,
            'deleted': deleted_count,
            'total': saved_count + updated_count,
            'joint_updates': joint_updates
        }
        
        if errors:
            result['errors'] = errors
            result['message'] = f'Сохранено {saved_count + updated_count} записей, удалено {deleted_count} записей. Ошибки: {len(errors)}'
            print(f"[DEBUG] Ошибки при сохранении: {errors}")
        else:
            result['message'] = f'Успешно сохранено {saved_count + updated_count} записей, удалено {deleted_count} записей'
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/delete_all_joints', methods=['POST'])
def delete_all_joints():
    """API для удаления всех записей о стыках для конкретного чертежа и линии"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        drawing_number = data.get('drawing_number', '').strip()
        line_number = data.get('line_number', '').strip()
        
        if not drawing_number:
            return jsonify({'error': 'Номер чертежа не может быть пустым'})
        
        if not line_number:
            return jsonify({'error': 'Номер линии не может быть пустым'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица pipeline_weld_joint_iso не найдена'})
        
        # Сначала подсчитываем количество записей для удаления
        cursor.execute('SELECT COUNT(*) as count FROM pipeline_weld_joint_iso WHERE "ISO" = ? AND "Линия" = ?', 
                      (drawing_number, line_number))
        count_result = cursor.fetchone()
        records_count = count_result['count'] if count_result else 0
        
        if records_count == 0:
            return jsonify({'success': True, 'message': 'Нет данных для удаления'})
        
        # Удаляем все записи для данного чертежа и линии
        cursor.execute('DELETE FROM pipeline_weld_joint_iso WHERE "ISO" = ? AND "Линия" = ?', 
                      (drawing_number, line_number))
        
        # Фиксируем изменения
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Успешно удалено {records_count} записей для чертежа {drawing_number} и линии {line_number}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/delete_joint', methods=['POST'])
def delete_joint():
    """API для удаления записи о стыке из таблицы pipeline_weld_joint_iso"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        joint_id = data.get('id')
        
        if not joint_id:
            return jsonify({'error': 'ID записи не указан'})
        
        cursor = conn.cursor()
        
        # Проверяем существование записи
        cursor.execute("SELECT id FROM pipeline_weld_joint_iso WHERE id = ?", (joint_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Запись не найдена'})
        
        # Удаляем запись
        cursor.execute("DELETE FROM pipeline_weld_joint_iso WHERE id = ?", (joint_id,))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Запись успешно удалена'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/condition_weld/bulk_delete_from_pipeline', methods=['POST'])
def bulk_delete_from_pipeline_by_condition():
    """Массовое удаление строк из pipeline_weld_joint_iso по выбранным позициям из condition_weld

    Ожидает JSON:
    {
      "records": [
        {"ISO": "...", "Линия": "...", "стык": "..."},
        ...
      ]
    }
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Ошибка подключения к БД'})

    try:
        data = request.get_json() or {}
        records = data.get('records', [])

        if not records:
            return jsonify({'success': False, 'message': 'Нет записей для удаления'})

        cursor = conn.cursor()

        # Проверим наличие таблицы назначения
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Таблица pipeline_weld_joint_iso не найдена'})

        deleted_count = 0
        errors = []

        # Подготовленный запрос удаления по составному ключу
        delete_sql = 'DELETE FROM pipeline_weld_joint_iso WHERE "ISO" = ? AND "Линия" = ? AND "стык" = ?'

        for rec in records:
            try:
                iso_val = (rec.get('ISO') or '').strip()
                line_val = (rec.get('Линия') or '').strip()
                joint_val = (rec.get('стык') or '').strip()

                if not iso_val or not line_val or not joint_val:
                    errors.append(f'Пропущены ключевые поля (ISO/Линия/стык): {rec}')
                    continue

                cursor.execute(delete_sql, (iso_val, line_val, joint_val))
                deleted_count += cursor.rowcount if cursor.rowcount is not None else 0
            except Exception as e:
                errors.append(f'Ошибка удаления {rec}: {str(e)}')

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/condition_weld/bulk_delete_by_ids', methods=['POST'])
def bulk_delete_pipeline_by_ids():
    """Массовое удаление из pipeline_weld_joint_iso по списку ID (из condition_weld.id)

    Ожидает JSON: { "record_ids": [1,2,3] }
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Ошибка подключения к БД'})

    try:
        data = request.get_json() or {}
        record_ids = data.get('record_ids', [])

        # Очистим и приведем к int
        safe_ids = []
        for rid in record_ids:
            try:
                if rid is None or str(rid).strip() == '':
                    continue
                safe_ids.append(int(str(rid)))
            except Exception:
                continue

        if not safe_ids:
            return jsonify({'success': False, 'message': 'Список ID пуст или некорректен'})

        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Таблица pipeline_weld_joint_iso не найдена'})

        # Удаление по пачке ID
        placeholders = ','.join(['?'] * len(safe_ids))
        delete_sql = f"DELETE FROM pipeline_weld_joint_iso WHERE id IN ({placeholders})"
        cursor.execute(delete_sql, safe_ids)
        deleted_count = cursor.rowcount if cursor.rowcount is not None else 0

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'deleted_count': deleted_count})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/open_raspulovka_file', methods=['POST'])
def open_raspulovka_file():
    """API для открытия файла распуловки"""
    try:
        data = request.get_json()
        drawing_number = data.get('drawing_number', '').strip()
        line_number = data.get('line_number', '').strip()
        
        if not drawing_number or not line_number:
            return jsonify({'error': 'Необходимо указать номер чертежа и линии'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'})
        
        cursor = conn.cursor()
        
        # Получаем путь к файлу распуловки из Log_Piping_PTO
        cursor.execute("""
            SELECT "распуловка" 
            FROM Log_Piping_PTO 
            WHERE "ISO" = ? AND "Линия" = ? 
            LIMIT 1
        """, (drawing_number, line_number))
        
        result = cursor.fetchone()
        conn.close()
        
        if not result or not result[0]:
            return jsonify({'error': 'Файл распуловки не найден для указанного чертежа и линии'})
        
        file_path = result[0]
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            return jsonify({'error': f'Файл не найден: {file_path}'})
        
        # Открываем файл с помощью системного приложения по умолчанию
        import subprocess
        import platform
        
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
            
            return jsonify({
                'success': True,
                'message': f'Файл открыт: {os.path.basename(file_path)}'
            })
        except Exception as e:
            return jsonify({'error': f'Ошибка при открытии файла: {str(e)}'})
            
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_joints_drawings', methods=['GET'])
def get_joints_drawings():
    """API для получения списка уникальных чертежей из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбец с ISO - возможные варианты названий
        iso_column = None
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
                break
        
        if not iso_column:
            return jsonify({'error': f'Столбец с ISO не найден. Доступные столбцы: {column_names}'})
        
        # Получаем уникальные чертежи из Log_Piping_PTO
        cursor.execute(f'SELECT DISTINCT `{iso_column}` FROM `Log_Piping_PTO` WHERE `{iso_column}` IS NOT NULL AND `{iso_column}` != "" ORDER BY `{iso_column}`')
        drawings = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'drawings': drawings})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_joints_lines', methods=['GET'])
def get_joints_lines():
    """API для получения списка уникальных линий из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбец с Линия - используем ТОЛЬКО столбец 'Линия'
        line_column = None
        for col in column_names:
            if col == 'Линия':
                line_column = col
                break
        
        if not line_column:
            return jsonify({'error': f'Столбец с Линия не найден. Доступные столбцы: {column_names}'})
        
        # Получаем уникальные линии из Log_Piping_PTO
        cursor.execute(f'SELECT DISTINCT `{line_column}` FROM `Log_Piping_PTO` WHERE `{line_column}` IS NOT NULL AND `{line_column}` != "" ORDER BY `{line_column}`')
        lines = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'lines': lines})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_joints_lines_by_drawing', methods=['POST'])
def get_joints_lines_by_drawing():
    """API для получения линий по выбранному чертежу из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        drawing_number = data.get('iso', '').strip()
        
        if not drawing_number:
            return jsonify({'error': 'Номер чертежа не указан'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбцы ISO и Линия (используем ТОЛЬКО 'Линия', игнорируем 'Линия2')
        iso_column = None
        line_column = None
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
            # Используем ТОЛЬКО столбец 'Линия'
            if col == 'Линия':
                line_column = col
        
        if not iso_column:
            return jsonify({'error': f'Столбец с ISO не найден. Доступные столбцы: {column_names}'})
        if not line_column:
            return jsonify({'error': f'Столбец с Линия не найден. Доступные столбцы: {column_names}'})
        
        # Получаем линии для выбранного чертежа из Log_Piping_PTO
        cursor.execute(f'SELECT DISTINCT `{line_column}` FROM `Log_Piping_PTO` WHERE `{iso_column}` = ? AND `{line_column}` IS NOT NULL AND `{line_column}` != "" ORDER BY `{line_column}`', (drawing_number,))
        lines = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'lines': lines})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_joints_drawings_by_line', methods=['POST'])
def get_joints_drawings_by_line():
    """API для получения чертежей по выбранной линии из таблицы Log_Piping_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        line_number = data.get('line', '').strip()
        
        if not line_number:
            return jsonify({'error': 'Номер линии не указан'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы Log_Piping_PTO
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log_Piping_PTO'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица Log_Piping_PTO не найдена'})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(Log_Piping_PTO)")
        columns_info = cursor.fetchall()
        column_names = [col[1] for col in columns_info]
        
        # Ищем столбцы ISO и Линия
        iso_column = None
        line_column = None
        for col in column_names:
            if 'iso' in col.lower() or col == 'ISO':
                iso_column = col
            # Используем ТОЛЬКО столбец 'Линия'
            if col == 'Линия':
                line_column = col
        
        if not iso_column:
            return jsonify({'error': f'Столбец с ISO не найден. Доступные столбцы: {column_names}'})
        if not line_column:
            return jsonify({'error': f'Столбец с Линия не найден. Доступные столбцы: {column_names}'})
        
        # Получаем чертежи для выбранной линии из Log_Piping_PTO
        # Ищем линии, содержащие введенный номер (для поддержки поиска по частичному совпадению)
        cursor.execute(f'SELECT DISTINCT `{iso_column}` FROM `Log_Piping_PTO` WHERE `{line_column}` LIKE ? AND `{iso_column}` IS NOT NULL AND `{iso_column}` != "" ORDER BY `{iso_column}`', (f'%{line_number}%',))
        drawings = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'drawings': drawings})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/logs')
def logs():
    """Страница журналов"""
    # Получаем краткую статистику для отображения на карточке
    status_statistics = get_status_statistics()
    
    # Формируем краткую статистику для карточки
    brief_stats = {
        'total_records': status_statistics.get('total_records', 0),
        'rk_statuses_count': len(status_statistics.get('rk_status_stats', [])),
        'vik_statuses_count': len(status_statistics.get('vik_status_stats', []))
    }
    
    return render_template('logs.html', brief_stats=brief_stats, version='v2.0')

@app.route('/system_logs')
def system_logs():
    """Страница системных логов"""
    try:
        # Получаем список лог-файлов
        logs_dir = os.path.join(Config.PROJECT_ROOT, 'logs')
        log_files = []
        
        logger.info(f"Поиск логов в директории: {logs_dir}")
        logger.info(f"Директория существует: {os.path.exists(logs_dir)}")
        
        if os.path.exists(logs_dir):
            all_files = os.listdir(logs_dir)
            logger.info(f"Всего файлов в директории: {len(all_files)}")
            
            for file in all_files:
                if file.endswith('.log'):
                    file_path = os.path.join(logs_dir, file)
                    file_stat = os.stat(file_path)
                    file_size = file_stat.st_size
                    file_mtime = datetime.fromtimestamp(file_stat.st_mtime)
                    
                    log_files.append({
                        'name': file,
                        'path': file_path,
                        'size': file_size,
                        'size_mb': round(file_size / (1024 * 1024), 2),
                        'modified': file_mtime.strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        # Сортируем по дате изменения (новые сверху)
        log_files.sort(key=lambda x: x['modified'], reverse=True)
        
        logger.info(f"Найдено лог-файлов: {len(log_files)}")
        
        return render_template('system_logs.html', log_files=log_files)
    except Exception as e:
        logger.error(f'Ошибка получения списка системных логов: {e}')
        flash(f'Ошибка получения списка системных логов: {e}', 'error')
        return render_template('system_logs.html', log_files=[])

@app.route('/reports')
def reports():
    """Страница отчетов"""
    try:
        # Получаем статистику из базы данных для отчетов
        conn = get_db_connection()
        logger.info(f"Подключение к БД: {'OK' if conn else 'FAILED'}")
        
        if not conn:
            flash('Ошибка подключения к базе данных', 'error')
            return render_template('reports.html', reports_data={})
        
        cursor = conn.cursor()
        reports_data = {}
        
        # Получаем список таблиц
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row['name'] for row in cursor.fetchall()]
        logger.info(f"Найдено таблиц в БД: {len(tables)}")
        
        # Статистика по основным таблицам
        for table in tables:
            if not table.startswith('sqlite_'):
                try:
                    cursor.execute(f"SELECT COUNT(*) as count FROM `{table}`")
                    count = cursor.fetchone()['count']
                    reports_data[table] = {
                        'name': table,
                        'count': count,
                        'description': get_table_description(table)
                    }
                except Exception as e:
                    logger.error(f"Ошибка получения статистики для таблицы {table}: {e}")
        
        conn.close()
        
        logger.info(f"Подготовлено данных для отчетов: {len(reports_data)} таблиц")
        
        return render_template('reports.html', reports_data=reports_data)
    except Exception as e:
        logger.error(f'Ошибка получения данных для отчетов: {e}')
        flash(f'Ошибка получения данных для отчетов: {e}', 'error')
        return render_template('reports.html', reports_data={})



def get_table_description(table_name):
    """Получает описание таблицы и полей"""
    table_descriptions = {
        'Log_Piping_PTO': 'Журнал трубопроводов ПТО',
        'NDT_Findings_Transmission_Register': 'Реестр результатов НК трансмиссии',
        'Pipeline_Test_Package': 'Пакет испытаний трубопроводов',
        'WL_Report_SMR': 'Отчет по СМР',
        'WL_China': 'Данные по Китаю',
        'Work_Order_Log_NDT': 'Журнал нарядов НК',
        'Staff_Titles_M_Kran': 'Штатное расписание М_Кран',
        'weld_repair_log': 'Журнал ремонтов сварных соединений',
        'logs_lnk': 'Основная таблица логов',
        'nk_results': 'Результаты неразрушающего контроля'
    }
    
    # Возвращаем словарь с описанием таблицы и полей
    return {
        '_table': table_descriptions.get(table_name, f'Таблица {table_name}'),
        # Здесь можно добавить описания конкретных полей для каждой таблицы
    }



@app.route('/api/get_table_columns/<table_name>')
def get_table_columns(table_name):
    """API для получения столбцов таблицы"""
    # Логируем в файл
    try:
        with open('debug.log', 'a', encoding='utf-8') as f:
            f.write(f"\n=== ПОЛУЧЕНИЕ СТОЛБЦОВ ТАБЛИЦЫ {table_name} === {datetime.now()}\n")
    except:
        pass
    
    # Убираем print statements, которые вызывают ошибку кодировки
    
    conn = get_db_connection()
    if not conn:
        try:
            with open('debug.log', 'a', encoding='utf-8') as f:
                f.write("Ошибка подключения к базе данных\n")
        except:
            pass
        return jsonify({'error': 'Ошибка подключения к базе данных'})
    
    try:
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info(`{table_name}`)")
        columns = [{'name': row['name'], 'type': row['type']} for row in cursor.fetchall()]
        conn.close()
        
        # Убираем print statements, которые вызывают ошибку кодировки
        
        return jsonify({'columns': columns})
    except Exception as e:
        try:
            with open('debug.log', 'a', encoding='utf-8') as f:
                f.write(f"Ошибка при получении столбцов для таблицы {table_name}: {e}\n")
        except:
            pass
        return jsonify({'error': str(e)})

@app.route('/filtered_data/<filter_type>')
def filtered_data(filter_type):
    """Страница с отфильтрованными данными по типу"""
    title = request.args.get('title', 'all')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search_term = request.args.get('search', '').strip()
    
    # Получаем фильтры по столбцам
    column_filters = {}
    for key, value in request.args.items():
        if key.startswith('filter_') and value.strip():
            column_name = key[7:]  # Убираем префикс 'filter_'
            column_filters[column_name] = value.strip()
    
    # Ограничиваем per_page разумными значениями
    if per_page < 10:
        per_page = 10
    elif per_page > 1000:
        per_page = 1000
    
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('index'))
    
    try:
        cursor = conn.cursor()
        
        # Определяем фильтр и заголовок на основе типа
        # ИСПРАВЛЕНО: Используем точно такие же критерии, как в статистике
        filter_conditions = {
            'rk_defects': {
                'condition': '("Статус_РК" LIKE "%Не годен%")',
                'title': 'РК Вырез/Ремонт (ОФИЦИАЛЬНЫЙ)',
                'description': 'Записи со статусом РК: Ремонт, Вырез, Вырезать',
                'table': 'logs_lnk'
            },
            'all_defects': {
                'condition': '("Статус_РК" LIKE "%Не годен%" OR "Статус_РК" = "Н/П")',
                'title': 'ВСЕГО НЕГОДНЫХ (по статистике)',
                'description': 'Записи со статусом РК: Ремонт, Вырез, Вырезать, Н/П (точно как в статистике)',
                'table': 'logs_lnk'
            },
            'logs': {
                'condition': '1=1',
                'title': 'Журнал НГС Эксперт',
                'description': 'Все записи из основного журнала',
                'table': 'logs_lnk'
            },
            'weld_log': {
                'condition': '1=1',
                'title': 'Журнал сварки',
                'description': 'Все записи из журнала сварки',
                'table': 'wl_china'
            },
            'repair_log': {
                'condition': '1=1',
                'title': 'Журнал ремонта',
                'description': 'Все записи из журнала ремонта',
                'table': 'weld_repair_log'
            }
        }
        
        if filter_type not in filter_conditions:
            flash('Неизвестный тип фильтра', 'error')
            return redirect(url_for('index'))
        
        filter_info = filter_conditions[filter_type]
        table_name = filter_info['table']
        
        # Условие для фильтрации по титулам
        title_condition = ""
        title_params = []
        if title and title != 'all':
            if table_name == 'wl_china':
                # Для таблицы wl_china используем столбец блок_N
                title_condition = 'AND "блок_N" LIKE ?'
            else:
                # Для других таблиц используем Титул
                title_condition = 'AND "Титул" LIKE ?'
            title_params.append(f'%{title}%')
        
        # Условие для поиска (как в view_table - регистронезависимый с двойным поиском)
        search_condition = ""
        search_params = []
        if search_term:
            if table_name == 'wl_china':
                # Для таблицы wl_china ищем по основным столбцам (как в view_table)
                search_columns = ["Номер_сварного_шва", "Номер_чертежа", "блок_N", "N_Линии"]
                search_conditions = []
                for col in search_columns:
                    # Используем двойной поиск: оригинальный термин + нижний регистр
                    search_conditions.append(f'(CAST("{col}" AS TEXT) LIKE ? OR LOWER(CAST("{col}" AS TEXT)) LIKE ?)')
                    search_params.append(f'%{search_term}%')
                    search_params.append(f'%{search_term.lower()}%')
                search_condition = f'AND ({' OR '.join(search_conditions)})'
            elif table_name == 'logs_lnk':
                # Для таблицы logs_lnk ищем по основным столбцам (как в view_table)
                search_columns = ["Чертеж", "Номер_стыка", "Линия", "Источник"]
                search_conditions = []
                for col in search_columns:
                    # Используем двойной поиск: оригинальный термин + нижний регистр
                    search_conditions.append(f'(CAST("{col}" AS TEXT) LIKE ? OR LOWER(CAST("{col}" AS TEXT)) LIKE ?)')
                    search_params.append(f'%{search_term}%')
                    search_params.append(f'%{search_term.lower()}%')
                search_condition = f'AND ({' OR '.join(search_conditions)})'
            elif table_name == 'weld_repair_log':
                # Для таблицы weld_repair_log ищем по основным столбцам (как в view_table)
                search_columns = ["Чертеж", "Номер_стыка", "Линия"]
                search_conditions = []
                for col in search_columns:
                    # Используем двойной поиск: оригинальный термин + нижний регистр
                    search_conditions.append(f'(CAST("{col}" AS TEXT) LIKE ? OR LOWER(CAST("{col}" AS TEXT)) LIKE ?)')
                    search_params.append(f'%{search_term}%')
                    search_params.append(f'%{search_term.lower()}%')
                search_condition = f'AND ({' OR '.join(search_conditions)})'
        
        # Добавляем фильтры по столбцам (как в view_table)
        column_filter_condition = ""
        if column_filters:
            logger.info(f"Фильтры по столбцам: {column_filters}")
            column_conditions = []
            for column_name, filter_value in column_filters.items():
                # Регистронезависимый поиск по конкретному столбцу (как в view_table)
                column_conditions.append(f'(CAST("{column_name}" AS TEXT) LIKE ? OR LOWER(CAST("{column_name}" AS TEXT)) LIKE ?)')
                search_params.append(f'%{filter_value}%')
                search_params.append(f'%{filter_value.lower()}%')
                logger.info(f"Добавлен фильтр для столбца '{column_name}': '{filter_value}'")
                logger.info(f"Добавлено условие: (CAST('{column_name}' AS TEXT) LIKE ? OR LOWER(CAST('{column_name}' AS TEXT)) LIKE ?)")
                logger.info(f"Добавлены параметры: %{filter_value}%, %{filter_value.lower()}%")
            
            if column_conditions:
                column_filter_condition = f' AND ({' AND '.join(column_conditions)})'
        
        # Объединяем параметры
        query_params = title_params + search_params
        
        # Отладочная информация для параметров
        logger.info(f"title_params: {title_params}")
        logger.info(f"search_params: {search_params}")
        logger.info(f"column_filter_condition: {column_filter_condition}")
        logger.info(f"query_params: {query_params}")
        logger.info(f"Количество параметров: {len(query_params)}")
        
        # Проверяем существование таблицы weld_repair_log и получаем ID уже перенесенных записей
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weld_repair_log'")
        weld_repair_table_exists = cursor.fetchone()
        
        exclude_condition = ""
        excluded_count = 0
        if weld_repair_table_exists and table_name == 'logs_lnk':
            # Получаем количество записей в weld_repair_log для отладки
            cursor.execute("SELECT COUNT(*) as count FROM weld_repair_log")
            excluded_count = cursor.fetchone()['count']
            logger.info(f"Таблица weld_repair_log существует, в ней {excluded_count} записей")
            
            # Показываем несколько ID для отладки
            cursor.execute("SELECT app_row_id FROM weld_repair_log LIMIT 5")
            sample_ids = [row['app_row_id'] for row in cursor.fetchall()]
            logger.info(f"Примеры ID в weld_repair_log: {sample_ids}")
            
            # ИСПРАВЛЕНО: Для all_defects исключаем записи, которые уже перенесены в журнал ремонта
            # Это нужно для показа только тех негодных записей, которые еще не обработаны
            if filter_type == 'all_defects':
                logger.info(f"Фильтр '{filter_type}' - показываем только негодные записи, НЕ перенесенные в журнал ремонта")
                exclude_condition = f' AND NOT EXISTS (SELECT 1 FROM weld_repair_log WHERE weld_repair_log.app_row_id = {table_name}.app_row_id)'
            else:
                # ИСПРАВЛЕННОЕ условие исключения - используем EXISTS вместо NOT IN для лучшей производительности
                exclude_condition = f' AND NOT EXISTS (SELECT 1 FROM weld_repair_log WHERE weld_repair_log.app_row_id = {table_name}.app_row_id)'
                logger.info(f"Фильтр '{filter_type}' - исключаем перенесенные записи")
            
            # Проверим сколько записей будет исключено для текущего фильтра
            if exclude_condition:
                test_exclude_query = f'''
                    SELECT COUNT(*) as count 
                    FROM {table_name} 
                    WHERE {filter_info["condition"]} {title_condition} {search_condition}
                    AND EXISTS (SELECT 1 FROM weld_repair_log WHERE weld_repair_log.app_row_id = {table_name}.app_row_id)
                '''
                cursor.execute(test_exclude_query, query_params)
                actually_excluded = cursor.fetchone()['count']
                logger.info(f"Записей будет исключено для фильтра '{filter_type}': {actually_excluded}")
            else:
                logger.info(f"Для фильтра '{filter_type}' исключения не применяются")
        else:
            logger.info("Таблица weld_repair_log не существует или используется другая таблица, исключения не применяются")
        
        # Получаем общее количество записей (исключая уже перенесенные в журнал ремонта)
        # Строим WHERE условие динамически (как в view_table)
        where_conditions = []
        
        # Добавляем базовое условие фильтра
        if filter_info["condition"] and filter_info["condition"].strip():
            where_conditions.append(filter_info["condition"])
        
        # Добавляем остальные условия, убирая начальный AND если есть
        if title_condition and title_condition.strip():
            condition = title_condition.strip()
            if condition.startswith('AND '):
                condition = condition[4:]  # Убираем 'AND '
            where_conditions.append(condition)
            
        if search_condition and search_condition.strip():
            condition = search_condition.strip()
            if condition.startswith('AND '):
                condition = condition[4:]  # Убираем 'AND '
            where_conditions.append(condition)
            
        if column_filter_condition and column_filter_condition.strip():
            condition = column_filter_condition.strip()
            if condition.startswith('AND '):
                condition = condition[4:]  # Убираем 'AND '
            where_conditions.append(condition)
            
        if exclude_condition and exclude_condition.strip():
            condition = exclude_condition.strip()
            if condition.startswith('AND '):
                condition = condition[4:]  # Убираем 'AND '
            where_conditions.append(condition)
        
        # Если нет условий, используем 1=1
        if not where_conditions:
            where_clause = '1=1'
        else:
            where_clause = ' AND '.join(where_conditions)
        
        count_query = f'''
            SELECT COUNT(*) as count 
            FROM {table_name} 
            WHERE {where_clause}
        '''
        
        # Отладочная информация для SQL-запроса
        logger.info(f"count_query: {count_query}")
        logger.info(f"Количество ? в запросе: {count_query.count('?')}")
        logger.info(f"Количество параметров: {len(query_params)}")
        logger.info(f"Параметры: {query_params}")
        logger.info(f"exclude_condition: '{exclude_condition}'")
        logger.info(f"where_conditions: {where_conditions}")
        logger.info(f"where_clause: {where_clause}")
        
        # Проверяем соответствие количества параметров
        param_count = count_query.count('?')
        if param_count != len(query_params):
            logger.error(f"НЕСООТВЕТСТВИЕ: {param_count} параметров в запросе, но {len(query_params)} значений")
            # Если параметров больше чем нужно, обрезаем
            if len(query_params) > param_count:
                query_params = query_params[:param_count]
                logger.info(f"Обрезали параметры до {param_count}")
        
        cursor.execute(count_query, query_params)
        total_records = cursor.fetchone()['count']
        
        # Отладочная информация
        logger.info(f"Фильтр: {filter_type}")
        logger.info(f"Таблица: {table_name}")
        logger.info(f"Условие фильтра: {filter_info['condition']}")
        logger.info(f"Условие исключения: {exclude_condition}")
        logger.info(f"Итоговый запрос: {count_query}")
        logger.info(f"Найдено записей после исключения: {total_records}")
        
        # Дополнительная проверка - получаем общее количество записей БЕЗ исключения
        count_query_no_exclude = f'''
            SELECT COUNT(*) as count 
            FROM {table_name} 
            WHERE {filter_info["condition"]} {title_condition} {search_condition}
        '''
        cursor.execute(count_query_no_exclude, query_params)
        total_without_exclusion = cursor.fetchone()['count']
        logger.info(f"Всего записей БЕЗ исключения: {total_without_exclusion}")
        logger.info(f"Разница (должно быть исключено): {total_without_exclusion - total_records}")
        
        # Определяем нужные столбцы для отображения в зависимости от таблицы
        if table_name == 'logs_lnk':
            display_columns = [
                'app_row_id',
                'Чертеж', 
                'Линия',
                'Лист',
                'Номер_стыка',
                'РК',
                'Статус_РК'
            ]
            # Для корректной работы переноса нужно получать дополнительные поля
            all_columns = display_columns + ['Диаметр_1', 'Толщина_1', 'Дата_сварки', 'Примечания_заключений', 'Заявленны_виды_контроля']
        elif table_name == 'wl_china':
            # Получаем все столбцы из таблицы wl_china
            cursor.execute("PRAGMA table_info(wl_china)")
            all_columns = [col['name'] for col in cursor.fetchall()]
            
            # Получаем выбранные столбцы из параметров запроса
            selected_columns = request.args.get('columns', '')
            
            if selected_columns and selected_columns.strip():
                # Очищаем строку от лишних символов
                cleaned_columns = selected_columns.strip()
                if cleaned_columns.startswith('[') and cleaned_columns.endswith(']'):
                    # Это JSON-подобная строка, извлекаем содержимое
                    import ast
                    try:
                        selected_columns_list = ast.literal_eval(cleaned_columns)
                    except:
                        # Если не удалось распарсить, используем split
                        selected_columns_list = [col.strip().strip("'\"") for col in cleaned_columns[1:-1].split(',')]
                else:
                    # Обычная строка через запятую
                    selected_columns_list = [col.strip() for col in selected_columns.split(',')]
                
                # Фильтруем только существующие столбцы и убираем пустые значения
                display_columns = [col for col in selected_columns_list if col in all_columns and col.strip()]
                
                # Если после фильтрации не осталось столбцов, показываем все
                if not display_columns:
                    display_columns = all_columns
            else:
                # По умолчанию показываем ВСЕ столбцы
                display_columns = all_columns
        elif table_name == 'weld_repair_log':
            display_columns = [
                'app_row_id',
                'Чертеж', 
                'Линия',
                'Лист',
                'Номер_стыка',
                'Тип_ремонта',
                'Статус'
            ]
            all_columns = display_columns
        else:
            # По умолчанию для неизвестных таблиц
            display_columns = ['*']
            all_columns = ['*']
        
        # Проверяем, что all_columns - это список
        if not isinstance(all_columns, list):
            logger.error(f"all_columns не является списком: {type(all_columns)} = {all_columns}")
            all_columns = ['*']
        
        columns_str = ', '.join([f'"{col}"' for col in all_columns]) if all_columns != ['*'] else '*'
        
        # Определяем порядок сортировки в зависимости от таблицы
        if table_name == 'logs_lnk':
            order_by = 'ORDER BY "Дата_загрузки" DESC, "Чертеж", "Номер_стыка"'
        elif table_name == 'wl_china':
            order_by = 'ORDER BY "Номер_сварного_шва", "Номер_чертежа"'
        elif table_name == 'weld_repair_log':
            order_by = 'ORDER BY "Дата_ремонта" DESC, "Чертеж", "Номер_стыка"'
        else:
            order_by = 'ORDER BY "Чертеж", "Номер_стыка"'
        
        data_query = f'''
            SELECT {columns_str}
            FROM {table_name} 
            WHERE {where_clause}
            {order_by}
            LIMIT {per_page} OFFSET {offset}
        '''
        cursor.execute(data_query, query_params)
        raw_records = cursor.fetchall()
        
        # Преобразуем Row объекты в словари для корректной JSON сериализации
        records = []
        for record in raw_records:
            if record is None:
                continue
            record_dict = {}
            # Заполняем все ожидаемые столбцы, включая отсутствующие
            for col in display_columns:
                try:
                    # Используем dict() для преобразования Row в словарь
                    value = dict(record).get(col, '')
                    # Преобразуем все значения в строки для безопасности
                    record_dict[col] = str(value) if value is not None else ''
                except (KeyError, TypeError, AttributeError) as e:
                    logger.error(f"Ошибка при обработке столбца {col}: {e}")
                    record_dict[col] = ''
            records.append(record_dict)
        
        # Используем предопределенный список столбцов
        columns = display_columns
        
        conn.close()
        
        total_pages = (total_records + per_page - 1) // per_page
        
        # Дополнительные параметры для wl_china
        template_params = {
            'filter_type': filter_type,
            'filter_title': filter_info['title'],
            'filter_description': filter_info['description'],
            'records': records,
            'columns': columns,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
            'total_records': total_records,
            'selected_title': title,
            'search_term': search_term,
            'table_name': table_name,  # Всегда передаем table_name
            'column_filters': column_filters  # Передаем фильтры по столбцам
        }
        
        # Добавляем параметры для выбора столбцов только для wl_china
        if table_name == 'wl_china':
            template_params.update({
                'all_columns': all_columns,
                'selected_columns': display_columns
            })
        
        return render_template('filtered_data.html', **template_params)
        
    except Exception as e:
        logger.error(f'Ошибка получения отфильтрованных данных: {e}')
        flash(f'Ошибка получения данных: {e}', 'error')
        return redirect(url_for('index'))

@app.route('/api/transfer_to_weld_repair_log', methods=['POST'])
def transfer_to_weld_repair_log():
    """API для переноса выбранных записей в таблицу weld_repair_log"""
    try:
        data = request.get_json()
        records = data.get('records', [])
        
        logger.info(f"Получен запрос на перенос {len(records)} записей")
        
        if not records:
            return jsonify({'success': False, 'message': 'Нет записей для переноса'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы weld_repair_log
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weld_repair_log'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            # Создаем таблицу weld_repair_log если её нет
            create_table_query = '''
                CREATE TABLE weld_repair_log (
                    app_row_id INTEGER PRIMARY KEY,
                    "Чертеж" TEXT,
                    "Линия" TEXT,
                    "Диаметр и толщина стенки" TEXT,
                    "№ стыка" TEXT,
                    "Дата сварки" TEXT,
                    "Размер выборки (длина, ширина, глубина), мм" TEXT,
                    "Способ и результаты контроля выборки" TEXT,
                    "Марка стали" TEXT,
                    "Фамилия, инициалы, клеймо сварщика допустившего брак" TEXT,
                    "Вид сварки" TEXT,
                    "Число ремонтов на одном участке" TEXT,
                    "Дата сварки после ремонта" TEXT,
                    "Ответственный" TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            '''
            cursor.execute(create_table_query)
            logger.info("Создана таблица weld_repair_log")
        else:
            # Проверяем и добавляем недостающие столбцы
            cursor.execute("PRAGMA table_info(weld_repair_log)")
            existing_columns = [col[1] for col in cursor.fetchall()]
            
            new_columns = [
                "Марка стали",
                "Фамилия, инициалы, клеймо сварщика допустившего брак",
                "Вид сварки",
                "Число ремонтов на одном участке",
                "Дата сварки после ремонта",
                "Ответственный"
            ]
            
            for column in new_columns:
                if column not in existing_columns:
                    try:
                        cursor.execute(f'ALTER TABLE weld_repair_log ADD COLUMN "{column}" TEXT')
                        logger.info(f"Добавлен столбец: {column}")
                    except Exception as e:
                        logger.warning(f"Ошибка добавления столбца {column}: {e}")
            
            conn.commit()
        
        transferred_count = 0
        skipped_count = 0
        
        for record in records:
            app_row_id = record.get('app_row_id')
            
            if not app_row_id:
                logger.warning("Запись без app_row_id пропущена")
                skipped_count += 1
                continue
            
            # Проверяем, существует ли уже запись с таким app_row_id
            cursor.execute('SELECT app_row_id FROM weld_repair_log WHERE app_row_id = ?', (app_row_id,))
            existing_record = cursor.fetchone()
            if existing_record:
                logger.info(f"Запись с app_row_id {app_row_id} уже существует в weld_repair_log, пропускаем")
                skipped_count += 1
                continue
            
            # ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА: Убеждаемся, что запись имеет правильный статус для переноса
            # Получаем статус записи из logs_lnk
            cursor.execute('SELECT "Статус_РК", "Статус_ВИК" FROM logs_lnk WHERE app_row_id = ?', (app_row_id,))
            status_record = cursor.fetchone()
            
            if not status_record:
                logger.warning(f"Запись с app_row_id {app_row_id} не найдена в logs_lnk, пропускаем")
                skipped_count += 1
                continue
            
            status_rk = status_record[0] or ''
            status_vik = status_record[1] or ''
            
            # Проверяем, соответствует ли запись критериям для переноса в weld_repair_log
            # Для фильтра "all_defects" переносим записи с "Не годен" или "Н/П" в РК, независимо от ВИК
            is_valid_for_transfer = (
                'Не годен' in status_rk or 
                status_rk == 'Н/П'
            )
            
            
            if not is_valid_for_transfer:
                logger.warning(f"Запись с app_row_id {app_row_id} имеет неподходящий статус РК: '{status_rk}' (требуется 'Не годен' или 'Н/П'), пропускаем")
                skipped_count += 1
                continue
            
            # Получаем полные данные записи из базы данных logs_lnk
            cursor.execute('''
                SELECT "Чертеж", "Линия", "Диаметр_1", "Толщина_1", "Номер_стыка", 
                       "Дата_сварки", "Примечания_заключений", "Заявленны_виды_контроля", "_Номер_сварного_шва_без_S_F_"
                FROM logs_lnk 
                WHERE app_row_id = ?
            ''', (app_row_id,))
            
            full_record = cursor.fetchone()
            if not full_record:
                logger.warning(f"Полные данные для записи {app_row_id} не найдены в logs_lnk")
                skipped_count += 1
                continue
            
            # Подготавливаем данные для вставки согласно сопоставлению
            chertezh = full_record[0] or ''
            liniya = full_record[1] or ''
            
            # Объединяем диаметр и толщину
            diametr = full_record[2] or ''
            tolshchina = full_record[3] or ''
            diametr_tolshchina = f"{diametr}х{tolshchina}" if diametr and tolshchina else (diametr or tolshchina or '')
            
            nomer_styka = full_record[4] or ''
            data_svarki = full_record[5] or ''
            
            # Размер выборки = Примечания_заключений
            razmer_vyborki = full_record[6] or ''
            
            # Способ и результаты контроля выборки = Заявленны_виды_контроля
            sposob_kontrolya = full_record[7] or ''
            joint_sf = ''
            try:
                joint_sf = full_record[8] or ''
            except Exception:
                joint_sf = ''

            # Попытка получить марку стали из wl_china.Базовый_материал_1 по ключам:
            # Чертеж (logs_lnk."Чертеж") = wl_china."Номер_чертежа"
            # и номер шва: строго wl_china."_Номер_сварного_шва_без_S_F_"
            marka_stali = ''
            try:
                cursor.execute("PRAGMA table_info(wl_china)")
                wl_china_columns = [row[1] for row in cursor.fetchall()]
                if 'Номер_чертежа' in wl_china_columns and 'Базовый_материал_1' in wl_china_columns and '_Номер_сварного_шва_без_S_F_' in wl_china_columns and joint_sf:
                    cursor.execute(
                        '''SELECT "Базовый_материал_1" FROM wl_china 
                           WHERE "Номер_чертежа" = ? AND "_Номер_сварного_шва_без_S_F_" = ? LIMIT 1''',
                        (chertezh, joint_sf)
                    )
                    row = cursor.fetchone()
                    if row and row[0] is not None:
                        marka_stали = row[0]
                    else:
                        cursor.execute(
                            '''SELECT "Базовый_материал_1" FROM wl_china 
                               WHERE "Номер_чертежа" = ? AND CAST("_Номер_сварного_шва_без_S_F_" AS TEXT) = CAST(? AS TEXT) LIMIT 1''',
                            (chertezh, joint_sf)
                        )
                        row = cursor.fetchone()
                        if row and row[0] is not None:
                            marka_stали = row[0]
            except Exception:
                # Тихо игнорируем, если таблицы/колонки нет — поле останется пустым
                pass
            
            
            # Вставляем запись
            insert_query = '''
                INSERT INTO weld_repair_log 
                (app_row_id, "Чертеж", "Линия", "Диаметр и толщина стенки", "№ стыка", "Дата сварки", "Размер выборки (длина, ширина, глубина), мм", "Способ и результаты контроля выборки", "Марка стали")
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            '''
            
            cursor.execute(insert_query, (
                app_row_id,
                chertezh,
                liniya, 
                diametr_tolshchina,
                nomer_styka,
                data_svarki,
                razmer_vyborki,
                sposob_kontrolya,
                marka_stali
            ))
            
            transferred_count += 1
        
        # Сохраняем изменения
        conn.commit()
        conn.close()
        
        message = f'Перенесено: {transferred_count}, пропущено: {skipped_count}'
        logger.info(f"Перенос в weld_repair_log завершен: {message}")
        
        return jsonify({
            'success': True, 
            'transferred_count': transferred_count,
            'skipped_count': skipped_count,
            'message': message
        })
        
    except Exception as e:
        logger.error(f'Ошибка переноса в weld_repair_log: {e}')
        return jsonify({'success': False, 'message': f'Ошибка переноса: {str(e)}'})





@app.route('/api/open_folder', methods=['POST'])
def open_folder():
    """API для открытия папки в проводнике"""
    import subprocess
    import platform
    
    try:
        data = request.get_json()
        folder_name = data.get('folder_name')
        
        if not folder_name:
            return jsonify({'success': False, 'message': 'Не указана папка'})
        
        # Обрабатываем путь к папке
        if folder_name.startswith('D:/'):
            # Абсолютный путь - используем как есть, заменив / на \
            folder_path = folder_name.replace('/', '\\')
        else:
            # Относительный путь - строим от PROJECT_ROOT
            folder_path = os.path.join(app.config['PROJECT_ROOT'], folder_name)
        
        if not os.path.exists(folder_path):
            return jsonify({'success': False, 'message': f'Папка не найдена: {folder_path}'})
        
        # Открываем папку в зависимости от ОС
        system = platform.system()
        if system == 'Windows':
            subprocess.run(['explorer', folder_path])
        elif system == 'Darwin':  # macOS
            subprocess.run(['open', folder_path])
        else:  # Linux
            subprocess.run(['xdg-open', folder_path])
        
        return jsonify({'success': True, 'message': f'Папка открыта: {folder_name}'})
        
    except Exception as e:
        logger.error(f'Ошибка открытия папки: {e}')
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

@app.route('/api/weld_repair_log')
def api_weld_repair_log():
    """API для получения данных weld_repair_log с пагинацией и фильтрами"""
    logger.info('Запрос списка записей weld_repair_log')
    
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        search = request.args.get('search', '').strip()
        drawing_filter = request.args.get('drawing', '').strip()
        date_filter = request.args.get('date_filter', '').strip()
        
        # Получаем фильтры по столбцам
        column_filters = {}
        for key, value in request.args.items():
            if key.startswith('filter_') and value.strip():
                column_name = key[7:]  # Убираем префикс 'filter_'
                column_filters[column_name] = value.strip()
        
        logger.info(f'Параметры запроса: page={page}, per_page={per_page}, search="{search}", drawing="{drawing_filter}", date_filter="{date_filter}"')
        
        conn = get_db_connection()
        if not conn:
            logger.error('Не удалось подключиться к базе данных')
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weld_repair_log'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            logger.warning('Таблица weld_repair_log не существует')
            return jsonify({'success': False, 'message': 'Таблица weld_repair_log не существует'})
        
        # Базовый запрос
        base_query = 'FROM weld_repair_log WHERE 1=1'
        params = []
        
        # Добавляем фильтры
        if search:
            base_query += ''' AND (
                "Чертеж" LIKE ? OR 
                "Линия" LIKE ? OR 
                "№ стыка" LIKE ? OR 
                "Дата сварки" LIKE ? OR
                "Размер выборки (длина, ширина, глубина), мм" LIKE ? OR
                "Способ и результаты контроля выборки" LIKE ?
            )'''
            search_param = f'%{search}%'
            params.extend([search_param] * 6)
        
        if drawing_filter:
            base_query += ' AND "Чертеж" = ?'
            params.append(drawing_filter)
        
        if date_filter:
            if date_filter == 'today':
                base_query += " AND date('Дата сварки') = date('now')"
            elif date_filter == 'week':
                base_query += " AND 'Дата сварки' >= datetime('now', '-7 days')"
            elif date_filter == 'month':
                base_query += " AND 'Дата сварки' >= datetime('now', '-30 days')"
        
        # Добавляем фильтры по столбцам
        if column_filters:
            for column_name, filter_value in column_filters.items():
                # Проверяем, что столбец существует в запросе
                valid_columns = ['app_row_id', 'Чертеж', 'Линия', 'Диаметр и толщина стенки', '№ стыка', 
                               'Дата сварки', 'Размер выборки (длина, ширина, глубина), мм', 
                               'Способ и результаты контроля выборки', 'Марка стали', 
                               'Фамилия, инициалы, клеймо сварщика допустившего брак', 'Вид сварки', 
                               'Число ремонтов на одном участке', 'Дата сварки после ремонта', 'Ответственный']
                if column_name in valid_columns:
                    base_query += f' AND CAST("{column_name}" AS TEXT) LIKE ?'
                    params.append(f'%{filter_value}%')
        
        # Получаем общее количество записей
        cursor.execute(f'SELECT COUNT(*) {base_query}', params)
        total_records = cursor.fetchone()[0]
        
        # Вычисляем пагинацию
        total_pages = (total_records + per_page - 1) // per_page
        offset = (page - 1) * per_page
        
        # Получаем записи для текущей страницы
        query = f'''
            SELECT app_row_id, "Чертеж", "Линия", "Диаметр и толщина стенки", 
                   "№ стыка", "Дата сварки", 
                   "Размер выборки (длина, ширина, глубина), мм",
                   "Способ и результаты контроля выборки",
                   "Марка стали", "Фамилия, инициалы, клеймо сварщика допустившего брак",
                   "Вид сварки", "Число ремонтов на одном участке", 
                   "Дата сварки после ремонта", "Ответственный"
            {base_query}
            ORDER BY app_row_id DESC
            LIMIT ? OFFSET ?
        '''
        
        cursor.execute(query, params + [per_page, offset])
        records = []
        
        for row in cursor.fetchall():
            records.append({
                'app_row_id': row[0],
                'Чертеж': row[1],
                'Линия': row[2],
                'Диаметр и толщина стенки': row[3],
                '№ стыка': row[4],
                'Дата сварки': row[5],
                'Размер выборки (длина, ширина, глубина), мм': row[6],
                'Способ и результаты контроля выборки': row[7],
                'Марка стали': row[8],
                'Фамилия, инициалы, клеймо сварщика допустившего брак': row[9],
                'Вид сварки': row[10],
                'Число ремонтов на одном участке': row[11],
                'Дата сварки после ремонта': row[12],
                'Ответственный': row[13]
            })
        
        # Обновляем статистику
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log')
        total_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM weld_repair_log")
        recent_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT "Чертеж") FROM weld_repair_log WHERE "Чертеж" IS NOT NULL')
        unique_drawings = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT "№ стыка") FROM weld_repair_log WHERE "№ стыка" IS NOT NULL')
        unique_joints = cursor.fetchone()[0]
        
        stats = {
            'total_count': total_count,
            'recent_count': recent_count,
            'unique_drawings': unique_drawings,
            'unique_joints': unique_joints
        }
        
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_records,
            'pages': total_pages
        }
        
        conn.close()
        
        logger.info(f'Успешно возвращено {len(records)} записей из {total_records} всего')
        
        return jsonify({
            'success': True,
            'records': records,
            'pagination': pagination,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f'Ошибка API weld_repair_log: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/weld_repair_log/import', methods=['POST'])
def api_weld_repair_log_import():
    """Импорт Excel/CSV для массового обновления weld_repair_log.
    Матч по app_row_id, если нет - создает новую запись.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Файл не передан'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Файл не выбран'}), 400

        # Читаем файл
        stream = io.BytesIO(file.read())
        filename_lower = file.filename.lower()
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(stream)
        else:
            df = pd.read_excel(stream)
        if df.empty:
            return jsonify({'success': False, 'message': 'Файл пустой'}), 400

        # Нормализация названий столбцов
        df.columns = [str(c).strip() for c in df.columns]

        # Проверяем наличие ключевого поля.
        # Приоритет всегда у app_row_id, fallback на id только если app_row_id отсутствует.
        lower_to_original = {str(c).strip().lower(): c for c in df.columns}
        id_col = lower_to_original.get('app_row_id') or lower_to_original.get('id')
        if not id_col:
            return jsonify({'success': False, 'message': 'Нет поля app_row_id или id для сопоставления'}), 400

        # Разрешенные к обновлению поля
        updatable_fields = [
            'Чертеж', 'Линия', 'Диаметр и толщина стенки', '№ стыка', 'Дата сварки',
            'Размер выборки (длина, ширина, глубина), мм', 'Способ и результаты контроля выборки',
            'Марка стали', 'Фамилия, инициалы, клеймо сварщика допустившего брак', 'Вид сварки',
            'Число ремонтов на одном участке', 'Дата сварки после ремонта', 'Ответственный'
        ]

        # Определяем какие поля есть в файле
        present_fields = [f for f in updatable_fields if f in df.columns]

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'}), 500

        cursor = conn.cursor()
        processed = updated = created = skipped_invalid_id = 0

        for _, r in df.iterrows():
            processed += 1
            
            record_id = r.get(id_col)
            
            if pd.isna(record_id) or record_id == '':
                # Создаем новую запись
                insert_fields = []
                insert_values = []
                
                for field in present_fields:
                    value = r.get(field)
                    if isinstance(value, float) and pd.isna(value):
                        value = ''
                    insert_fields.append(field)
                    insert_values.append(str(value) if value is not None else '')
                
                # Добавляем обязательные поля с значениями по умолчанию
                insert_fields.extend(['created_at'])
                insert_values.extend([datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                
                # Создаем SQL для вставки
                placeholders = ', '.join(['?' for _ in insert_values])
                field_names = ', '.join([f'"{f}"' for f in insert_fields])
                insert_sql = f'INSERT INTO weld_repair_log ({field_names}) VALUES ({placeholders})'
                
                cursor.execute(insert_sql, tuple(insert_values))
                created += 1
            else:
                # app_row_id в БД текстовый, поэтому сравниваем как строку.
                record_id_str = str(record_id).strip()
                if not record_id_str:
                    skipped_invalid_id += 1
                    continue

                cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id_str,))
                if cursor.fetchone()[0] == 0:
                    continue
                
                # Формируем SQL для обновления
                update_parts = []
                update_values = []
                
                for field in present_fields:
                    value = r.get(field)
                    if isinstance(value, float) and pd.isna(value):
                        value = ''
                    update_parts.append(f'"{field}" = ?')
                    update_values.append(str(value) if value is not None else '')
                
                if update_parts:
                    update_values.append(record_id_str)
                    update_sql = f'UPDATE weld_repair_log SET {", ".join(update_parts)} WHERE app_row_id = ?'
                    cursor.execute(update_sql, tuple(update_values))
                    updated += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': (
                f'Импорт завершен. Обработано: {processed}, обновлено: {updated}, '
                f'создано: {created}, пропущено с пустым ID: {skipped_invalid_id}'
            ),
            'processed': processed,
            'updated': updated,
            'created': created,
            'skipped_invalid_id': skipped_invalid_id
        })

    except Exception as e:
        logger.error(f'Ошибка импорта weld_repair_log: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/weld_repair_log/responsibles')
def api_weld_repair_log_responsibles():
    """Возвращает уникальные непустые значения поля 'Ответственный'."""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'}), 500

        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT "Ответственный"
            FROM weld_repair_log
            WHERE "Ответственный" IS NOT NULL
              AND TRIM("Ответственный") != ''
            ORDER BY "Ответственный" COLLATE NOCASE
        ''')
        responsibles = [row[0] for row in cursor.fetchall() if row[0] is not None]
        conn.close()

        return jsonify({
            'success': True,
            'responsibles': responsibles
        })
    except Exception as e:
        logger.error(f'Ошибка получения списка ответственных: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/weld_repair_log/<int:record_id>')
def api_get_weld_repair_record(record_id):
    """API для получения одной записи weld_repair_log"""
    logger.info(f'Запрос записи weld_repair_log с ID: {record_id}')
    
    try:
        conn = get_db_connection()
        if not conn:
            logger.error('Не удалось подключиться к базе данных')
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weld_repair_log'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            logger.warning('Таблица weld_repair_log не существует')
            return jsonify({'success': False, 'message': 'Таблица weld_repair_log не существует'})
        
        # Проверяем количество записей в таблице
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log')
        total_count = cursor.fetchone()[0]
        logger.info(f'Всего записей в таблице weld_repair_log: {total_count}')
        
        cursor.execute('''
            SELECT app_row_id, "Чертеж", "Линия", "Диаметр и толщина стенки", 
                   "№ стыка", "Дата сварки", 
                   "Размер выборки (длина, ширина, глубина), мм",
                   "Способ и результаты контроля выборки",
                   "Марка стали", "Фамилия, инициалы, клеймо сварщика допустившего брак",
                   "Вид сварки", "Число ремонтов на одном участке", 
                   "Дата сварки после ремонта", "Ответственный"
            FROM weld_repair_log WHERE app_row_id = ?
        ''', (record_id,))
        
        row = cursor.fetchone()
        if not row:
            logger.warning(f'Запись с ID {record_id} не найдена')
            return jsonify({'success': False, 'message': f'Запись с ID {record_id} не найдена'})
        
        record = {
            'app_row_id': row[0],
            'Чертеж': row[1],
            'Линия': row[2],
            'Диаметр и толщина стенки': row[3],
            '№ стыка': row[4],
            'Дата сварки': row[5],
            'Размер выборки (длина, ширина, глубина), мм': row[6],
            'Способ и результаты контроля выборки': row[7],
            'Марка стали': row[8],
            'Фамилия, инициалы, клеймо сварщика допустившего брак': row[9],
            'Вид сварки': row[10],
            'Число ремонтов на одном участке': row[11],
            'Дата сварки после ремонта': row[12],
            'Ответственный': row[13]
        }
        
        logger.info(f'Успешно получена запись с ID {record_id}')
        conn.close()
        return jsonify({'success': True, 'record': record})
        
    except Exception as e:
        logger.error(f'Ошибка получения записи {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/weld_repair_log', methods=['POST'])
def api_create_weld_repair_record():
    """API для создания новой записи weld_repair_log"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Нет данных для создания записи'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Разрешенные для редактирования столбцы
        editable_columns = [
            "Марка стали",
            "Фамилия, инициалы, клеймо сварщика допустившего брак", 
            "Вид сварки",
            "Число ремонтов на одном участке",
            "Дата сварки после ремонта",
            "Ответственный"
        ]
        
        # Вставляем новую запись (только редактируемые поля)
        cursor.execute('''
            INSERT INTO weld_repair_log 
            ("Марка стали", "Фамилия, инициалы, клеймо сварщика допустившего брак",
             "Вид сварки", "Число ремонтов на одном участке", 
             "Дата сварки после ремонта", "Ответственный")
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data.get('Марка стали', ''),
            data.get('Фамилия, инициалы, клеймо сварщика допустившего брак', ''),
            data.get('Вид сварки', ''),
            data.get('Число ремонтов на одном участке', ''),
            data.get('Дата сварки после ремонта', ''),
            data.get('Ответственный', '')
        ))
        
        conn.commit()
        new_id = cursor.lastrowid
        conn.close()
        
        logger.info(f'Создана новая запись weld_repair_log с ID: {new_id}')
        return jsonify({'success': True, 'message': 'Запись успешно создана', 'record_id': new_id})
        
    except Exception as e:
        logger.error(f'Ошибка создания записи weld_repair_log: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/weld_repair_log/<int:record_id>', methods=['PUT'])
def api_update_weld_repair_record(record_id):
    """API для обновления записи weld_repair_log"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Нет данных для обновления записи'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование записи
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id,))
        if cursor.fetchone()[0] == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Обновляем запись (только редактируемые поля)
        cursor.execute('''
            UPDATE weld_repair_log SET 
            "Марка стали" = ?, 
            "Фамилия, инициалы, клеймо сварщика допустившего брак" = ?,
            "Вид сварки" = ?, 
            "Число ремонтов на одном участке" = ?,
            "Дата сварки после ремонта" = ?, 
            "Ответственный" = ?
            WHERE app_row_id = ?
        ''', (
            data.get('Марка стали', ''),
            data.get('Фамилия, инициалы, клеймо сварщика допустившего брак', ''),
            data.get('Вид сварки', ''),
            data.get('Число ремонтов на одном участке', ''),
            data.get('Дата сварки после ремонта', ''),
            data.get('Ответственный', ''),
            record_id
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f'Обновлена запись weld_repair_log с ID: {record_id}')
        return jsonify({'success': True, 'message': 'Запись успешно обновлена'})
        
    except Exception as e:
        logger.error(f'Ошибка обновления записи {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/weld_repair_log/<int:record_id>', methods=['DELETE'])
def api_delete_weld_repair_record(record_id):
    """API для удаления записи weld_repair_log"""
    conn = None
    try:
        logger.info(f'Попытка удаления записи weld_repair_log с ID: {record_id}')
        
        conn = get_db_connection()
        if not conn:
            logger.error('Не удалось подключиться к базе данных')
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование записи
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id,))
        record_exists = cursor.fetchone()[0]
        logger.info(f'Запись с ID {record_id} существует: {record_exists > 0}')
        
        if record_exists == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Удаляем запись
        logger.info(f'Выполняем DELETE для записи с ID: {record_id}')
        cursor.execute('DELETE FROM weld_repair_log WHERE app_row_id = ?', (record_id,))
        
        # Проверяем результат удаления
        deleted_count = cursor.rowcount
        logger.info(f'Результат удаления: {deleted_count} записей')
        
        if deleted_count == 0:
            logger.warning(f'DELETE не удалил ни одной записи для ID: {record_id}')
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': 'Не удалось удалить запись'})
        
        # Сохраняем изменения
        conn.commit()
        logger.info(f'Изменения сохранены в базе данных')
        
        # Проверяем, что запись действительно удалена
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id,))
        still_exists = cursor.fetchone()[0]
        logger.info(f'Запись после удаления все еще существует: {still_exists > 0}')
        
        conn.close()
        
        if still_exists > 0:
            logger.error(f'Запись с ID {record_id} не была удалена')
            return jsonify({'success': False, 'message': 'Ошибка при удалении записи'})
        
        logger.info(f'Успешно удалена запись weld_repair_log с ID: {record_id}')
        return jsonify({'success': True, 'message': 'Запись успешно удалена'})
        
    except sqlite3.OperationalError as e:
        logger.error(f'Ошибка SQL при удалении записи {record_id}: {e}')
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'message': f'Ошибка базы данных: {str(e)}'})
    except sqlite3.IntegrityError as e:
        logger.error(f'Ошибка целостности при удалении записи {record_id}: {e}')
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'message': f'Ошибка целостности данных: {str(e)}'})
    except Exception as e:
        logger.error(f'Неожиданная ошибка при удалении записи {record_id}: {e}')
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        return jsonify({'success': False, 'message': f'Ошибка сервера: {str(e)}'})

@app.route('/api/weld_repair_log/bulk_delete', methods=['POST'])
def api_bulk_delete_weld_repair_records():
    """API для массового удаления записей weld_repair_log"""
    conn = None
    try:
        data = request.get_json()
        if not data or 'record_ids' not in data:
            return jsonify({'success': False, 'message': 'Не указаны ID записей для удаления'})
        
        record_ids = data['record_ids']
        if not record_ids:
            return jsonify({'success': False, 'message': 'Список ID пуст'})
        
        logger.info(f'Попытка массового удаления {len(record_ids)} записей: {record_ids}')
        
        conn = get_db_connection()
        if not conn:
            logger.error('Не удалось подключиться к базе данных')
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем, сколько записей существует
        placeholders = ','.join(['?' for _ in record_ids])
        cursor.execute(f'SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id IN ({placeholders})', record_ids)
        existing_count = cursor.fetchone()[0]
        logger.info(f'Найдено существующих записей: {existing_count} из {len(record_ids)}')
        
        if existing_count == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Не найдено записей для удаления'})
        
        # Удаляем записи
        cursor.execute(f'DELETE FROM weld_repair_log WHERE app_row_id IN ({placeholders})', record_ids)
        
        deleted_count = cursor.rowcount
        logger.info(f'Результат удаления: {deleted_count} записей')
        
        if deleted_count == 0:
            logger.warning('DELETE не удалил ни одной записи')
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': 'Не удалось удалить записи'})
        
        # Сохраняем изменения
        conn.commit()
        logger.info(f'Изменения сохранены в базе данных')
        
        conn.close()
        
        logger.info(f'Успешно удалено {deleted_count} записей weld_repair_log из {len(record_ids)} запрошенных')
        return jsonify({
            'success': True, 
            'message': f'Удалено записей: {deleted_count} из {len(record_ids)}', 
            'deleted_count': deleted_count,
            'requested_count': len(record_ids)
        })
        
    except sqlite3.OperationalError as e:
        logger.error(f'Ошибка SQL при массовом удалении: {e}')
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'message': f'Ошибка базы данных: {str(e)}'})
    except sqlite3.IntegrityError as e:
        logger.error(f'Ошибка целостности при массовом удалении: {e}')
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'message': f'Ошибка целостности данных: {str(e)}'})
    except Exception as e:
        logger.error(f'Неожиданная ошибка при массовом удалении: {e}')
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        return jsonify({'success': False, 'message': f'Ошибка сервера: {str(e)}'})

@app.route('/api/weld_repair_log/export')
def api_export_weld_repair_log():
    """API для экспорта weld_repair_log в Excel с учетом фильтров"""
    try:
        # Получаем параметры фильтров
        search = request.args.get('search', '')
        drawing = request.args.get('drawing', '')
        date_filter = request.args.get('date_filter', '')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Формируем SQL запрос с фильтрами
        query = '''
            SELECT 
                app_row_id,
                "Чертеж",
                "Линия",
                "Диаметр и толщина стенки",
                "№ стыка",
                "Дата сварки",
                "Размер выборки (длина, ширина, глубина), мм",
                "Способ и результаты контроля выборки",
                "Марка стали",
                "Фамилия, инициалы, клеймо сварщика допустившего брак",
                "Вид сварки",
                "Число ремонтов на одном участке",
                "Дата сварки после ремонта",
                "Ответственный"
            FROM weld_repair_log
            WHERE 1=1
        '''
        
        params = []
        
        # Добавляем фильтр поиска
        if search:
            query += ''' AND (
                "Чертеж" LIKE ? OR 
                "Линия" LIKE ? OR 
                "№ стыка" LIKE ? OR 
                "Марка стали" LIKE ? OR 
                "Фамилия, инициалы, клеймо сварщика допустившего брак" LIKE ? OR 
                "Вид сварки" LIKE ? OR 
                "Ответственный" LIKE ?
            )'''
            search_param = f'%{search}%'
            params.extend([search_param] * 7)
        
        # Добавляем фильтр по чертежу
        if drawing:
            query += ' AND "Чертеж" = ?'
            params.append(drawing)
        
        # Добавляем фильтр по дате
        if date_filter:
            if date_filter == 'today':
                query += ' AND DATE("Дата сварки") = DATE("now")'
            elif date_filter == 'week':
                query += ' AND DATE("Дата сварки") >= DATE("now", "-7 days")'
            elif date_filter == 'month':
                query += ' AND DATE("Дата сварки") >= DATE("now", "-1 month")'
        
        query += ' ORDER BY app_row_id DESC'
        
        # Выполняем запрос
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # Получаем названия колонок
        columns = [description[0] for description in cursor.description]
        
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'Нет данных для экспорта'})
        
        # Создаем DataFrame
        df = pd.DataFrame(records, columns=columns)
        
        # Переименовываем колонки для лучшей читаемости
        column_mapping = {
            'app_row_id': 'ID',
            'Чертеж': 'Чертеж',
            'Линия': 'Линия',
            'Диаметр и толщина стенки': 'Диаметр и толщина стенки',
            '№ стыка': '№ стыка',
            'Дата сварки': 'Дата сварки',
            'Размер выборки (длина, ширина, глубина), мм': 'Размер выборки (длина, ширина, глубина), мм',
            'Способ и результаты контроля выборки': 'Способ и результаты контроля выборки',
            'Марка стали': 'Марка стали',
            'Фамилия, инициалы, клеймо сварщика допустившего брак': 'Фамилия, инициалы, клеймо сварщика допустившего брак',
            'Вид сварки': 'Вид сварки',
            'Число ремонтов на одном участке': 'Число ремонтов на одном участке',
            'Дата сварки после ремонта': 'Дата сварки после ремонта',
            'Ответственный': 'Ответственный'
        }
        
        df = df.rename(columns=column_mapping)
        
        # Очищаем данные от проблемных символов и NaN значений
        df = clean_data_for_excel(df)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Журнал ремонта сварных швов', index=False)
            
            # Получаем рабочий лист для форматирования
            worksheet = writer.sheets['Журнал ремонта сварных швов']
            
            # Автоматически подгоняем ширину колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Формируем имя файла с датой
        filename = f'weld_repair_log_{get_filename_timestamp()}.xlsx'
        
        logger.info(f'Экспортировано {len(records)} записей weld_repair_log в Excel')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f'Ошибка экспорта weld_repair_log: {e}')
        return jsonify({'success': False, 'message': str(e)})













@app.route('/nk_results')
def nk_results():
    """Страница для внесения результатов НК"""
    try:
        conn = get_db_connection()
        if not conn:
            return "Ошибка подключения к базе данных", 500
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы weld_repair_log
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weld_repair_log'")
        if not cursor.fetchone():
            conn.close()
            return "Таблица weld_repair_log не найдена", 404
        
        # Добавляем новые столбцы для НК, если их нет
        cursor.execute("PRAGMA table_info(weld_repair_log)")
        existing_columns = [col[1] for col in cursor.fetchall()]
        
        new_nk_columns = [
            "Дата заявки о качестве ремонта",
            "Статус после ремонта", 
            "Рез ФАКТ"
        ]
        
        for column in new_nk_columns:
            if column not in existing_columns:
                try:
                    cursor.execute(f'ALTER TABLE weld_repair_log ADD COLUMN "{column}" TEXT')
                    logger.info(f"Добавлен столбец НК: {column}")
                except Exception as e:
                    logger.warning(f"Ошибка добавления столбца НК {column}: {e}")
        
        # Получаем статистику
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log')
        total_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE "Рез ФАКТ" IS NOT NULL AND "Рез ФАКТ" != ""')
        nk_completed = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE "Рез ФАКТ" IS NULL OR "Рез ФАКТ" = ""')
        nk_pending = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT "Чертеж") FROM weld_repair_log WHERE "Чертеж" IS NOT NULL')
        unique_drawings = cursor.fetchone()[0]
        
        # Получаем уникальные чертежи для фильтра
        cursor.execute('SELECT DISTINCT "Чертеж" FROM weld_repair_log WHERE "Чертеж" IS NOT NULL ORDER BY "Чертеж"')
        unique_drawings_list = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        stats = {
            'total_count': total_count,
            'nk_completed': nk_completed,
            'nk_pending': nk_pending,
            'unique_drawings': unique_drawings
        }
        
        return render_template('nk_results.html', 
                             stats=stats, 
                             unique_drawings=unique_drawings_list)
        
    except Exception as e:
        logger.error(f'Ошибка загрузки страницы НК: {e}')
        return f"Ошибка: {str(e)}", 500

@app.route('/api/nk_results')
def api_nk_results():
    """API для получения записей с результатами НК"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        search = request.args.get('search', '')
        drawing = request.args.get('drawing', '')
        status_filter = request.args.get('status_filter', '')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Формируем SQL запрос с фильтрами
        query = '''
            SELECT 
                app_row_id,
                "Чертеж",
                "Линия",
                "Диаметр и толщина стенки",
                "№ стыка",
                "Дата сварки",
                "Марка стали",
                "Вид сварки",
                "Число ремонтов на одном участке",
                "Дата сварки после ремонта",
                "Ответственный",
                "Дата заявки о качестве ремонта",
                "Статус после ремонта",
                "Рез ФАКТ"
            FROM weld_repair_log
            WHERE 1=1
        '''
        
        params = []
        
        # Добавляем фильтр поиска
        if search:
            query += ''' AND (
                "Чертеж" LIKE ? OR 
                "Линия" LIKE ? OR 
                "№ стыка" LIKE ? OR 
                "Марка стали" LIKE ? OR 
                "Вид сварки" LIKE ? OR 
                "Ответственный" LIKE ? OR
                "Статус после ремонта" LIKE ? OR
                "Рез ФАКТ" LIKE ?
            )'''
            search_param = f'%{search}%'
            params.extend([search_param] * 8)
        
        # Добавляем фильтр по чертежу
        if drawing:
            query += ' AND "Чертеж" = ?'
            params.append(drawing)
        
        # Добавляем фильтр по статусу НК
        if status_filter:
            if status_filter == 'completed':
                query += ' AND ("Рез ФАКТ" IS NOT NULL AND "Рез ФАКТ" != "")'
            elif status_filter == 'pending':
                query += ' AND ("Рез ФАКТ" IS NULL OR "Рез ФАКТ" = "")'
        
        # Получаем общее количество записей
        count_query = '''
            SELECT COUNT(*)
            FROM weld_repair_log
            WHERE 1=1
        '''
        
        count_params = []
        
        # Добавляем те же фильтры для подсчета
        if search:
            count_query += ''' AND (
                "Чертеж" LIKE ? OR 
                "Линия" LIKE ? OR 
                "№ стыка" LIKE ? OR 
                "Марка стали" LIKE ? OR 
                "Вид сварки" LIKE ? OR 
                "Ответственный" LIKE ? OR
                "Статус после ремонта" LIKE ? OR
                "Рез ФАКТ" LIKE ?
            )'''
            search_param = f'%{search}%'
            count_params.extend([search_param] * 8)
        
        if drawing:
            count_query += ' AND "Чертеж" = ?'
            count_params.append(drawing)
        
        if status_filter:
            if status_filter == 'completed':
                count_query += ' AND ("Рез ФАКТ" IS NOT NULL AND "Рез ФАКТ" != "")'
            elif status_filter == 'pending':
                count_query += ' AND ("Рез ФАКТ" IS NULL OR "Рез ФАКТ" = "")'
        
        cursor.execute(count_query, count_params)
        total = cursor.fetchone()[0]
        
        # Добавляем пагинацию и сортировку
        query += ' ORDER BY app_row_id DESC LIMIT ? OFFSET ?'
        params.extend([per_page, (page - 1) * per_page])
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # Получаем названия колонок
        columns = [description[0] for description in cursor.description]
        
        # Преобразуем в список словарей
        records_list = []
        for record in records:
            record_dict = {}
            for i, column in enumerate(columns):
                record_dict[column] = record[i]
            records_list.append(record_dict)
        
        # Статистика
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log')
        total_count = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE "Рез ФАКТ" IS NOT NULL AND "Рез ФАКТ" != ""')
        nk_completed = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE "Рез ФАКТ" IS NULL OR "Рез ФАКТ" = ""')
        nk_pending = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(DISTINCT "Чертеж") FROM weld_repair_log WHERE "Чертеж" IS NOT NULL')
        unique_drawings = cursor.fetchone()[0]
        
        conn.close()
        
        # Пагинация
        pages = (total + per_page - 1) // per_page
        
        pagination = {
            'page': page,
            'pages': pages,
            'per_page': per_page,
            'total': total
        }
        
        stats = {
            'total_count': total_count,
            'nk_completed': nk_completed,
            'nk_pending': nk_pending,
            'unique_drawings': unique_drawings
        }
        
        return jsonify({
            'success': True,
            'records': records_list,
            'pagination': pagination,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f'Ошибка получения записей НК: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/nk_results/<int:record_id>')
def api_get_nk_record(record_id):
    """API для получения одной записи НК"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                app_row_id,
                "Чертеж",
                "Линия",
                "Диаметр и толщина стенки",
                "№ стыка",
                "Дата сварки",
                "Марка стали",
                "Вид сварки",
                "Число ремонтов на одном участке",
                "Дата сварки после ремонта",
                "Ответственный",
                "Дата заявки о качестве ремонта",
                "Статус после ремонта",
                "Рез ФАКТ"
            FROM weld_repair_log 
            WHERE app_row_id = ?
        ''', (record_id,))
        
        record = cursor.fetchone()
        conn.close()
        
        if not record:
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Преобразуем в словарь
        columns = [description[0] for description in cursor.description]
        record_dict = {}
        for i, column in enumerate(columns):
            record_dict[column] = record[i]
        
        return jsonify({'success': True, 'record': record_dict})
        
    except Exception as e:
        logger.error(f'Ошибка получения записи НК {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/nk_results/<int:record_id>', methods=['PUT'])
def api_update_nk_record(record_id):
    """API для обновления записи НК"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Данные не получены'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование записи
        cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id,))
        if cursor.fetchone()[0] == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Обновляем только поля НК
        cursor.execute('''
            UPDATE weld_repair_log SET 
            "Дата заявки о качестве ремонта" = ?, 
            "Статус после ремонта" = ?,
            "Рез ФАКТ" = ?
            WHERE app_row_id = ?
        ''', (
            data.get('Дата заявки о качестве ремонта', ''),
            data.get('Статус после ремонта', ''),
            data.get('Рез ФАКТ', ''),
            record_id
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f'Обновлена запись НК с ID: {record_id}')
        return jsonify({'success': True, 'message': 'Результаты НК успешно обновлены'})
        
    except Exception as e:
        logger.error(f'Ошибка обновления записи НК {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/nk_results/export')
def api_export_nk_results():
    """API для экспорта результатов НК в Excel"""
    try:
        # Получаем параметры фильтров
        search = request.args.get('search', '')
        drawing = request.args.get('drawing', '')
        status_filter = request.args.get('status_filter', '')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Формируем SQL запрос с фильтрами
        query = '''
            SELECT 
                app_row_id,
                "Чертеж",
                "Линия",
                "Диаметр и толщина стенки",
                "№ стыка",
                "Дата сварки",
                "Марка стали",
                "Вид сварки",
                "Число ремонтов на одном участке",
                "Дата сварки после ремонта",
                "Ответственный",
                "Дата заявки о качестве ремонта",
                "Статус после ремонта",
                "Рез ФАКТ"
            FROM weld_repair_log
            WHERE 1=1
        '''
        
        params = []
        
        # Добавляем фильтр поиска
        if search:
            query += ''' AND (
                "Чертеж" LIKE ? OR 
                "Линия" LIKE ? OR 
                "№ стыка" LIKE ? OR 
                "Марка стали" LIKE ? OR 
                "Вид сварки" LIKE ? OR 
                "Ответственный" LIKE ? OR
                "Статус после ремонта" LIKE ? OR
                "Рез ФАКТ" LIKE ?
            )'''
            search_param = f'%{search}%'
            params.extend([search_param] * 8)
        
        # Добавляем фильтр по чертежу
        if drawing:
            query += ' AND "Чертеж" = ?'
            params.append(drawing)
        
        # Добавляем фильтр по статусу НК
        if status_filter:
            if status_filter == 'completed':
                query += ' AND ("Рез ФАКТ" IS NOT NULL AND "Рез ФАКТ" != "")'
            elif status_filter == 'pending':
                query += ' AND ("Рез ФАКТ" IS NULL OR "Рез ФАКТ" = "")'
        
        query += ' ORDER BY app_row_id DESC'
        
        # Выполняем запрос
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # Получаем названия колонок
        columns = [description[0] for description in cursor.description]
        
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'message': 'Нет данных для экспорта'})
        
        # Создаем DataFrame
        df = pd.DataFrame(records, columns=columns)
        
        # Переименовываем колонки для лучшей читаемости
        column_mapping = {
            'app_row_id': 'ID',
            'Чертеж': 'Чертеж',
            'Линия': 'Линия',
            'Диаметр и толщина стенки': 'Диаметр и толщина стенки',
            '№ стыка': '№ стыка',
            'Дата сварки': 'Дата сварки',
            'Марка стали': 'Марка стали',
            'Вид сварки': 'Вид сварки',
            'Число ремонтов на одном участке': 'Число ремонтов на одном участке',
            'Дата сварки после ремонта': 'Дата сварки после ремонта',
            'Ответственный': 'Ответственный',
            'Дата заявки о качестве ремонта': 'Дата заявки о качестве ремонта',
            'Статус после ремонта': 'Статус после ремонта',
            'Рез ФАКТ': 'Рез ФАКТ'
        }
        
        df = df.rename(columns=column_mapping)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Результаты НК', index=False)
            
            # Получаем рабочий лист для форматирования
            worksheet = writer.sheets['Результаты НК']
            
            # Автоматически подгоняем ширину колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Формируем имя файла с датой
        filename = f'nk_results_{get_filename_timestamp()}.xlsx'
        
        logger.info(f'Экспортировано {len(records)} записей результатов НК в Excel')
        
        # Логируем активность экспорта
        log_activity('Экспорт данных', f'Экспортировано {len(records)} записей НК в {filename}', 'download')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f'Ошибка экспорта результатов НК: {e}')
        return jsonify({'success': False, 'message': str(e)})



@app.route('/api/nk_results/import', methods=['POST'])
def api_import_nk_results():
    """Импорт результатов НК из Excel/CSV. Обновление по app_row_id (ID)."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Файл не передан'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Файл не выбран'}), 400

        filename_lower = file.filename.lower()
        if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls') or filename_lower.endswith('.csv')):
            return jsonify({'success': False, 'message': 'Поддерживаются только .xlsx, .xls, .csv'}), 400

        # Читаем файл в DataFrame
        file_stream = io.BytesIO(file.read())
        if filename_lower.endswith('.csv'):
            df = pd.read_csv(file_stream)
        else:
            df = pd.read_excel(file_stream)

        if df.empty:
            return jsonify({'success': False, 'message': 'Файл пустой'}), 400

        # Приводим названия столбцов
        normalized_cols = {str(c).strip(): str(c).strip() for c in df.columns}
        df = df.rename(columns=normalized_cols)

        # Поддерживаем варианты названия ID.
        # Приоритет у app_row_id, затем fallback на ID/id.
        id_column = None
        for candidate in ['app_row_id', 'App_Row_Id', 'AppRowId', 'ID', 'id']:
            if candidate in df.columns:
                id_column = candidate
                break

        if not id_column:
            return jsonify({'success': False, 'message': 'В файле нет столбца ID/app_row_id'}), 400

        # Оставляем только интересующие поля
        updatable_fields = [
            'Дата заявки о качестве ремонта',
            'Статус после ремонта',
            'Рез ФАКТ'
        ]
        present_fields = [c for c in updatable_fields if c in df.columns]
        if not present_fields:
            return jsonify({'success': False, 'message': 'В файле нет столбцов для обновления'}), 400

        # Нормализуем формат даты к YYYY-MM-DD как в БД (input type=date)
        if 'Дата заявки о качестве ремонта' in present_fields:
            try:
                # Пробуем распарсить разные форматы (в т.ч. DD.MM.YYYY, Excel-даты)
                parsed = pd.to_datetime(df['Дата заявки о качестве ремонта'], errors='coerce', dayfirst=True)
                df['Дата заявки о качестве ремонта'] = parsed.dt.strftime('%Y-%m-%d')
                df['Дата заявки о качестве ремонта'] = df['Дата заявки о качестве ремонта'].fillna('')
            except Exception as _e:
                # Если парсинг не удался, оставляем как есть; NaN обработаем ниже
                pass

        # Соединение с БД
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'}), 500
        cursor = conn.cursor()

        processed = 0
        updated = 0

        # Обновляем построчно
        for _, row in df.iterrows():
            processed += 1
            record_id = row.get(id_column)
            if pd.isna(record_id):
                continue
            record_id_str = str(record_id).strip()
            if not record_id_str:
                continue

            cursor.execute('SELECT COUNT(*) FROM weld_repair_log WHERE app_row_id = ?', (record_id_str,))
            if cursor.fetchone()[0] == 0:
                continue

            set_parts = []
            values = []
            for field in present_fields:
                value = row.get(field)
                # Преобразуем NaN в пустую строку
                if isinstance(value, float) and pd.isna(value):
                    value = ''
                set_parts.append(f'"{field}" = ?')
                values.append(value)

            if set_parts:
                values.append(record_id_str)
                sql = f'UPDATE weld_repair_log SET ' + ', '.join(set_parts) + ' WHERE app_row_id = ?'
                cursor.execute(sql, tuple(values))
                updated += 1

        conn.commit()
        conn.close()

        logger.info(f'Импорт НК: обработано {processed}, обновлено {updated}')
        log_activity('Импорт данных', f'Импорт НК: обработано {processed}, обновлено {updated}', 'upload')

        return jsonify({'success': True, 'processed': int(processed), 'updated': int(updated)})

    except Exception as e:
        logger.error(f'Ошибка импорта результатов НК: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/api/shutdown', methods=['POST'])
def shutdown_server():
    """Останавливает сервер"""
    try:
        logger.info('Получен запрос на остановку сервера')
        # Отправляем сигнал завершения
        os._exit(0)
        return jsonify({'success': True, 'message': 'Сервер остановлен'})
    except Exception as e:
        logger.error(f'Ошибка остановки сервера: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/logs_lnk_table')
def logs_lnk_table():
    """Страница с таблицей logs_lnk с возможностью управления столбцами"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # Ограничиваем per_page разумными значениями
    if per_page < 10:
        per_page = 10
    elif per_page > 1000:
        per_page = 1000
    
    offset = (page - 1) * per_page
    search_term = request.args.get('search', '')
    visible_columns = request.args.get('columns', '').split(',') if request.args.get('columns') else []
    
    # Получаем фильтры по столбцам
    column_filters = {}
    column_filter_types = {}
    for key, value in request.args.items():
        if key.startswith('filter_') and not key.startswith('filter_type_') and value.strip():
            column_name = key[7:]  # Убираем префикс 'filter_'
            column_filters[column_name] = value.strip()
            
            # Получаем тип фильтра
            filter_type_key = f'filter_type_{column_name}'
            filter_type = request.args.get(filter_type_key, 'contains')
            column_filter_types[column_name] = filter_type
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('index'))
    
    try:
        cursor = conn.cursor()
        
        # Получаем все столбцы таблицы logs_lnk
        cursor.execute("PRAGMA table_info(logs_lnk)")
        all_columns = [col[1] for col in cursor.fetchall()]
        
        # Определяем столбцы для отображения
        if not visible_columns or visible_columns == ['']:
            # По умолчанию показываем основные столбцы
            default_columns = [
                'app_row_id', 'Чертеж', 'Линия', 'Лист', 'Номер_стыка', 
                'Диаметр_1', 'Толщина_1', 'Дата_сварки', 'ВИК', 'Статус_ВИК', 
                'РК', 'Статус_РК', 'Титул', 'Зона'
            ]
            visible_columns = [col for col in default_columns if col in all_columns]
        else:
            # Фильтруем только существующие столбцы и убираем пустые значения
            visible_columns = [col for col in visible_columns if col in all_columns and col.strip()]
        
        # Если после фильтрации не осталось столбцов, используем значения по умолчанию
        if not visible_columns:
            default_columns = [
                'app_row_id', 'Чертеж', 'Линия', 'Лист', 'Номер_стыка', 
                'Диаметр_1', 'Толщина_1', 'Дата_сварки', 'ВИК', 'Статус_ВИК', 
                'РК', 'Статус_РК', 'Титул', 'Зона'
            ]
            visible_columns = [col for col in default_columns if col in all_columns]
        
        # Формируем условие поиска и фильтрации
        where_conditions = []
        search_params = []
        
        # Добавляем общий поиск
        if search_term:
            # Поиск по всем видимым столбцам (регистронезависимый)
            search_conditions = []
            print(f"[DEBUG] logs_lnk - Поисковый термин: '{search_term}'")
            for col in visible_columns:
                # Используем двойной поиск: оригинальный термин + нижний регистр
                search_conditions.append(f'(CAST("{col}" AS TEXT) LIKE ? OR LOWER(CAST("{col}" AS TEXT)) LIKE ?)')
                search_params.append(f'%{search_term}%')
                search_params.append(f'%{search_term.lower()}%')
            where_conditions.append(f"({' OR '.join(search_conditions)})")
            print(f"[DEBUG] logs_lnk - Условие поиска: {' OR '.join(search_conditions)}")
        
        # Добавляем фильтры по столбцам
        if column_filters:
            print(f"[DEBUG] logs_lnk - Фильтры по столбцам: {column_filters}")
            print(f"[DEBUG] logs_lnk - Типы фильтров: {column_filter_types}")
            for column_name, filter_value in column_filters.items():
                if column_name in all_columns:  # Проверяем, что столбец существует
                    filter_type = column_filter_types.get(column_name, 'contains')
                    
                    if filter_type == 'contains':
                        # Регистронезависимый поиск по конкретному столбцу
                        where_conditions.append(f'(CAST("{column_name}" AS TEXT) LIKE ? OR LOWER(CAST("{column_name}" AS TEXT)) LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'not_contains':
                        # Исключение записей, содержащих значение
                        where_conditions.append(f'(CAST("{column_name}" AS TEXT) NOT LIKE ? AND LOWER(CAST("{column_name}" AS TEXT)) NOT LIKE ?)')
                        search_params.append(f'%{filter_value}%')
                        search_params.append(f'%{filter_value.lower()}%')
                    elif filter_type == 'empty':
                        # Показывать только пустые значения (NULL, пустая строка, только пробелы, "-", "None")
                        where_conditions.append(f'("{column_name}" IS NULL OR CAST("{column_name}" AS TEXT) = \'\' OR TRIM(CAST("{column_name}" AS TEXT)) = \'\' OR CAST("{column_name}" AS TEXT) = \'-\' OR CAST("{column_name}" AS TEXT) = \'None\')')
                    elif filter_type == 'not_empty':
                        # Показывать только непустые значения (не NULL, не пустая строка, не только пробелы, не "-", не "None")
                        where_conditions.append(f'("{column_name}" IS NOT NULL AND CAST("{column_name}" AS TEXT) != \'\' AND TRIM(CAST("{column_name}" AS TEXT)) != \'\' AND CAST("{column_name}" AS TEXT) != \'-\' AND CAST("{column_name}" AS TEXT) != \'None\')')
                    
                    print(f"[DEBUG] logs_lnk - Добавлен фильтр для столбца '{column_name}': '{filter_value}' (тип: {filter_type})")
        
        # Формируем итоговое условие WHERE
        if where_conditions:
            search_condition = f"WHERE {' AND '.join(where_conditions)}"
        else:
            search_condition = ""
        
        print(f"[DEBUG] logs_lnk - Итоговое SQL условие: {search_condition}")
        print(f"[DEBUG] logs_lnk - Параметры: {search_params}")
        
        # Получаем общее количество записей
        count_query = f"SELECT COUNT(*) as count FROM logs_lnk {search_condition}"
        cursor.execute(count_query, search_params)
        total_records = cursor.fetchone()['count']
        
        # Получаем данные с пагинацией
        columns_str = ', '.join([f'"{col}"' for col in visible_columns])
        data_query = f"""
            SELECT {columns_str}
            FROM logs_lnk 
            {search_condition}
            ORDER BY "Дата_загрузки" DESC, "Чертеж", "Номер_стыка"
            LIMIT {per_page} OFFSET {offset}
        """
        cursor.execute(data_query, search_params)
        records = cursor.fetchall()
        
        # Преобразуем в список словарей
        records_list = []
        for record in records:
            record_dict = {}
            for i, col in enumerate(visible_columns):
                record_dict[col] = record[i]
            records_list.append(record_dict)
        
        total_pages = (total_records + per_page - 1) // per_page
        
        conn.close()
        
        return render_template('logs_lnk_table.html',
                             records=records_list,
                             columns=visible_columns,
                             all_columns=all_columns,
                             page=page,
                             per_page=per_page,
                             total_pages=total_pages,
                             total_records=total_records,
                             search_term=search_term,
                             visible_columns=visible_columns,
                             column_filters=column_filters,
                             column_filter_types=column_filter_types)
        
    except Exception as e:
        logger.error(f'Ошибка при загрузке таблицы logs_lnk: {e}')
        flash(f'Ошибка при загрузке данных: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/backup', methods=['POST'])
def create_backup():
    """Создание резервной копии системы"""
    try:
        json_data = request.json or {}
        backup_type = json_data.get('type', 'critical')
        backup_name = json_data.get('name', None)
        
        # Добавляем путь к backup_system.py в sys.path
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        backup_system_path = os.path.join(project_root, 'backup_system.py')
        
        if not os.path.exists(backup_system_path):
            return jsonify({'success': False, 'message': 'Система резервного копирования не найдена'})
        
        # Импортируем систему бэкапа
        import sys
        sys.path.insert(0, project_root)
        
        try:
            from backup_system import BackupSystem
        except ImportError as e:
            return jsonify({'success': False, 'message': f'Ошибка импорта системы бэкапа: {str(e)}'})
        
        # Создаем экземпляр системы бэкапа
        backup_system = BackupSystem(project_root)
        
        # Создаем бэкап в зависимости от типа
        if backup_type == 'full':
            backup_path = backup_system.create_full_backup(backup_name or '')
        elif backup_type == 'database':
            backup_path = backup_system.create_database_backup(backup_name or '')
        else:  # critical
            backup_path = backup_system.create_critical_backup(backup_name or '')
        
        if backup_path:
            # Получаем размер файла
            file_size = os.path.getsize(backup_path)
            size_mb = file_size / (1024 * 1024)
            
            # Логируем активность
            log_activity('Резервное копирование', f'Создан {backup_type} бэкап: {os.path.basename(backup_path)} ({size_mb:.1f} MB)', 'save')
            
            return jsonify({
                'success': True, 
                'message': f'Резервная копия создана успешно',
                'backup_path': backup_path,
                'backup_name': os.path.basename(backup_path),
                'size_mb': round(size_mb, 1),
                'type': backup_type
            })
        else:
            return jsonify({'success': False, 'message': 'Не удалось создать резервную копию'})
            
    except Exception as e:
        logger.error(f'Ошибка создания резервной копии: {e}')
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

@app.route('/api/backups', methods=['GET'])
def list_backups():
    """Получение списка доступных резервных копий"""
    try:
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        backup_system_path = os.path.join(project_root, 'backup_system.py')
        
        if not os.path.exists(backup_system_path):
            return jsonify({'success': False, 'message': 'Система резервного копирования не найдена'})
        
        import sys
        sys.path.insert(0, project_root)
        
        try:
            from backup_system import BackupSystem
        except ImportError as e:
            return jsonify({'success': False, 'message': f'Ошибка импорта системы бэкапа: {str(e)}'})
        
        backup_system = BackupSystem(project_root)
        backups = backup_system.list_backups()
        
        return jsonify({'success': True, 'backups': backups})
        
    except Exception as e:
        logger.error(f'Ошибка получения списка бэкапов: {e}')
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

@app.route('/id_formation')
def id_formation_page():
    """Страница формирования ИД"""
    return render_template('id_formation.html')

@app.route('/folder_formation')
def folder_formation_page():
    """Страница формирования ИД в папку"""
    return render_template('folder_formation.html')

@app.route('/backups/<filename>')
def download_backup(filename):
    """Скачивание резервной копии"""
    try:
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        backup_path = os.path.join(project_root, 'backups', filename)
        
        if not os.path.exists(backup_path):
            flash('Файл резервной копии не найден', 'error')
            return redirect(url_for('index'))
        
        # Логируем скачивание
        log_activity('Скачивание бэкапа', f'Скачан файл: {filename}', 'download')
        
        return send_file(backup_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f'Ошибка скачивания бэкапа {filename}: {e}')
        flash(f'Ошибка скачивания: {str(e)}', 'error')
        return redirect(url_for('index'))



@app.route('/api/backup/<filename>', methods=['DELETE'])
def delete_backup(filename):
    """Удаление резервной копии"""
    try:
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        backup_path = os.path.join(project_root, 'backups', filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'Файл не найден'})
        
        # Удаляем файл и метаданные
        os.remove(backup_path)
        
        # Удаляем JSON файл метаданных если есть
        metadata_path = backup_path.replace('.zip', '.json')
        if os.path.exists(metadata_path):
            os.remove(metadata_path)
        
        # Логируем удаление
        log_activity('Удаление бэкапа', f'Удален файл: {filename}', 'trash')
        
        return jsonify({'success': True, 'message': 'Бэкап удален успешно'})
        
    except Exception as e:
        logger.error(f'Ошибка удаления бэкапа {filename}: {e}')
        return jsonify({'success': False, 'message': f'Ошибка удаления: {str(e)}'})

@app.route('/enhanced_filter')
def enhanced_filter():
    """Улучшенная страница фильтрации с реальными данными"""
    
    # Получаем параметры фильтрации
    status_filter = request.args.get('status', 'all')
    title_filter = request.args.get('title', 'all')
    line_filter = request.args.get('line', 'all')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    if not conn:
        flash('Ошибка подключения к базе данных', 'error')
        return redirect(url_for('index'))
    
    try:
        cursor = conn.cursor()
        
        # Получаем уникальные значения для фильтров
        cursor.execute('SELECT DISTINCT "Статус_РК" FROM condition_weld WHERE "Статус_РК" IS NOT NULL AND "Статус_РК" != "None" ORDER BY "Статус_РК"')
        available_statuses = [row[0] for row in cursor.fetchall()]
        
        # Используем новую функцию для получения частей титулов
        available_titles = get_title_parts_list()
        
        cursor.execute('SELECT DISTINCT "Линия" FROM condition_weld WHERE "Линия" IS NOT NULL AND "Линия" != "None" ORDER BY "Линия"')
        available_lines = [row[0] for row in cursor.fetchall()]
        
        # Формируем условия фильтрации
        conditions = []
        params = []
        
        if status_filter and status_filter != 'all':
            conditions.append('"Статус_РК" = ?')
            params.append(status_filter)
        
        if title_filter and title_filter != 'all':
            conditions.append('"Титул" LIKE ?')
            params.append(f'%{title_filter}%')
        
        if line_filter and line_filter != 'all':
            conditions.append('"Линия" = ?')
            params.append(line_filter)
        
        # Добавляем условие исключения записей где Код_удаления содержит 'R' (оставляем NULL и пустые)
        conditions.append('("Код_удаления" IS NULL OR "Код_удаления" = "" OR "Код_удаления" = "None" OR "Код_удаления" NOT LIKE "%R%")')
        
        # Формируем WHERE условие
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        # Получаем общее количество записей
        count_query = f'SELECT COUNT(*) FROM condition_weld WHERE {where_clause}'
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]
        
        # Получаем отфильтрованные данные
        data_query = f'''
            SELECT 
                "id",
                "ISO", 
                "Линия",
                "Титул",
                "стык",
                "Статус_РК",
                "Статус_ВИК",
                "Дата_сварки"
            FROM condition_weld 
            WHERE {where_clause}
            ORDER BY "Дата_сварки" DESC, "ISO", "стык"
            LIMIT {per_page} OFFSET {offset}
        '''
        
        cursor.execute(data_query, params)
        rows = cursor.fetchall()
        
        # Конвертируем в список словарей
        data = [dict(row) for row in rows]
        
        # Вычисляем пагинацию
        total_pages = (total_records + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages
        
        # Статистика
        stats = {
            'total_records': total_records,
            'filtered_records': len(data),
            'current_page': page,
            'total_pages': total_pages,
            'per_page': per_page
        }
        
        # Получаем статистику по статусам
        status_statistics = get_status_statistics()
        
        return render_template('enhanced_filter.html',
                             data=data,
                             stats=stats,
                             status_statistics=status_statistics,
                             available_statuses=available_statuses,
                             available_titles=available_titles,
                             available_lines=available_lines,
                             current_status=status_filter,
                             current_title=title_filter,
                             current_line=line_filter,
                             has_prev=has_prev,
                             has_next=has_next,
                             prev_page=page-1,
                             next_page=page+1)
        
    except Exception as e:
        logger.error(f"Ошибка в enhanced_filter: {e}")
        flash(f'Ошибка фильтрации: {e}', 'error')
        return redirect(url_for('index'))
    finally:
        conn.close()

@app.route('/api/status_statistics')
def api_status_statistics():
    """API для получения статистики по статусам РК и ВИК"""
    try:
        # Получаем параметры фильтрации
        status_filter = request.args.get('status', 'all')
        title_filter = request.args.get('title', 'all')
        line_filter = request.args.get('line', 'all')
        
        status_statistics = get_status_statistics_filtered(status_filter, title_filter, line_filter)
        return jsonify({
            'success': True,
            'data': status_statistics
        })
    except Exception as e:
        logger.error(f"Ошибка получения статистики статусов: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/results_statistics')
def api_results_statistics():
    """API для получения статистики по результатам заключений РК и ВИК с фильтром"""
    try:
        # Получаем параметры фильтрации
        status_filter = request.args.get('status', 'all')
        title_filter = request.args.get('title', 'all')
        line_filter = request.args.get('line', 'all')
        
        results_statistics = get_results_statistics(status_filter, title_filter, line_filter)
        return jsonify({
            'success': True,
            'data': results_statistics
        })
    except Exception as e:
        logger.error(f"Ошибка получения статистики результатов: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/export_filtered_data', methods=['POST'])
def export_filtered_data():
    """Экспорт отфильтрованных данных из enhanced_filter в Excel"""
    try:
        # Получаем параметры фильтрации из формы
        status_filter = request.form.get('status', 'all')
        title_filter = request.form.get('title', 'all')
        line_filter = request.form.get('line', 'all')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к базе данных'}), 500
        
        try:
            cursor = conn.cursor()
            
            # Формируем условия фильтрации
            conditions = []
            params = []
            
            if status_filter and status_filter != 'all':
                conditions.append('"Статус_РК" = ?')
                params.append(status_filter)
            
            if title_filter and title_filter != 'all':
                conditions.append('"Титул" LIKE ?')
                params.append(f'%{title_filter}%')
            
            if line_filter and line_filter != 'all':
                conditions.append('"Линия" = ?')
                params.append(line_filter)
            
            # Добавляем условие исключения записей с Код_удаления = 'R'
            conditions.append('("Код_удаления" IS NULL OR "Код_удаления" != "R")')
            
            # Формируем WHERE условие
            where_clause = ' AND '.join(conditions) if conditions else '1=1'
            
            # Получаем данные для экспорта
            query = f'''
                SELECT 
                    "id",
                    "ISO", 
                    "Линия",
                    "Титул",
                    "стык",
                    "Статус_РК",
                    "Статус_ВИК",
                    "Результаты_Заключения_РК",
                    "Результаты_АКТ_ВИК",
                    "Дата_сварки"
                FROM condition_weld 
                WHERE {where_clause}
                ORDER BY "Дата_сварки" DESC, "ISO", "стык"
            '''
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Проверяем, есть ли данные
            if not rows:
                return jsonify({'error': 'Нет данных для экспорта по заданным фильтрам'}), 404
            
            # Создаем Excel файл
            import pandas as pd
            from io import BytesIO
            
            # Получаем имена столбцов
            columns = [description[0] for description in cursor.description]
            
            # Конвертируем в DataFrame
            df = pd.DataFrame(rows, columns=columns)
            
            # Создаем Excel файл в памяти
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Отфильтрованные данные', index=False)
            
            output.seek(0)
            
            # Формируем имя файла
            timestamp = get_filename_timestamp()
            filename = f'enhanced_filter_export_{timestamp}.xlsx'
            
            conn.close()
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            logger.error(f"Ошибка экспорта данных: {e}")
            conn.close()
            return jsonify({'error': f'Ошибка экспорта: {e}'}), 500
            
    except Exception as e:
        logger.error(f"Ошибка в export_filtered_data: {e}")
        return jsonify({'error': f'Ошибка обработки запроса: {e}'}), 500

@app.route('/export_cards_pdf', methods=['POST'])
def export_cards_pdf():
    """Экспорт карточек статистики в PDF"""
    try:
        # Получаем параметры фильтрации из формы
        status_filter = request.form.get('status', 'all')
        title_filter = request.form.get('title', 'all')
        line_filter = request.form.get('line', 'all')
        
        # Получаем статистику статусов
        status_statistics = get_status_statistics_filtered(status_filter, title_filter, line_filter)
        
        # Получаем статистику результатов
        results_statistics = get_results_statistics(status_filter, title_filter, line_filter)
        
        # Создаем PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import datetime
        import os
        
        # Регистрируем шрифт для поддержки кириллицы
        try:
            # Пытаемся найти системный шрифт Arial
            font_paths = [
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/ARIAL.TTF',
                '/System/Library/Fonts/Arial.ttf',  # macOS
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'  # Linux
            ]
            
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Arial', font_path))
                    font_registered = True
                    break
            
            if not font_registered:
                # Используем встроенный шрифт DejaVu Sans
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                pdfmetrics.registerFont(UnicodeCIDFont('DejaVuSans'))
        except Exception as e:
            logger.warning(f"Не удалось зарегистрировать шрифт: {e}")
            # Используем встроенный шрифт
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('DejaVuSans'))
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Создаем стили с поддержкой кириллицы (компактные для одного листа)
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=12,
            spaceAfter=15,
            alignment=1,  # Center alignment
            fontName='Arial'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=10,
            spaceAfter=5,
            fontName='Arial'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Arial'
        )
        
        # Информация о фильтрах
        filter_info = []
        if status_filter != 'all':
            filter_info.append(f"Статус РК: {status_filter}")
        if title_filter != 'all':
            filter_info.append(f"Титул: {title_filter}")
        if line_filter != 'all':
            filter_info.append(f"Линия: {line_filter}")
        
        filter_text = "Все записи" if not filter_info else ", ".join(filter_info)
        
        story.append(Paragraph("Статистика по сварным соединениям", title_style))
        story.append(Paragraph(f"Фильтры: {filter_text}", normal_style))
        story.append(Paragraph(f"Дата создания: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
        story.append(Spacer(1, 5))
        
        # Статистика по статусам РК
        story.append(Paragraph("Заключения_РК НГС", heading_style))
        story.append(Spacer(1, 5))
        
        rk_data = [['Статус', 'Количество', 'Процент']]
        for stat in status_statistics.get('rk_status_stats', []):
            rk_data.append([stat['status'], str(stat['count']), f"{stat['percentage']}%"])
        
        rk_table = Table(rk_data)
        rk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(rk_table)
        story.append(Spacer(1, 8))
        
        # Статистика по статусам ВИК
        story.append(Paragraph("Заключения_ВИК НГС", heading_style))
        story.append(Spacer(1, 3))
        
        vik_data = [['Статус', 'Количество', 'Процент']]
        for stat in status_statistics.get('vik_status_stats', []):
            vik_data.append([stat['status'], str(stat['count']), f"{stat['percentage']}%"])
        
        vik_table = Table(vik_data)
        vik_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(vik_table)
        story.append(Spacer(1, 8))
        
        # Статистика по результатам РК
        story.append(Paragraph("Заключения_РК из Китайского", heading_style))
        story.append(Spacer(1, 3))
        
        rk_results_data = [['Результат', 'Количество', 'Процент']]
        for stat in results_statistics.get('rk_results_stats', []):
            rk_results_data.append([stat['result'], str(stat['count']), f"{stat['percentage']}%"])
        
        rk_results_table = Table(rk_results_data)
        rk_results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(rk_results_table)
        story.append(Spacer(1, 8))
        
        # Статистика по результатам ВИК
        story.append(Paragraph("Заключения_ВИК из Китайского", heading_style))
        story.append(Spacer(1, 3))
        
        vik_results_data = [['Результат', 'Количество', 'Процент']]
        for stat in results_statistics.get('vik_results_stats', []):
            vik_results_data.append([stat['result'], str(stat['count']), f"{stat['percentage']}%"])
        
        vik_results_table = Table(vik_results_data)
        vik_results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(vik_results_table)
        
        # Создаем PDF
        doc.build(story)
        buffer.seek(0)
        
        # Возвращаем PDF файл
        from flask import Response
        
        filename = f"statistics_cards_{get_filename_timestamp()}.pdf"
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/pdf'
            }
        )
        
    except Exception as e:
        logger.error(f"Ошибка экспорта карточек в PDF: {e}")
        return jsonify({'error': f'Ошибка экспорта: {e}'}), 500

@app.route('/vue')
def vue_app():
    """Маршрут для Vue.js приложения"""
    return render_template('vue_app.html')

# API endpoints для Vue.js
@app.route('/api/vue/stats')
def vue_stats():
    """API для получения статистики для Vue.js"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'}), 500
        
        cursor = conn.cursor()
        
        # Получаем статистику по таблицам
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        
        total_records = 0
        table_stats = []
        
        for table in tables:
            table_name = table[0]
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            total_records += count
            
            table_stats.append({
                'name': table_name,
                'records': count
            })
        
        # Получаем правильную статистику по сваркам из таблицы logs_lnk
        successful_welds = 0
        in_progress = 0
        
        try:
            # Успешные сварки - РК Годен
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" = "Годен"')
            successful_welds = cursor.fetchone()[0]
        except Exception as e:
            logger.warning(f"Ошибка при подсчете успешных сварки: {e}")
        
        try:
            # В обработке - РК Заявлен (отставание в контроле)
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE LOWER("РК") LIKE "%явлен%"')
            in_progress = cursor.fetchone()[0]
        except Exception as e:
            logger.warning(f"Ошибка при подсчете сварки в обработке: {e}")
        
        # Дополнительная статистика для более полной картины
        vik_good = 0
        rk_defects = 0
        rt_requested = 0
        
        try:
            # ВИК Годен
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_ВИК" = "Годен"')
            vik_good = cursor.fetchone()[0]
        except Exception as e:
            logger.warning(f"Ошибка при подсчете ВИК годен: {e}")
        
        try:
            # РК дефекты
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Статус_РК" LIKE "%Не годен%"')
            rk_defects = cursor.fetchone()[0]
        except Exception as e:
            logger.warning(f"Ошибка при подсчете РК дефекты: {e}")
        
        try:
            # Заявки на РК
            cursor.execute('SELECT COUNT(*) FROM logs_lnk WHERE "Заявленны_виды_контроля" LIKE "%RT%"')
            rt_requested = cursor.fetchone()[0]
        except Exception as e:
            logger.warning(f"Ошибка при подсчете заявок на РК: {e}")
        
        stats = {
            'totalRecords': total_records,
            'successfulWelds': successful_welds,
            'inProgress': in_progress,
            'vikGood': vik_good,
            'rkDefects': rk_defects,
            'rtRequested': rt_requested,
            'tables': len(tables),
            'tableStats': table_stats
        }
        
        conn.close()
        return jsonify(stats)
        
    except Exception as e:
        logger.error(f'Ошибка получения статистики для Vue.js: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/vue/logs')
def vue_logs():
    """API для получения логов для Vue.js"""
    try:
        limit = request.args.get('limit', 100, type=int)
        
        # Читаем логи из файла
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'staff_log')
        
        logs = []
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()[-limit:]  # Последние записи
                
                for i, line in enumerate(lines):
                    try:
                        # Парсим лог (простая реализация)
                        parts = line.strip().split(' - ')
                        if len(parts) >= 3:
                            timestamp = parts[0]
                            level = 'INFO'
                            module = 'system'
                            message = ' - '.join(parts[1:])
                            
                            logs.append({
                                'id': i + 1,
                                'timestamp': timestamp,
                                'level': level,
                                'module': module,
                                'message': message
                            })
                    except:
                        continue
        
        return jsonify(logs)
        
    except Exception as e:
        logger.error(f'Ошибка получения логов для Vue.js: {e}')
        return jsonify({'error': str(e)}), 500

def get_table_display_name(table_name):
    """Функция для получения читаемого названия таблицы"""
    table_names_mapping = {
        '2024_2025': 'Календарь',
        'Log_Piping_PTO': 'Список ISO от ПТО',
        'Pipeline_Test_Package': 'Тест Пакеты',
        'defects_NDT': 'Дефекты РК',
        'folder_NDT_Report': 'Заключения pdf по РК',
        'logs_lnk': 'Журнал НК от НГС Эксперт',
        'pto_ndt_volume_register': 'Объем РК от ПТО(с изм)',
        'weld_repair_log': 'Журнал Ремонтов',
        'Daily_Staff_Allocation': 'Расстановка',
        'NDT_Findings_Transmission_Register': 'Реестр переданных заключений НГС Эксперт(excel)',
        'tks_registry': 'Реестр ТКС',
        'wl_china': 'ЖСР Китайский',
        'condition_weld': 'Статус стыка(послед)',
        'pipeline_weld_joint_iso': 'Реестр ПРОЕКТНЫХ Стыков от М Кран',
        'type_weld': 'Тип шва',
        'wl_report_smr': 'Отчет СМР площадка',
        'work_order_log_NDT': 'Заявки по НК от М Кран',
        'основнаяНК': 'Основная НК',
        'слов_клейм_факт': 'Словарь клейм факт'
    }
    return table_names_mapping.get(table_name, table_name)

@app.route('/api/vue/database')
def vue_database():
    """API для получения информации о базе данных для Vue.js"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к БД'}), 500
        
        cursor = conn.cursor()
        
        # Получаем список таблиц
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        
        db_stats = {
            'tables': len(tables),
            'records': 0,
            'size': 0,
            'lastUpdate': datetime.now().isoformat()
        }
        
        table_info = []
        
        for table in tables:
            table_name = table[0]
            
            # Количество записей
            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}`")
            records = cursor.fetchone()[0]
            db_stats['records'] += records
            
            # Размер таблицы (приблизительно)
            size = records * 1024  # Примерная оценка
            
            table_info.append({
                'name': table_name,
                'displayName': get_table_display_name(table_name),
                'records': records,
                'size': size,
                'lastUpdate': datetime.now().isoformat()
            })
        
        # Размер БД (приблизительно)
        db_stats['size'] = db_stats['records'] * 1024
        
        conn.close()
        
        return jsonify({
            'stats': db_stats,
            'tables': table_info
        })
        
    except Exception as e:
        logger.error(f'Ошибка получения информации о БД для Vue.js: {e}')
        return jsonify({'error': str(e)}), 500



@app.route('/api/open_file')
def open_file():
    """API для открытия файлов через системное приложение"""
    try:
        file_path = request.args.get('path', '')
        
        if not file_path:
            return jsonify({'error': 'Путь к файлу не указан'}), 400
        
        # Нормализуем путь для корректного сравнения
        file_path = os.path.normpath(file_path)
        
        logger.info(f'Попытка открытия файла: {file_path}')
        
        # Проверяем, что файл существует
        if not os.path.exists(file_path):
            logger.error(f'Файл не найден: {file_path}')
            return jsonify({'error': f'Файл не найден: {file_path}'}), 404
        
        # Проверяем размер файла
        file_size = os.path.getsize(file_path)
        logger.info(f'Размер файла: {file_size} байт')
        
        # Проверяем права доступа
        if not os.access(file_path, os.R_OK):
            logger.error(f'Нет прав на чтение файла: {file_path}')
            return jsonify({'error': f'Нет прав на чтение файла: {file_path}'}), 403
        
        # Проверяем расширение файла
        file_ext = os.path.splitext(file_path)[1].lower()
        logger.info(f'Расширение файла: {file_ext}')
        
        # Проверяем, что файл не заблокирован другим процессом
        try:
            with open(file_path, 'rb') as test_file:
                test_file.read(1)
            logger.info('Файл доступен для чтения')
        except PermissionError:
            logger.error(f'Файл заблокирован другим процессом: {file_path}')
            return jsonify({'error': f'Файл заблокирован другим процессом: {file_path}'}), 403
        except Exception as e:
            logger.error(f'Ошибка при проверке доступности файла: {e}')
            return jsonify({'error': f'Ошибка при проверке доступности файла: {str(e)}'}), 500
        
        # Проверяем, что путь безопасный (не выходит за пределы разрешенных директорий)
        allowed_paths = [
            'D:\\МК_Кран',
            'D:\\МК_Кран\\script_M_Kran',
            'D:\\МК_Кран\\МК_Кран_Кингесеп',
            'C:\\Users',
            'E:\\'
        ]
        
        is_allowed = False
        for allowed_path in allowed_paths:
            # Нормализуем разрешенный путь для сравнения
            normalized_allowed = os.path.normpath(allowed_path)
            if file_path.startswith(normalized_allowed):
                is_allowed = True
                break
        
        if not is_allowed:
            logger.warning(f'Попытка доступа к запрещенному файлу: {file_path}')
            logger.warning(f'Разрешенные пути: {allowed_paths}')
            return jsonify({'error': f'Доступ к файлу запрещен. Путь: {file_path}'}), 403
        
        # Открываем файл через системное приложение
        try:
            import subprocess
            import platform
            
            system = platform.system()
            logger.info(f'Операционная система: {system}')
            
            if system == "Windows":
                logger.info(f'Открываем файл в Windows: {file_path}')
                
                # Определяем расширение файла
                file_ext = os.path.splitext(file_path)[1].lower()
                
                try:
                    # Для Excel файлов пробуем открыть через Excel с параметрами
                    if file_ext in ['.xlsx', '.xls']:
                        logger.info('Пробуем открыть Excel файл через Excel с параметрами...')
                        try:
                            # Пробуем открыть через PowerShell с параметрами для полноэкранного режима
                            ps_command = f'Start-Process "{file_path}" -WindowStyle Maximized'
                            result = subprocess.run(['powershell', '-Command', ps_command], 
                                                  capture_output=True, text=True, shell=True)
                            if result.returncode == 0:
                                logger.info('Excel файл успешно открыт через PowerShell в полноэкранном режиме')
                            else:
                                # Если не сработало, пробуем обычный способ
                                os.startfile(file_path)
                                logger.info('Excel файл открыт через os.startfile')
                        except Exception as ps_error:
                            logger.warning(f'PowerShell не сработал: {ps_error}')
                            # Пробуем обычный способ
                            os.startfile(file_path)
                            logger.info('Excel файл открыт через os.startfile (fallback)')
                    else:
                        # Для остальных файлов тоже пробуем PowerShell для полноэкранного режима
                        logger.info('Пробуем открыть файл через PowerShell в полноэкранном режиме...')
                        try:
                            ps_command = f'Start-Process "{file_path}" -WindowStyle Maximized'
                            result = subprocess.run(['powershell', '-Command', ps_command], 
                                                  capture_output=True, text=True, shell=True)
                            if result.returncode == 0:
                                logger.info('Файл успешно открыт через PowerShell в полноэкранном режиме')
                            else:
                                # Если не сработало, используем обычный способ
                                os.startfile(file_path)
                                logger.info('Файл открыт через os.startfile')
                        except Exception as ps_error:
                            logger.warning(f'PowerShell не сработал: {ps_error}')
                            # Используем обычный способ
                            os.startfile(file_path)
                            logger.info('Файл открыт через os.startfile (fallback)')
                        
                except Exception as startfile_error:
                    logger.warning(f'os.startfile не сработал: {startfile_error}')
                    logger.info('Пробуем альтернативный способ через subprocess...')
                    try:
                        # Альтернативный способ для Windows
                        result = subprocess.run(['cmd', '/c', 'start', '', file_path], 
                                              capture_output=True, text=True, shell=True)
                        if result.returncode != 0:
                            logger.error(f'Ошибка при открытии файла через subprocess: {result.stderr}')
                            # Не возвращаем ошибку, просто логируем
                            logger.warning('Продолжаем выполнение несмотря на ошибку subprocess')
                        else:
                            logger.info('Файл успешно отправлен в системное приложение через subprocess')
                    except Exception as subprocess_error:
                        logger.error(f'Ошибка при открытии файла через subprocess: {subprocess_error}')
                        # Не возвращаем ошибку, просто логируем
                        logger.warning('Продолжаем выполнение несмотря на ошибку subprocess')
            elif system == "Darwin":  # macOS
                logger.info(f'Открываем файл через open: {file_path}')
                try:
                    result = subprocess.run(["open", file_path], capture_output=True, text=True)
                    if result.returncode != 0:
                        logger.error(f'Ошибка при открытии файла: {result.stderr}')
                        logger.warning('Продолжаем выполнение несмотря на ошибку open')
                    else:
                        logger.info('Файл успешно открыт через open')
                except Exception as open_error:
                    logger.error(f'Исключение при открытии файла через open: {open_error}')
                    logger.warning('Продолжаем выполнение несмотря на ошибку open')
            else:  # Linux
                logger.info(f'Открываем файл через xdg-open: {file_path}')
                try:
                    result = subprocess.run(["xdg-open", file_path], capture_output=True, text=True)
                    if result.returncode != 0:
                        logger.error(f'Ошибка при открытии файла: {result.stderr}')
                        logger.warning('Продолжаем выполнение несмотря на ошибку xdg-open')
                    else:
                        logger.info('Файл успешно открыт через xdg-open')
                except Exception as xdg_error:
                    logger.error(f'Исключение при открытии файла через xdg-open: {xdg_error}')
                    logger.warning('Продолжаем выполнение несмотря на ошибку xdg-open')
            
            file_name = os.path.basename(file_path)
            logger.info(f'Файл успешно открыт: {file_name}')
            
            return jsonify({
                'success': True,
                'message': f'Файл открыт: {file_name}',
                'file_path': file_path,
                'file_size': file_size
            })
            
        except Exception as e:
            logger.error(f'Исключение при открытии файла: {str(e)}')
            import traceback
            logger.error(f'Traceback: {traceback.format_exc()}')
            return jsonify({'error': f'Не удалось открыть файл: {str(e)}'}), 500
        
    except Exception as e:
        logger.error(f'Общая ошибка при открытии файла: {e}')
        import traceback
        logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/open_file_location')
def open_file_location():
    """API для открытия папки с файлом в проводнике"""
    try:
        file_path = request.args.get('path', '')
        
        if not file_path:
            return jsonify({'error': 'Путь к файлу не указан'}), 400
        
        # Нормализуем путь для корректного сравнения
        file_path = os.path.normpath(file_path)
        
        logger.info(f'Попытка открытия папки с файлом: {file_path}')
        
        # Проверяем, что файл существует
        if not os.path.exists(file_path):
            logger.error(f'Файл не найден: {file_path}')
            return jsonify({'error': f'Файл не найден: {file_path}'}), 404
        
        # Проверяем, что путь безопасный
        allowed_paths = [
            'D:\\МК_Кран',
            'D:\\МК_Кран\\script_M_Kran',
            'D:\\МК_Кран\\МК_Кран_Кингесеп',
            'C:\\Users',
            'E:\\'
        ]
        
        is_allowed = False
        for allowed_path in allowed_paths:
            normalized_allowed = os.path.normpath(allowed_path)
            if file_path.startswith(normalized_allowed):
                is_allowed = True
                break
        
        if not is_allowed:
            logger.warning(f'Попытка доступа к запрещенному файлу: {file_path}')
            return jsonify({'error': f'Доступ к файлу запрещен. Путь: {file_path}'}), 403
        
        # Открываем папку с файлом
        try:
            import subprocess
            import platform
            
            system = platform.system()
            logger.info(f'Открываем папку в системе: {system}')
            
            if system == "Windows":
                # Открываем папку и выделяем файл
                try:
                    result = subprocess.run(['explorer', '/select,', file_path], 
                                          capture_output=True, text=True, shell=True)
                    if result.returncode != 0:
                        logger.error(f'Ошибка при открытии папки: {result.stderr}')
                        # Не возвращаем ошибку, просто логируем
                        logger.warning('Продолжаем выполнение несмотря на ошибку explorer')
                    else:
                        logger.info('Папка успешно открыта через explorer')
                except Exception as explorer_error:
                    logger.error(f'Исключение при открытии папки через explorer: {explorer_error}')
                    # Не возвращаем ошибку, просто логируем
                    logger.warning('Продолжаем выполнение несмотря на ошибку explorer')
            elif system == "Darwin":  # macOS
                # Открываем папку в Finder
                folder_path = os.path.dirname(file_path)
                try:
                    result = subprocess.run(['open', folder_path], capture_output=True, text=True)
                    if result.returncode != 0:
                        logger.error(f'Ошибка при открытии папки: {result.stderr}')
                        logger.warning('Продолжаем выполнение несмотря на ошибку open')
                    else:
                        logger.info('Папка успешно открыта через open')
                except Exception as open_error:
                    logger.error(f'Исключение при открытии папки через open: {open_error}')
                    logger.warning('Продолжаем выполнение несмотря на ошибку open')
            else:  # Linux
                # Открываем папку в файловом менеджере
                folder_path = os.path.dirname(file_path)
                try:
                    result = subprocess.run(['xdg-open', folder_path], capture_output=True, text=True)
                    if result.returncode != 0:
                        logger.error(f'Ошибка при открытии папки: {result.stderr}')
                        logger.warning('Продолжаем выполнение несмотря на ошибку xdg-open')
                    else:
                        logger.info('Папка успешно открыта через xdg-open')
                except Exception as xdg_error:
                    logger.error(f'Исключение при открытии папки через xdg-open: {xdg_error}')
                    logger.warning('Продолжаем выполнение несмотря на ошибку xdg-open')
            
            folder_path = os.path.dirname(file_path)
            file_name = os.path.basename(file_path)
            logger.info(f'Папка успешно открыта: {folder_path}')
            
            return jsonify({
                'success': True,
                'message': f'Папка открыта: {folder_path}',
                'file_name': file_name,
                'folder_path': folder_path
            })
            
        except Exception as e:
            logger.error(f'Исключение при открытии папки: {str(e)}')
            import traceback
            logger.error(f'Traceback: {traceback.format_exc()}')
            return jsonify({'error': f'Не удалось открыть папку: {str(e)}'}), 500
        
    except Exception as e:
        logger.error(f'Общая ошибка при открытии папки: {e}')
        import traceback
        logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500



@app.route('/api/check_joint_duplicate', methods=['POST'])
def check_joint_duplicate():
    """API для проверки дублирования стыка в базе данных"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        drawing_number = data.get('drawing_number', '').strip()
        line_number = data.get('line_number', '').strip()
        joint_number = data.get('joint_number', '').strip()
        current_joint_id = data.get('current_joint_id', '').strip()
        
        if not drawing_number or not line_number or not joint_number:
            return jsonify({'error': 'Не указаны все необходимые параметры'})
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pipeline_weld_joint_iso'")
        if not cursor.fetchone():
            return jsonify({'error': 'Таблица pipeline_weld_joint_iso не найдена'})
        
        # Проверяем дубликат
        if current_joint_id:
            # Если это обновление существующей записи, исключаем её из проверки
            cursor.execute('''
                SELECT id, стык FROM pipeline_weld_joint_iso 
                WHERE ISO = ? AND Линия = ? AND стык = ? AND id != ?
            ''', (drawing_number, line_number, joint_number, current_joint_id))
        else:
            # Если это новая запись, проверяем все записи
            cursor.execute('''
                SELECT id, стык FROM pipeline_weld_joint_iso 
                WHERE ISO = ? AND Линия = ? AND стык = ?
            ''', (drawing_number, line_number, joint_number))
        
        existing_joints = cursor.fetchall()
        
        conn.close()
        
        if existing_joints:
            return jsonify({
                'duplicate': True,
                'message': f'Стык "{joint_number}" уже существует в базе данных',
                'existing_ids': [row[0] for row in existing_joints]
            })
        else:
            return jsonify({
                'duplicate': False,
                'message': 'Стык уникален'
            })
            
    except Exception as e:
        return jsonify({'error': str(e)})
    finally:
        if conn:
            conn.close()

def create_list_id_pto_table():
    """Создает таблицу list_ID_PTO если она не существует"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='list_ID_PTO'")
        table_exists = cursor.fetchone()
        
        if table_exists:
            # Проверяем структуру существующей таблицы
            cursor.execute("PRAGMA table_info(list_ID_PTO)")
            columns = cursor.fetchall()
            column_names = [col[1] for col in columns]
            
            # Если структура неправильная, пересоздаем таблицу
            expected_columns = ['id', 'Номер_ИД', 'Проект', 'Ответственный', 'Дата_начала', 
                              'Дата_окончания', 'Статус', 'Описание', 'Дата_создания', 'Дата_обновления']
            
            if not all(col in column_names for col in expected_columns):
                logger.warning("Структура таблицы list_ID_PTO неправильная, пересоздаем...")
                cursor.execute("DROP TABLE IF EXISTS list_ID_PTO")
                table_exists = False
        
        if not table_exists:
            # Создаем таблицу list_ID_PTO
            cursor.execute('''
                CREATE TABLE list_ID_PTO (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    Номер_ИД TEXT UNIQUE NOT NULL,
                    Проект TEXT,
                    Ответственный TEXT,
                    Дата_начала TEXT,
                    Дата_окончания TEXT,
                    Статус TEXT,
                    Описание TEXT,
                    Дата_создания TEXT DEFAULT (datetime('now', 'localtime')),
                    Дата_обновления TEXT DEFAULT (datetime('now', 'localtime'))
                )
            ''')
            logger.info("Таблица list_ID_PTO создана успешно")
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Ошибка создания таблицы list_ID_PTO: {e}")
        conn.close()
        return False

@app.route('/api/save_id_to_database', methods=['POST'])
def save_id_to_database():
    """API для сохранения ИД в таблицу list_ID_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        # Создаем таблицу если её нет
        if not create_list_id_pto_table():
            return jsonify({'error': 'Не удалось создать таблицу list_ID_PTO'})
        
        data = request.get_json()
        
        # Валидация данных - только номер ИД обязателен
        if not data.get('number'):
            return jsonify({'error': 'Номер ИД обязателен для заполнения'})
        
        # Проверяем дублирование по Номер_ИД
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM list_ID_PTO WHERE Номер_ИД = ?', (data['number'],))
        existing_record = cursor.fetchone()
        
        if existing_record:
            # Даже если ИД уже существует, создаем папки
            folder_success, folder_message = create_id_folders(data['number'])
            
            if folder_success:
                return jsonify({
                    'error': f'ИД с номером "{data["number"]}" уже существует в базе данных. {folder_message}',
                    'duplicate': True,
                    'folder_created': True
                })
            else:
                return jsonify({
                    'error': f'ИД с номером "{data["number"]}" уже существует в базе данных. Предупреждение: {folder_message}',
                    'duplicate': True,
                    'folder_warning': folder_message,
                    'folder_created': False
                })
        
        # Вставляем новую запись (без поля id, так как оно автоинкрементное)
        cursor.execute('''
            INSERT INTO list_ID_PTO (
                Номер_ИД, Проект, Ответственный, Дата_начала, 
                Дата_окончания, Статус, Описание
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['number'],
            data['project'],
            data['responsible'],
            data['startDate'],
            data.get('endDate', ''),
            data.get('status', 'pending'),
            data.get('description', '')
        ))
        
        conn.commit()
        new_id = cursor.lastrowid
        
        conn.close()
        
        # Создаем папки для ИД
        folder_success, folder_message = create_id_folders(data['number'])
        
        if folder_success:
            return jsonify({
                'success': True,
                'message': f'ИД успешно сохранен в базе данных. {folder_message}',
                'id': new_id,
                'folder_created': True
            })
        else:
            # ИД сохранен в БД, но папки не созданы
            logger.warning(f"ИД сохранен в БД, но папки не созданы: {folder_message}")
            return jsonify({
                'success': True,
                'message': f'ИД успешно сохранен в базе данных. Предупреждение: {folder_message}',
                'id': new_id,
                'folder_warning': folder_message,
                'folder_created': False
            })
        
    except Exception as e:
        logger.error(f"Ошибка сохранения ИД в БД: {e}")
        return jsonify({'error': f'Ошибка сохранения: {str(e)}'})
    finally:
        if conn:
            conn.close()

@app.route('/api/update_id_in_database', methods=['POST'])
def update_id_in_database():
    """API для обновления ИД в таблице list_ID_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        
        # Валидация данных
        if not data.get('id'):
            return jsonify({'error': 'ID записи обязателен для обновления'})
        
        if not data.get('number'):
            return jsonify({'error': 'Номер ИД обязателен для заполнения'})
        
        # Проверяем дублирование по Номер_ИД (исключая текущую запись)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM list_ID_PTO WHERE Номер_ИД = ? AND id != ?', 
                      (data['number'], data['id']))
        existing_record = cursor.fetchone()
        
        if existing_record:
            return jsonify({
                'error': f'ИД с номером "{data["number"]}" уже существует в базе данных',
                'duplicate': True
            })
        
        # Обновляем запись
        cursor.execute('''
            UPDATE list_ID_PTO SET 
                Номер_ИД = ?, Проект = ?, Ответственный = ?, 
                Дата_начала = ?, Дата_окончания = ?, Статус = ?, 
                Описание = ?, Дата_обновления = (datetime('now', 'localtime'))
            WHERE id = ?
        ''', (
            data['number'],
            data['project'],
            data['responsible'],
            data['startDate'],
            data.get('endDate', ''),
            data.get('status', 'pending'),
            data.get('description', ''),
            data['id']
        ))
        
        if cursor.rowcount == 0:
            return jsonify({'error': 'Запись не найдена для обновления'})
        
        conn.commit()
        conn.close()
        
        # Создаем папки для ИД (если изменился номер)
        folder_success, folder_message = create_id_folders(data['number'])
        
        if folder_success:
            return jsonify({
                'success': True,
                'message': f'ИД успешно обновлен в базе данных. {folder_message}',
                'folder_created': True
            })
        else:
            # ИД обновлен в БД, но папки не созданы
            logger.warning(f"ИД обновлен в БД, но папки не созданы: {folder_message}")
            return jsonify({
                'success': True,
                'message': f'ИД успешно обновлен в базе данных. Предупреждение: {folder_message}',
                'folder_warning': folder_message,
                'folder_created': False
            })
        
    except Exception as e:
        logger.error(f"Ошибка обновления ИД в БД: {e}")
        return jsonify({'error': f'Ошибка обновления: {str(e)}'})
    finally:
        if conn:
            conn.close()

@app.route('/api/load_ids_from_database')
def load_ids_from_database():
    """API для загрузки всех ИД из таблицы list_ID_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='list_ID_PTO'")
        if not cursor.fetchone():
            return jsonify({'ids': []})
        
        # Сначала проверим структуру таблицы
        cursor.execute("PRAGMA table_info(list_ID_PTO)")
        columns = cursor.fetchall()
        logger.info(f"Структура таблицы list_ID_PTO: {columns}")
        
        # Проверяем, есть ли поле 'id' в таблице
        column_names = [col[1] for col in columns]
        logger.info(f"Имена столбцов в таблице: {column_names}")
        
        if 'id' not in column_names:
            logger.error("Поле 'id' отсутствует в таблице list_ID_PTO!")
            # Принудительно пересоздаем таблицу
            cursor.execute("DROP TABLE IF EXISTS list_ID_PTO")
            cursor.execute('''
                CREATE TABLE list_ID_PTO (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    Номер_ИД TEXT UNIQUE NOT NULL,
                    Проект TEXT,
                    Ответственный TEXT,
                    Дата_начала TEXT,
                    Дата_окончания TEXT,
                    Статус TEXT,
                    Описание TEXT,
                    Дата_создания TEXT DEFAULT (datetime('now', 'localtime')),
                    Дата_обновления TEXT DEFAULT (datetime('now', 'localtime'))
                )
            ''')
            conn.commit()
            logger.info("Таблица list_ID_PTO пересоздана с правильной структурой")
            return jsonify({'ids': []})
        
        # Загружаем все записи
        cursor.execute('''
            SELECT id, Номер_ИД, Проект, Ответственный, Дата_начала, 
                   Дата_окончания, Статус, Описание, Дата_создания, Дата_обновления
            FROM list_ID_PTO 
            ORDER BY Дата_создания DESC
        ''')
        
        records = cursor.fetchall()
        ids = []
        
        for record in records:
            ids.append({
                'id': record[0],
                'number': record[1],
                'project': record[2],
                'responsible': record[3],
                'startDate': record[4],
                'endDate': record[5] if record[5] else None,
                'status': record[6],
                'description': record[7] if record[7] else '',
                'createdAt': record[8],
                'updatedAt': record[9]
            })
        
        conn.close()
        return jsonify({'ids': ids})
        
    except Exception as e:
        logger.error(f"Ошибка загрузки ИД из БД: {e}")
        return jsonify({'error': f'Ошибка загрузки: {str(e)}'})
    finally:
        if conn:
            conn.close()


@app.route('/api/recreate_folder', methods=['POST'])
def recreate_folder():
    """API для пересоздания папки ИД"""
    try:
        data = request.get_json()
        
        if not data.get('idNumber'):
            return jsonify({'error': 'Номер ИД обязателен для пересоздания папки'})
        
        # Пересоздаем папку
        folder_success, folder_message = create_id_folders(data['idNumber'])
        
        if folder_success:
            return jsonify({
                'success': True,
                'message': f'Папка для ИД {data["idNumber"]} пересоздана успешно'
            })
        else:
            return jsonify({
                'success': False,
                'error': folder_message
            })
        
    except Exception as e:
        logger.error(f"Ошибка пересоздания папки: {e}")
        return jsonify({'error': f'Ошибка пересоздания папки: {str(e)}'})

@app.route('/api/delete_id_from_database', methods=['POST'])
def delete_id_from_database():
    """API для удаления ИД из таблицы list_ID_PTO"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Ошибка подключения к БД'})
    
    try:
        data = request.get_json()
        
        if not data.get('id'):
            return jsonify({'error': 'ID записи обязателен для удаления'})
        
        cursor = conn.cursor()
        cursor.execute('DELETE FROM list_ID_PTO WHERE id = ?', (data['id'],))
        
        if cursor.rowcount == 0:
            return jsonify({'error': 'Запись не найдена для удаления'})
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'ИД успешно удален из базы данных'
        })
        
    except Exception as e:
        logger.error(f"Ошибка удаления ИД из БД: {e}")
        return jsonify({'error': f'Ошибка удаления: {str(e)}'})
    finally:
        if conn:
            conn.close()


# API маршруты для редактирования condition_weld
@app.route('/api/condition_weld/<int:record_id>', methods=['GET'])
def api_get_condition_weld_record(record_id):
    """Получение данных записи condition_weld по ID (GET запрос)"""
    logger.info(f'Запрос записи condition_weld с ID: {record_id}')
    
    try:
        conn = get_db_connection()
        if not conn:
            logger.error('Не удалось подключиться к базе данных')
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'}), 500
        
        cursor = conn.cursor()
        
        # Проверяем существование таблицы
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='condition_weld'")
        table_exists = cursor.fetchone()
        
        if not table_exists:
            logger.warning('Таблица condition_weld не существует')
            return jsonify({'success': False, 'message': 'Таблица condition_weld не существует'}), 404
        
        # Получаем данные записи
        cursor.execute("""
            SELECT 
                id, Титул, ISO, Линия, стык, Код_удаления, Тип_шва,
                ID_RT, РК, Статус_РК, Дата_контроля_РК,
                ID_VT, ВИК, Статус_ВИК, Дата_контроля_ВИК,
                ID_WC, Заключение_РК_N, Результаты_Заключения_РК,
                Дата_Заключения_РК, АКТ_ВИК_N, Дата_АКТ_ВИК, Результаты_АКТ_ВИК,
                Дата_сварки
            FROM condition_weld 
            WHERE id = ?
        """, (record_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            logger.warning(f'Запись с ID {record_id} не найдена')
            return jsonify({'success': False, 'message': f'Запись с ID {record_id} не найдена'}), 404
        
        # Преобразуем в словарь
        record_dict = {
            'id': row[0],
            'Титул': row[1],
            'ISO': row[2],
            'Линия': row[3],
            'стык': row[4],
            'Код_удаления': row[5],
            'Тип_шва': row[6],
            'ID_RT': row[7],
            'РК': row[8],
            'Статус_РК': row[9],
            'Дата_контроля_РК': row[10],
            'ID_VT': row[11],
            'ВИК': row[12],
            'Статус_ВИК': row[13],
            'Дата_контроля_ВИК': row[14],
            'ID_WC': row[15],
            'Заключение_РК_N': row[16],
            'Результаты_Заключения_РК': row[17],
            'Дата_Заключения_РК': row[18],
            'АКТ_ВИК_N': row[19],
            'Дата_АКТ_ВИК': row[20],
            'Результаты_АКТ_ВИК': row[21],
            'Дата_сварки': row[22]
        }
        
        logger.info(f'Успешно получена запись condition_weld с ID {record_id}')
        return jsonify({'success': True, 'record': record_dict})
        
    except Exception as e:
        logger.error(f'Ошибка получения записи condition_weld {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/get_condition_weld_by_id', methods=['POST'])
def get_condition_weld_by_id():
    """Получение данных записи condition_weld по ID для редактирования"""
    try:
        data = request.get_json()
        record_id = data.get('id')
        
        if not record_id:
            return jsonify({'error': 'ID записи обязателен'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Получаем данные записи
        cursor.execute("""
            SELECT 
                id, Титул, ISO, Линия, стык, Код_удаления, Тип_шва,
                ID_RT, РК, Статус_РК, Дата_контроля_РК,
                ID_VT, ВИК, Статус_ВИК, Дата_контроля_ВИК,
                ID_WC, Заключение_РК_N, Результаты_Заключения_РК,
                Дата_Заключения_РК, АКТ_ВИК_N, Дата_АКТ_ВИК, Результаты_АКТ_ВИК,
                Дата_сварки
            FROM condition_weld 
            WHERE id = ?
        """, (record_id,))
        
        record = cursor.fetchone()
        conn.close()
        
        if not record:
            return jsonify({'error': 'Запись не найдена'})
        
        # Преобразуем в словарь
        record_dict = {
            'id': record[0],
            'Титул': record[1],
            'ISO': record[2],
            'Линия': record[3],
            'стык': record[4],
            'Код_удаления': record[5],
            'Тип_шва': record[6],
            'ID_RT': record[7],
            'РК': record[8],
            'Статус_РК': record[9],
            'Дата_контроля_РК': record[10],
            'ID_VT': record[11],
            'ВИК': record[12],
            'Статус_ВИК': record[13],
            'Дата_контроля_ВИК': record[14],
            'ID_WC': record[15],
            'Заключение_РК_N': record[16],
            'Результаты_Заключения_РК': record[17],
            'Дата_Заключения_РК': record[18],
            'АКТ_ВИК_N': record[19],
            'Дата_АКТ_ВИК': record[20],
            'Результаты_АКТ_ВИК': record[21],
            'Дата_сварки': record[22]
        }
        
        return jsonify({'success': True, 'record': record_dict})
        
    except Exception as e:
        logger.error(f"Ошибка получения данных condition_weld: {e}")
        return jsonify({'error': f'Ошибка получения данных: {str(e)}'})

@app.route('/api/update_condition_weld', methods=['POST'])
def update_condition_weld():
    """Обновление данных в таблицах на основе condition_weld"""
    try:
        data = request.get_json()
        record_id = data.get('id')
        
        if not record_id:
            return jsonify({'error': 'ID записи обязателен'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Получаем исходные данные записи
        cursor.execute("""
            SELECT ISO, стык, ID_RT, ID_VT, ID_WC
            FROM condition_weld 
            WHERE id = ?
        """, (record_id,))
        
        original_record = cursor.fetchone()
        if not original_record:
            conn.close()
            return jsonify({'error': 'Запись не найдена'})
        
        iso, styk, id_rt, id_vt, id_wc = original_record
        
        # Обновляем данные в pipeline_weld_joint_iso
        if 'Код_удаления' in data or 'Тип_шва' in data or 'стык' in data:
            update_fields = []
            update_values = []
            
            if 'Код_удаления' in data:
                update_fields.append('Код_удаления = ?')
                update_values.append(data['Код_удаления'])
            
            if 'Тип_шва' in data:
                update_fields.append('"Тип_соединения_российский_стандарт" = ?')
                update_values.append(data['Тип_шва'])
            
            if 'стык' in data:
                update_fields.append('стык = ?')
                update_values.append(data['стык'])
            
            if update_fields:
                update_values.append(iso)
                update_values.append(styk)
                
                update_sql = f"""
                    UPDATE pipeline_weld_joint_iso 
                    SET {', '.join(update_fields)}
                    WHERE ISO = ? AND стык = ?
                """
                cursor.execute(update_sql, update_values)
        

        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Данные успешно обновлены'
        })
        
    except Exception as e:
        logger.error(f"Ошибка обновления condition_weld: {e}")
        return jsonify({'error': f'Ошибка обновления: {str(e)}'})

@app.route('/api/recreate_condition_weld_table', methods=['POST'])
def recreate_condition_weld_table():
    """Пересоздание таблицы condition_weld после обновления данных"""
    try:
        # Импортируем функцию создания таблицы
        import sys
        import os
        sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', 'scripts', 'data_loaders'))
        
        from create_condition_weld_table import create_condition_weld_table
        
        # Пересоздаем таблицу
        success = create_condition_weld_table()
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Таблица condition_weld успешно пересоздана'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Ошибка при пересоздании таблицы condition_weld'
            })
        
    except Exception as e:
        logger.error(f"Ошибка пересоздания таблицы condition_weld: {e}")
        return jsonify({'error': f'Ошибка пересоздания таблицы: {str(e)}'})

# API endpoints для таблицы слов_клейм_факт
@app.route('/api/slov_kleimo_fact/<int:record_id>', methods=['GET'])
def api_get_slov_kleimo_fact_record(record_id):
    """API для получения записи из таблицы слов_клейм_факт"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Получаем запись с JOIN для отображения полного ФИО
        cursor.execute('''
            SELECT s.*, f.ФИО as ФИО_полное
            FROM слов_клейм_факт s
            LEFT JOIN ФИО_свар f ON s.ФИО = f.id_fio
            WHERE s.id = ?
        ''', (record_id,))
        record = cursor.fetchone()
        
        if not record:
            conn.close()
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Получаем названия столбцов основной таблицы
        cursor.execute("PRAGMA table_info(слов_клейм_факт)")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Создаем словарь записи из основной таблицы
        record_dict = {}
        for i, col in enumerate(columns):
            record_dict[col] = record[i]
        
        # Добавляем полное ФИО из JOIN
        if len(record) > len(columns):
            record_dict['ФИО_полное'] = record[len(columns)]
        
        # Сохраняем исходный id_fio для JavaScript
        original_fio_id = record_dict.get('ФИО')
        
        # Заменяем id_fio на полное ФИО для отображения
        if record_dict.get('ФИО_полное'):
            record_dict['ФИО'] = record_dict['ФИО_полное']
            record_dict['ФИО_id'] = original_fio_id  # Сохраняем исходный id_fio
        
        conn.close()
        
        return jsonify({'success': True, 'record': record_dict})
        
    except Exception as e:
        logger.error(f'Ошибка получения записи слов_клейм_факт {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/slov_kleimo_fact/<int:record_id>', methods=['PUT'])
def api_update_slov_kleimo_fact_record(record_id):
    """API для обновления записи в таблице слов_клейм_факт"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Нет данных для обновления записи'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем существование записи
        cursor.execute('SELECT COUNT(*) FROM слов_клейм_факт WHERE id = ?', (record_id,))
        if cursor.fetchone()[0] == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Запись не найдена'})
        
        # Обновляем запись (поля ФИО и Примечание)
        cursor.execute('''
            UPDATE слов_клейм_факт SET 
            ФИО = ?,
            Примечание = ?
            WHERE id = ?
        ''', (
            data.get('ФИО', ''),
            data.get('Примечание', ''),
            record_id
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f'Обновлена запись слов_клейм_факт с ID: {record_id}')
        return jsonify({'success': True, 'message': 'Запись успешно обновлена'})
        
    except Exception as e:
        logger.error(f'Ошибка обновления записи слов_клейм_факт {record_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

# API endpoints для таблицы ФИО_свар
@app.route('/api/fio_svar', methods=['GET'])
def api_get_fio_svar_list():
    """API для получения списка ФИО из таблицы ФИО_свар"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Получаем все ФИО
        cursor.execute('SELECT id_fio, ФИО FROM ФИО_свар ORDER BY ФИО')
        fio_list = cursor.fetchall()
        
        # Преобразуем в список словарей
        fio_data = [{'id_fio': row[0], 'ФИО': row[1]} for row in fio_list]
        
        conn.close()
        
        return jsonify({'success': True, 'fio_list': fio_data})
        
    except Exception as e:
        logger.error(f'Ошибка получения списка ФИО: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/fio_svar', methods=['POST'])
def api_add_fio_svar():
    """API для добавления нового ФИО в таблицу ФИО_свар"""
    try:
        data = request.get_json()
        if not data or 'ФИО' not in data:
            return jsonify({'success': False, 'message': 'Не указано ФИО для добавления'})
        
        fio_text = data['ФИО'].strip()
        if not fio_text:
            return jsonify({'success': False, 'message': 'ФИО не может быть пустым'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Проверяем, не существует ли уже такое ФИО
        cursor.execute('SELECT id_fio FROM ФИО_свар WHERE ФИО = ?', (fio_text,))
        existing = cursor.fetchone()
        if existing:
            conn.close()
            return jsonify({'success': True, 'id_fio': existing[0], 'message': 'ФИО уже существует'})
        
        # Добавляем новое ФИО
        cursor.execute('INSERT INTO ФИО_свар (ФИО) VALUES (?)', (fio_text,))
        new_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        logger.info(f'Добавлено новое ФИО: {fio_text} с ID: {new_id}')
        return jsonify({'success': True, 'id_fio': new_id, 'message': 'ФИО успешно добавлено'})
        
    except Exception as e:
        logger.error(f'Ошибка добавления ФИО: {e}')
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/fio_svar/<int:fio_id>', methods=['GET'])
def api_get_fio_svar(fio_id):
    """API для получения конкретного ФИО по ID"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
        
        cursor = conn.cursor()
        
        # Получаем ФИО по ID
        cursor.execute('SELECT id_fio, ФИО FROM ФИО_свар WHERE id_fio = ?', (fio_id,))
        fio_record = cursor.fetchone()
        
        if not fio_record:
            conn.close()
            return jsonify({'success': False, 'message': 'ФИО не найдено'})
        
        fio_data = {'id_fio': fio_record[0], 'ФИО': fio_record[1]}
        
        conn.close()
        
        return jsonify({'success': True, 'fio': fio_data})
        
    except Exception as e:
        logger.error(f'Ошибка получения ФИО {fio_id}: {e}')
        return jsonify({'success': False, 'message': str(e)})

if __name__ == '__main__':
    import sys
    
    # Логируем запуск системы
    try:
        log_activity('Запуск системы', 'Веб-интерфейс M_Kran запущен', 'play')
        logger.info('Система M_Kran запущена')
    except Exception as e:
        logger.error(f'Ошибка при логировании запуска системы: {e}')
    
    # Отключаем автоматическое открытие браузера в режиме отладки
    # Отключаем предупреждение о development server
    import logging
    logging.getLogger('werkzeug').setLevel(logging.ERROR)
    
    app.run(
        debug=False,
        host='127.0.0.1',
        port=int(os.environ.get('MKRAN_PORT', '5000')),
        use_reloader=False
    )